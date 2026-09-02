import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIG
# ============================================================

INPUT_FILE = (
    r"output/Thriveni/"
    r"Thriveni_DHN_UNIVERSITY_DATA_01-09-2026/"
    r"Thriveni_DHN_UNIVERSITY_DATA_01-09-2026.xlsx"
)

OUTPUT_DIR = (
    r"cleaned_output/Thriveni/"
    r"Thriveni_DHN_UNIVERSITY_DATA_01-09-2026"
)

OUTPUT_FILENAME = (
    "Thriveni_DHN_UNIVERSITY_DATA_01-09-2026.xlsx"
)

SHEET_NAME = "Scraped Data"

EMAIL_COLUMN = "email"
NAME_COLUMN = "name"

# Email is mandatory
DROP_INVALID_EMAILS = True

# If name is missing/wrong, derive from email
NAME_FROM_EMAIL = True

# If firstname.lastname email clearly conflicts with bad scraped name,
# use the email-derived name.
FIX_MISMATCHED_NAMES_FROM_EMAIL = True

# Keep shared/generic emails such as info@, admin@ etc.
KEEP_SHARED_EMAILS = True

# Remove duplicate primary emails
REMOVE_DUPLICATE_EMAILS = True


# ============================================================
# EXPECTED COLUMNS
# ============================================================

EXPECTED_COLUMNS = [
    "name",
    "email",
    "country",
    "alternate_emails",
    "email_type",
    "email_conflict",
    "employee_name",
    "employee_email",
    "page_type",
    "journal_name",
    "editorial_role",
    "academic_title",
    "academic_rank",
    "specialty",
    "affiliation",
    "university",
    "faculty",
    "school",
    "department",
    "institute",
    "division",
    "city",
    "address",
    "phone",
    "orcid",
    "google_scholar",
    "scopus_author_id",
    "researcher_id",
    "pubmed",
    "profile_url",
    "personal_homepage",
    "source_url",
    "country_source",
    "confidence",
    "extraction_method",
    "scrape_status",
]


# ============================================================
# EMAIL RULES
# ============================================================

EMAIL_REGEX = re.compile(
    r"^[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,24}$",
    re.I,
)

FAKE_TLDS = {
    "turn",
    "human",
    "results",
    "his",
    "page",
    "html",
    "text",
    "example",
    "invalid",
    "test",
}

FAKE_EXACT_EMAILS = {
    "or@all.turn",
    "study@midgestation.human",
    "also@delivery.results",
    "analyst@musc.his",
    "person@email.cz",
}

SOCIAL_ARTIFACT_EMAILS = {
    "bluesky@bsky.social",
}

GENERIC_LOCAL_EXACT = {
    "info",
    "contact",
    "office",
    "admin",
    "administrator",
    "webmaster",
    "support",
    "help",
    "reception",
    "secretary",
    "faculty",
    "staff",
    "team",
    "department",
    "dept",
    "editor",
    "editorial",
    "enquiries",
    "inquiries",
    "communications",
    "media",
}

GENERIC_LOCAL_CONTAINS = {
    "webmaster",
    "contact",
    "support",
    "admin",
    "obsgyn",
    "gyn-",
    "faculty",
    "staff",
    "department",
    "editorial",
    "secretary",
    "reception",
}


# ============================================================
# NAME RULES
# ============================================================

NAME_LABEL_EXACT = {
    "",
    "telephone",
    "phone",
    "email",
    "contact",
    "contact us",
    "faculty",
    "staff",
    "team",
    "research team",
    "leadership",
    "administrative leadership",
    "view profile",
    "view full profile",
    "profile",
    "read more",
    "learn more",
    "home",
    "website",
    "location",
    "follow us on bluesky",
    "bluesky",
}

NAME_PREFIX_REGEX = re.compile(
    r"""
    ^\s*
    (?:
        professor |
        prof |
        dr |
        doctor |
        doc |
        mudr |
        mr |
        mrs |
        miss |
        ms
    )
    \.?
    \s+
    """,
    re.I | re.X,
)

CREDENTIAL_PATTERN = (
    r"(?:"
    r"MD|M\.D\.|"
    r"PhD|Ph\.D\.|"
    r"DO|D\.O\.|"
    r"PsyD|"
    r"EdD|"
    r"JD|"
    r"MPH|"
    r"MHA|"
    r"MPA|"
    r"MBA|"
    r"MSW|"
    r"MEd|"
    r"MSc|"
    r"MS|"
    r"MA|"
    r"BSc|"
    r"BA|"
    r"MBBS|"
    r"MBChB|"
    r"BMedSci|"
    r"ChB|"
    r"MRCP|"
    r"MRCPCH|"
    r"FRCPCH|"
    r"CSc|CSc\.|"
    r"ScD|"
    r"DDS|"
    r"DMD|"
    r"RN|"
    r"PharmD|"
    r"FACOG|"
    r"FACS|"
    r"FAPA"
    r")"
)


# ============================================================
# COUNTRY RULES
# ============================================================

SOURCE_COUNTRY_RULES = [
    ("uiowa.edu", "United States"),
    ("wustl.edu", "United States"),
    ("washington.edu", "United States"),
    ("uw.edu", "United States"),
    ("american.edu", "United States"),
    ("umaryland.edu", "United States"),
    ("nyulangone.org", "United States"),
    ("columbia.edu", "United States"),

    ("liverpool.ac.uk", "United Kingdom"),
    ("ox.ac.uk", "United Kingdom"),
    ("nottingham.ac.uk", "United Kingdom"),

    ("lunduniversity.lu.se", "Sweden"),
    ("med.lu.se", "Sweden"),

    ("rug.nl", "Netherlands"),
    ("umcg.nl", "Netherlands"),

    ("uio.no", "Norway"),

    ("rsu.lv", "Latvia"),

    ("lfmotol.cuni.cz", "Czech Republic"),
    ("lf2.cuni.cz", "Czech Republic"),
    ("lf3.cuni.cz", "Czech Republic"),

    ("um.edu.mt", "Malta"),

    ("euc.ac.cy", "Cyprus"),

    ("hunimed.eu", "Italy"),
]

EMAIL_TLD_COUNTRY_RULES = {
    ".ac.cy": "Cyprus",
    ".edu.mt": "Malta",
    ".ac.uk": "United Kingdom",
    ".cuni.cz": "Czech Republic",
    ".rsu.lv": "Latvia",
    ".lu.se": "Sweden",
    ".umcg.nl": "Netherlands",
}


# ============================================================
# BASIC HELPERS
# ============================================================

def clean_text(value):
    if value is None:
        return ""

    text = str(value)

    text = text.replace("\u00a0", " ")
    text = text.replace("\u200b", "")
    text = text.replace("\ufeff", "")

    text = re.sub(r"\s+", " ", text)

    return text.strip()


def ascii_normalize(value):
    value = clean_text(value)

    return (
        unicodedata.normalize("NFKD", value)
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )


def normalize_email(value):
    email = clean_text(value).lower()

    if not email:
        return ""

    email = re.sub(
        r"^mailto:",
        "",
        email,
        flags=re.I,
    )

    email = email.strip(
        " <>[](){};,:'\""
    )

    email = re.sub(
        r"\s*@\s*",
        "@",
        email,
    )

    email = re.sub(
        r"\s*\.\s*",
        ".",
        email,
    )

    return email.strip()


# ============================================================
# EMAIL VALIDATION
# ============================================================

def is_valid_email(email):
    email = normalize_email(email)

    if not email:
        return False

    if email in FAKE_EXACT_EMAILS:
        return False

    if email in SOCIAL_ARTIFACT_EMAILS:
        return False

    if not EMAIL_REGEX.fullmatch(email):
        return False

    if ".." in email:
        return False

    try:
        local, domain = email.rsplit("@", 1)
    except ValueError:
        return False

    if not local:
        return False

    if not domain:
        return False

    if "." not in domain:
        return False

    tld = domain.rsplit(".", 1)[-1]

    if tld in FAKE_TLDS:
        return False

    suspicious_domains = {
        "all.turn",
        "midgestation.human",
        "delivery.results",
    }

    if domain in suspicious_domains:
        return False

    return True


# ============================================================
# SHARED EMAIL CHECK
# ============================================================

def is_shared_email(email):
    email = normalize_email(email)

    if "@" not in email:
        return False

    local = email.split("@", 1)[0].lower()

    compact = re.sub(
        r"[^a-z]",
        "",
        local,
    )

    if compact in GENERIC_LOCAL_EXACT:
        return True

    for token in GENERIC_LOCAL_CONTAINS:
        if token in local:
            return True

    return False


# ============================================================
# NAME FROM EMAIL
# ============================================================

def email_local_parts(email):
    email = normalize_email(email)

    if "@" not in email:
        return []

    local = email.split("@", 1)[0]

    parts = re.split(
        r"[._+\-]+",
        local,
    )

    cleaned = []

    for part in parts:
        part = re.sub(
            r"\d+$",
            "",
            part,
        )

        part = re.sub(
            r"[^A-Za-zÀ-ÖØ-öø-ÿĀ-ž]",
            "",
            part,
        )

        if part:
            cleaned.append(part)

    return cleaned


def pretty_email_part(part):
    if not part:
        return ""

    if len(part) == 1:
        return part.upper() + "."

    return (
        part[0].upper()
        + part[1:].lower()
    )


def name_from_email(email):
    """
    john.smith@domain.edu
        -> John Smith

    j.smith@domain.edu
        -> J. Smith

    michael.brown3@domain.edu
        -> Michael Brown

    info@domain.edu
        -> Info
    """

    parts = email_local_parts(email)

    if not parts:
        local = normalize_email(
            email
        ).split("@")[0]

        return local.title()

    return " ".join(
        pretty_email_part(part)
        for part in parts
    )


def strong_full_name_from_email(email):
    parts = email_local_parts(email)

    meaningful = [
        part
        for part in parts
        if len(part) >= 3
    ]

    if len(meaningful) < 2:
        return ""

    return " ".join(
        pretty_email_part(part)
        for part in parts
    )


# ============================================================
# NAME CLEANING
# ============================================================

def remove_name_prefixes(name):
    name = clean_text(name)

    previous = None

    while name and previous != name:
        previous = name

        name = NAME_PREFIX_REGEX.sub(
            "",
            name,
        ).strip()

    return name


def remove_credentials(name):
    name = clean_text(name)

    for _ in range(10):
        old = name

        name = re.sub(
            rf"""
            (?:
                [,\s]+
                {CREDENTIAL_PATTERN}
                \.?
            )
            \s*$
            """,
            "",
            name,
            flags=re.I | re.X,
        )

        name = re.sub(
            r"\s*\((?:Hons|Honours)\)\s*$",
            "",
            name,
            flags=re.I,
        )

        name = name.strip(
            " ,;-"
        )

        if name == old:
            break

    return clean_text(name)


def clean_initials(name):
    return re.sub(
        r"(?<=[A-Za-z])\.(?=[A-Za-z])",
        ". ",
        name,
    )


def convert_surname_first(name):
    """
    Baker, Nicky
        -> Nicky Baker

    Care, Dr Angharad
        -> Angharad Care

    Bakker, dr. M.K.
        -> M. K. Bakker
    """

    if "," not in name:
        return name

    left, right = name.split(
        ",",
        1,
    )

    left = clean_text(left)
    right = clean_text(right)

    if not left or not right:
        return left or right

    right = re.sub(
        r"""
        ^
        (?:
            (?:professor|prof|drs?|doc|mr|mrs|miss|ms)
            \.?
            \s*
        )+
        """,
        "",
        right,
        flags=re.I | re.X,
    )

    right = remove_credentials(
        right
    )

    right = clean_initials(
        right
    )

    if not right:
        return left

    return f"{right} {left}"


def is_bad_name(name):
    name = clean_text(name)

    if not name:
        return True

    lower = name.lower()

    if lower in NAME_LABEL_EXACT:
        return True

    if "@" in name:
        return True

    if "[at]" in lower:
        return True

    if "(at)" in lower:
        return True

    if "[dot]" in lower:
        return True

    if "(dot)" in lower:
        return True

    if len(name) > 100:
        return True

    if re.fullmatch(
        r"[\W\d_]+",
        name,
    ):
        return True

    if lower.startswith(
        "view full profile"
    ):
        return True

    if lower.startswith(
        "follow us"
    ):
        return True

    return False


def clean_person_name(name):
    name = clean_text(name)

    if not name:
        return ""

    name = re.sub(
        r"^\s*profile\s+of\s+",
        "",
        name,
        flags=re.I,
    )

    name = remove_name_prefixes(
        name
    )

    name = convert_surname_first(
        name
    )

    name = remove_name_prefixes(
        name
    )

    name = remove_credentials(
        name
    )

    name = clean_initials(
        name
    )

    name = re.sub(
        r"\s+",
        " ",
        name,
    )

    return name.strip(
        " ,;-|"
    )


# ============================================================
# NAME / EMAIL MATCHING
# ============================================================

def tokenise_name(name):
    value = ascii_normalize(
        name
    )

    tokens = re.findall(
        r"[a-z]+",
        value,
    )

    ignore = {
        "dr",
        "doctor",
        "prof",
        "professor",
        "mr",
        "mrs",
        "miss",
        "ms",
        "md",
        "phd",
        "do",
        "jd",
    }

    return [
        token
        for token in tokens
        if len(token) >= 2
        and token not in ignore
    ]


def should_replace_name_from_email(
    name,
    email,
):
    strong_name = (
        strong_full_name_from_email(
            email
        )
    )

    if not strong_name:
        return False

    if is_bad_name(name):
        return True

    current_tokens = tokenise_name(
        name
    )

    email_tokens = tokenise_name(
        strong_name
    )

    if (
        len(email_tokens) < 2
        or not current_tokens
    ):
        return False

    email_first = email_tokens[0]
    email_last = email_tokens[-1]

    # Existing first/last matches email -> keep source name.
    if email_first in current_tokens:
        return False

    if email_last in current_tokens:
        return False

    # firstname.lastname email provides strong evidence.
    if (
        len(email_first) >= 3
        and len(email_last) >= 3
    ):
        return True

    return False


# ============================================================
# COUNTRY CLEANING
# ============================================================

def country_from_source(
    source_url
):
    source_url = clean_text(
        source_url
    ).lower()

    for token, country in SOURCE_COUNTRY_RULES:
        if token in source_url:
            return country

    return ""


def country_from_email(
    email
):
    email = normalize_email(
        email
    )

    if "@" not in email:
        return ""

    domain = email.split(
        "@",
        1,
    )[1]

    for suffix, country in EMAIL_TLD_COUNTRY_RULES.items():
        if domain.endswith(
            suffix
        ):
            return country

    return ""


# ============================================================
# ALTERNATE EMAIL CLEANUP
# ============================================================

def clean_alternate_emails(
    value,
    primary_email,
):
    value = clean_text(
        value
    )

    if not value:
        return ""

    candidates = re.split(
        r"[;,|\s]+",
        value,
    )

    output = []
    seen = set()

    primary_email = normalize_email(
        primary_email
    )

    for candidate in candidates:
        candidate = normalize_email(
            candidate
        )

        if not is_valid_email(
            candidate
        ):
            continue

        if candidate == primary_email:
            continue

        if candidate in seen:
            continue

        seen.add(
            candidate
        )

        output.append(
            candidate
        )

    return "; ".join(
        output
    )


# ============================================================
# URL CLEANING
# ============================================================

def clean_url(value):
    value = clean_text(
        value
    )

    if not value:
        return ""

    if value.lower() in {
        "none",
        "null",
        "n/a",
        "na",
        "-",
    }:
        return ""

    return value


# ============================================================
# PHONE CLEANUP
# ============================================================

def clean_phone(value):
    value = clean_text(
        value
    )

    if not value:
        return ""

    value = re.sub(
        r"\s+",
        " ",
        value,
    )

    return value[:150]


# ============================================================
# ROW QUALITY
# ============================================================

def row_quality(row):
    score = 0

    if clean_text(
        row.get("name")
    ):
        score += 30

    if clean_text(
        row.get("profile_url")
    ):
        score += 25

    if clean_text(
        row.get("affiliation")
    ):
        score += 10

    if clean_text(
        row.get("academic_title")
    ):
        score += 10

    if clean_text(
        row.get("country")
    ):
        score += 5

    if row.get(
        "email_type"
    ) == "personal":
        score += 5

    method = clean_text(
        row.get("extraction_method")
    )

    if "profile_v11" in method:
        score += 20

    if "mailto" in method:
        score += 5

    return score


# ============================================================
# CLEAN ONE ROW
# ============================================================

def clean_row(row):
    email = normalize_email(
        row.get("email")
    )

    # --------------------------------------------------------
    # EMAIL IS MANDATORY
    # --------------------------------------------------------

    if not is_valid_email(
        email
    ):
        return None, "invalid_email"

    row["email"] = email

    # --------------------------------------------------------
    # NAME
    # --------------------------------------------------------

    original_name = clean_text(
        row.get("name")
    )

    name = clean_person_name(
        original_name
    )

    name_changed = False
    name_generated_from_email = False

    if is_bad_name(name):
        if NAME_FROM_EMAIL:
            name = name_from_email(
                email
            )

            name_changed = True
            name_generated_from_email = True

    elif FIX_MISMATCHED_NAMES_FROM_EMAIL:

        if should_replace_name_from_email(
            name,
            email,
        ):
            derived_name = (
                strong_full_name_from_email(
                    email
                )
            )

            if derived_name:
                name = derived_name

                name_changed = True
                name_generated_from_email = True

    # Guarantee name when email is valid.
    if (
        not name
        and NAME_FROM_EMAIL
    ):
        name = name_from_email(
            email
        )

        name_changed = True
        name_generated_from_email = True

    row["name"] = name

    # --------------------------------------------------------
    # EMAIL TYPE
    # --------------------------------------------------------

    if is_shared_email(
        email
    ):
        row["email_type"] = (
            "shared/role"
        )
    else:
        row["email_type"] = (
            "personal"
        )

    row["email_conflict"] = "no"

    # --------------------------------------------------------
    # ALTERNATE EMAILS
    # --------------------------------------------------------

    row["alternate_emails"] = (
        clean_alternate_emails(
            row.get(
                "alternate_emails"
            ),
            email,
        )
    )

    # --------------------------------------------------------
    # COUNTRY
    # --------------------------------------------------------

    existing_country = clean_text(
        row.get("country")
    )

    source_country = (
        country_from_source(
            row.get(
                "source_url"
            )
        )
    )

    email_country = (
        country_from_email(
            email
        )
    )

    if source_country:
        row["country"] = (
            source_country
        )

        if (
            source_country
            != existing_country
        ):
            row["country_source"] = (
                "source_domain_cleaning"
            )

    elif (
        not existing_country
        and email_country
    ):
        row["country"] = (
            email_country
        )

        row["country_source"] = (
            "email_domain_cleaning"
        )

    else:
        row["country"] = (
            existing_country
        )

    # --------------------------------------------------------
    # NORMAL TEXT FIELDS
    # --------------------------------------------------------

    text_fields = [
        "journal_name",
        "editorial_role",
        "academic_title",
        "academic_rank",
        "specialty",
        "affiliation",
        "university",
        "faculty",
        "school",
        "department",
        "institute",
        "division",
        "city",
        "address",
    ]

    for field in text_fields:
        row[field] = clean_text(
            row.get(field)
        )

    row["phone"] = clean_phone(
        row.get("phone")
    )

    # --------------------------------------------------------
    # URL / IDENTIFIER FIELDS
    # --------------------------------------------------------

    url_fields = [
        "orcid",
        "google_scholar",
        "scopus_author_id",
        "researcher_id",
        "pubmed",
        "profile_url",
        "personal_homepage",
        "source_url",
    ]

    for field in url_fields:
        row[field] = clean_url(
            row.get(field)
        )

    # --------------------------------------------------------
    # PROVENANCE
    # --------------------------------------------------------

    if name_changed:

        method = clean_text(
            row.get(
                "extraction_method"
            )
        )

        if name_generated_from_email:
            marker = (
                "cleaned_name_from_email"
            )
        else:
            marker = (
                "cleaned_name"
            )

        if marker not in method:

            if method:
                method = (
                    f"{method}+{marker}"
                )
            else:
                method = marker

        row["extraction_method"] = (
            method
        )

        # Shared email derived names have lower identity confidence.
        if (
            row["email_type"]
            == "shared/role"
        ):
            try:
                confidence = float(
                    row.get(
                        "confidence"
                    )
                    or 100
                )

                row["confidence"] = min(
                    confidence,
                    80,
                )

            except Exception:
                row["confidence"] = 80

    row["scrape_status"] = (
        "accepted"
    )

    return row, None


# ============================================================
# READ EXCEL
# ============================================================

def read_excel(path):
    workbook = load_workbook(
        path
    )

    if SHEET_NAME in workbook.sheetnames:
        worksheet = workbook[
            SHEET_NAME
        ]
    else:
        worksheet = workbook[
            workbook.sheetnames[0]
        ]

    headers = []

    for cell in worksheet[1]:

        header = clean_text(
            cell.value
        )

        headers.append(
            header
        )

    header_index = {
        header: index
        for index, header
        in enumerate(headers)
        if header
    }

    if EMAIL_COLUMN not in header_index:

        raise RuntimeError(
            f"\nRequired column '{EMAIL_COLUMN}' not found.\n"
            f"Found columns:\n{headers}"
        )

    rows = []

    for excel_row_number, values in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):

        record = {}

        for index, header in enumerate(
            headers
        ):

            if not header:
                continue

            if index < len(values):
                record[header] = (
                    values[index]
                )
            else:
                record[header] = None

        record["_excel_row"] = (
            excel_row_number
        )

        rows.append(
            record
        )

    return (
        workbook,
        worksheet,
        headers,
        rows,
    )


# ============================================================
# MAIN CLEANER
# ============================================================

def clean_workbook(
    input_file,
    output_dir,
    output_filename,
):

    input_path = Path(
        input_file
    )

    if not input_path.exists():

        raise FileNotFoundError(
            "\nInput file not found:\n"
            f"{input_path.resolve()}\n"
        )

    output_directory = Path(
        output_dir
    )

    output_directory.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_path = (
        output_directory
        / output_filename
    )

    (
        workbook,
        worksheet,
        headers,
        raw_rows,
    ) = read_excel(
        input_path
    )

    print()
    print("=" * 80)
    print("UNIVERSITY DATA CLEANER")
    print("=" * 80)

    print(
        f"Input file : {input_path}"
    )

    print(
        f"Output dir : {output_directory}"
    )

    print(
        f"Input rows : {len(raw_rows)}"
    )

    print()

    cleaned_rows = []
    rejected_rows = []

    names_changed = 0
    names_from_email = 0
    shared_emails = 0

    # ========================================================
    # CLEAN
    # ========================================================

    for row in raw_rows:

        original_name = clean_text(
            row.get("name")
        )

        cleaned_row, error = clean_row(
            row.copy()
        )

        if cleaned_row is None:

            rejected_rows.append(
                {
                    "excel_row":
                        row.get(
                            "_excel_row"
                        ),

                    "name":
                        original_name,

                    "email":
                        normalize_email(
                            row.get(
                                "email"
                            )
                        ),

                    "reason":
                        error,
                }
            )

            continue

        final_name = clean_text(
            cleaned_row.get(
                "name"
            )
        )

        if (
            final_name
            != original_name
        ):

            names_changed += 1

            method = clean_text(
                cleaned_row.get(
                    "extraction_method"
                )
            )

            if (
                "cleaned_name_from_email"
                in method
            ):
                names_from_email += 1

        if (
            cleaned_row.get(
                "email_type"
            )
            == "shared/role"
        ):
            shared_emails += 1

        cleaned_rows.append(
            cleaned_row
        )

    # ========================================================
    # DEDUPE BY PRIMARY EMAIL
    # ========================================================

    duplicate_count = 0

    if REMOVE_DUPLICATE_EMAILS:

        best_by_email = {}

        for row in cleaned_rows:

            email = normalize_email(
                row.get(
                    "email"
                )
            )

            if email not in best_by_email:

                best_by_email[
                    email
                ] = row

                continue

            duplicate_count += 1

            current = best_by_email[
                email
            ]

            if (
                row_quality(row)
                > row_quality(current)
            ):

                best_by_email[
                    email
                ] = row

        cleaned_rows = list(
            best_by_email.values()
        )

    # ========================================================
    # SORT
    # ========================================================

    cleaned_rows.sort(
        key=lambda row: (
            clean_text(
                row.get(
                    "name"
                )
            ).casefold(),

            clean_text(
                row.get(
                    "email"
                )
            ).casefold(),
        )
    )

    # ========================================================
    # REMOVE OLD DATA
    # ========================================================

    if worksheet.max_row > 1:

        worksheet.delete_rows(
            2,
            worksheet.max_row - 1,
        )

    # ========================================================
    # WRITE CLEANED DATA
    # ========================================================

    for row in cleaned_rows:

        values = []

        for header in headers:

            if not header:

                values.append(
                    ""
                )

                continue

            value = row.get(
                header
            )

            if value is None:
                value = ""

            values.append(
                value
            )

        worksheet.append(
            values
        )

    # ========================================================
    # EXCEL FORMATTING
    # ========================================================

    worksheet.freeze_panes = (
        "A2"
    )

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    widths = {
        "name": 30,
        "email": 40,
        "country": 20,
        "alternate_emails": 45,
        "email_type": 16,
        "email_conflict": 16,
        "employee_name": 20,
        "employee_email": 35,
        "page_type": 22,
        "journal_name": 35,
        "editorial_role": 30,
        "academic_title": 40,
        "academic_rank": 25,
        "specialty": 40,
        "affiliation": 60,
        "university": 45,
        "faculty": 40,
        "school": 40,
        "department": 45,
        "institute": 45,
        "division": 40,
        "city": 22,
        "address": 60,
        "phone": 25,
        "orcid": 35,
        "google_scholar": 50,
        "scopus_author_id": 30,
        "researcher_id": 30,
        "pubmed": 50,
        "profile_url": 65,
        "personal_homepage": 65,
        "source_url": 70,
        "country_source": 25,
        "confidence": 15,
        "extraction_method": 70,
        "scrape_status": 18,
    }

    for column_number, header in enumerate(
        headers,
        start=1,
    ):

        column_letter = (
            get_column_letter(
                column_number
            )
        )

        width = widths.get(
            header,
            22,
        )

        worksheet.column_dimensions[
            column_letter
        ].width = width

    # ========================================================
    # SAVE
    # ========================================================

    workbook.save(
        output_path
    )

    # ========================================================
    # SUMMARY
    # ========================================================

    unique_emails = len(
        {
            normalize_email(
                row.get(
                    "email"
                )
            )
            for row in cleaned_rows
        }
    )

    blank_names = sum(
        1
        for row in cleaned_rows
        if not clean_text(
            row.get(
                "name"
            )
        )
    )

    print()
    print("=" * 80)
    print("CLEANING SUMMARY")
    print("=" * 80)

    print(
        f"Original rows              : {len(raw_rows)}"
    )

    print(
        f"Invalid/junk emails removed: {len(rejected_rows)}"
    )

    print(
        f"Duplicate emails removed   : {duplicate_count}"
    )

    print(
        f"Final rows                 : {len(cleaned_rows)}"
    )

    print(
        f"Unique emails              : {unique_emails}"
    )

    print(
        f"Names changed              : {names_changed}"
    )

    print(
        f"Names generated from email : {names_from_email}"
    )

    print(
        f"Shared/role emails         : {shared_emails}"
    )

    print(
        f"Blank names remaining      : {blank_names}"
    )

    # ========================================================
    # REJECTED EMAILS
    # ========================================================

    if rejected_rows:

        print()
        print("-" * 80)
        print("REMOVED INVALID/JUNK EMAILS")
        print("-" * 80)

        for item in rejected_rows:

            print(
                f"Row {item['excel_row']} | "
                f"{item['name']} | "
                f"{item['email']} | "
                f"{item['reason']}"
            )

    print()
    print("=" * 80)

    print(
        "CLEANED FILE SAVED TO:"
    )

    print(
        output_path.resolve()
    )

    print("=" * 80)
    print()

    return output_path


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":

    try:

        clean_workbook(
            input_file=INPUT_FILE,
            output_dir=OUTPUT_DIR,
            output_filename=OUTPUT_FILENAME,
        )

    except Exception as exc:

        print()
        print("=" * 80)
        print("ERROR")
        print("=" * 80)

        print(
            str(exc)
        )

        print()

        raise