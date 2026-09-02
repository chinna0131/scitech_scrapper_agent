from __future__ import annotations

import argparse
import logging
import random
import re
import time
from pathlib import Path
from urllib.parse import quote_plus, urljoin

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import sync_playwright


# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE = Path("queries.xlsx")

OUTPUT_DIR = Path("google_output")
OUTPUT_FILE = OUTPUT_DIR / "parneetha_google_maps_results.xlsx"
CHECKPOINT_FILE = OUTPUT_DIR / "parneetha_google_maps_results_checkpoint.csv"

PROFILE_DIR = (
    Path("browser_profiles")
    / "google_maps_multi_query"
)

MAX_SCROLLS = 300
NO_NEW_LIMIT = 12

MIN_PROFILE_DELAY = 2.0
MAX_PROFILE_DELAY = 4.5

MIN_WEBSITE_DELAY = 2.5
MAX_WEBSITE_DELAY = 5.5

MIN_QUERY_DELAY = 8.0
MAX_QUERY_DELAY = 15.0

LONG_BREAK_EVERY = 20
LONG_BREAK_MIN = 20.0
LONG_BREAK_MAX = 40.0


COLUMNS = [
    "search_query",
    "company_name",
    "category",
    "rating",
    "address",
    "telephone",
    "website",
    "email",
    "google_profile_link",
    "latitude",
    "longitude",
    "search_url",
]


# ============================================================
# BASIC HELPERS
# ============================================================

def clean(value: str | None) -> str:
    return re.sub(
        r"\s+",
        " ",
        str(value or ""),
    ).strip()


def human_delay(
    minimum: float,
    maximum: float,
    label: str,
) -> None:
    seconds = random.uniform(
        minimum,
        maximum,
    )

    logging.info(
        "%s delay: %.1f seconds",
        label,
        seconds,
    )

    time.sleep(seconds)


def build_search_url(
    query: str,
) -> str:
    return (
        "https://www.google.com/maps/search/"
        + quote_plus(query)
        + "?hl=en"
    )


# ============================================================
# INPUT EXCEL
# ============================================================

def load_queries() -> list[str]:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Input Excel not found: "
            f"{INPUT_FILE.resolve()}"
        )

    dataframe = pd.read_excel(
        INPUT_FILE,
        dtype=str,
    ).fillna("")

    dataframe.columns = [
        clean(column).casefold()
        for column in dataframe.columns
    ]

    if "query" not in dataframe.columns:
        raise ValueError(
            "Input Excel must contain a column named 'query'."
        )

    queries: list[str] = []
    seen: set[str] = set()

    for value in dataframe["query"]:
        query = clean(value)

        if not query:
            continue

        key = query.casefold()

        if key in seen:
            continue

        seen.add(key)
        queries.append(query)

    if not queries:
        raise ValueError(
            "No valid queries found in queries.xlsx."
        )

    return queries


# ============================================================
# GOOGLE MAPS VERIFICATION
# ============================================================

def is_verification_page(page) -> bool:
    url = page.url.lower()

    if (
        "consent.google." in url
        or "/sorry/" in url
        or "recaptcha" in url
    ):
        return True

    selectors = [
        'iframe[src*="recaptcha"]',
        'form[action*="/sorry/"]',
        "#captcha-form",
    ]

    for selector in selectors:
        try:
            locator = page.locator(
                selector
            ).first

            if (
                locator.count()
                and locator.is_visible()
            ):
                return True

        except Exception:
            continue

    return False


def wait_for_maps_results(
    page,
    timeout_seconds: int = 900,
) -> None:
    deadline = (
        time.time()
        + timeout_seconds
    )

    while time.time() < deadline:
        if is_verification_page(page):
            logging.info(
                "Google verification detected. "
                "Complete it manually in Chrome."
            )

            page.wait_for_timeout(
                3000
            )

            continue

        result_links = page.locator(
            'a[href*="/maps/place/"]'
        ).count()

        feed_count = page.locator(
            'div[role="feed"]'
        ).count()

        if (
            result_links > 0
            or feed_count > 0
        ):
            logging.info(
                "Google Maps results loaded."
            )

            return

        page.wait_for_timeout(
            2500
        )

    raise RuntimeError(
        "Google Maps results did not load "
        "before timeout."
    )


# ============================================================
# GOOGLE MAPS RESULT COLLECTION
# ============================================================

def get_result_feed(page):
    feed = page.locator(
        'div[role="feed"]'
    ).first

    try:
        if (
            feed.count()
            and feed.is_visible()
        ):
            return feed

    except Exception:
        pass

    candidates = [
        'div[aria-label*="Results for" i]',
        'div[aria-label*="Search results" i]',
        'div[role="main"]',
    ]

    for selector in candidates:
        locator = page.locator(
            selector
        ).first

        try:
            if (
                locator.count()
                and locator.is_visible()
            ):
                return locator

        except Exception:
            continue

    return None


def collect_profile_links(
    page,
    max_scrolls: int,
) -> list[str]:
    feed = get_result_feed(page)

    links: dict[str, None] = {}

    previous_count = 0
    no_new_rounds = 0

    for scroll_index in range(
        1,
        max_scrolls + 1,
    ):
        anchors = page.locator(
            'a[href*="/maps/place/"]'
        ).all()

        for anchor in anchors:
            try:
                href = clean(
                    anchor.get_attribute(
                        "href"
                    )
                )

            except Exception:
                continue

            if not href:
                continue

            if href.startswith("/"):
                href = urljoin(
                    "https://www.google.com",
                    href,
                )

            normalized = (
                href
                .split("&", 1)[0]
                .rstrip("/")
            )

            if normalized:
                links[normalized] = None

        current_count = len(
            links
        )

        logging.info(
            "Scroll %d | unique companies=%d",
            scroll_index,
            current_count,
        )

        if current_count <= previous_count:
            no_new_rounds += 1
        else:
            no_new_rounds = 0

        previous_count = (
            current_count
        )

        body_text = ""

        try:
            body_text = clean(
                page.locator(
                    "body"
                ).inner_text()
            ).casefold()

        except Exception:
            pass

        reached_end = any(
            marker in body_text
            for marker in [
                "you've reached the end of the list",
                "you have reached the end of the list",
                "end of results",
                "no more results",
            ]
        )

        if reached_end:
            logging.info(
                "Google Maps reported "
                "the end of the list."
            )

            break

        if no_new_rounds >= NO_NEW_LIMIT:
            logging.info(
                "No additional companies loaded "
                "after %d rounds.",
                NO_NEW_LIMIT,
            )

            break

        try:
            if feed:
                feed.evaluate(
                    """
                    (element) => {
                        element.scrollTo(
                            0,
                            element.scrollHeight
                        );
                    }
                    """
                )

            else:
                page.mouse.wheel(
                    0,
                    3000,
                )

        except Exception:
            page.mouse.wheel(
                0,
                3000,
            )

        page.wait_for_timeout(
            random.randint(
                1800,
                3500,
            )
        )

    return list(
        links.keys()
    )


# ============================================================
# PAGE EXTRACTION HELPERS
# ============================================================

def first_visible_text(
    page,
    selectors: list[str],
) -> str:
    for selector in selectors:
        try:
            locator = page.locator(
                selector
            ).first

            if (
                locator.count()
                and locator.is_visible()
            ):
                value = clean(
                    locator.inner_text()
                )

                if value:
                    return value

        except Exception:
            continue

    return ""


def first_visible_attribute(
    page,
    selectors: list[str],
    attribute: str,
) -> str:
    for selector in selectors:
        try:
            locator = page.locator(
                selector
            ).first

            if (
                locator.count()
                and locator.is_visible()
            ):
                value = clean(
                    locator.get_attribute(
                        attribute
                    )
                )

                if value:
                    return value

        except Exception:
            continue

    return ""


def extract_coordinates(
    url: str,
) -> tuple[str, str]:
    patterns = [
        r"@(-?\d+\.\d+),(-?\d+\.\d+)",
        r"!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)",
    ]

    for pattern in patterns:
        match = re.search(
            pattern,
            url,
        )

        if match:
            return (
                match.group(1),
                match.group(2),
            )

    return "", ""


def normalize_rating(
    value: str,
) -> str:
    match = re.search(
        r"[1-5](?:[.,]\d+)?",
        value,
    )

    if not match:
        return ""

    return match.group(
        0
    ).replace(
        ",",
        ".",
    )


def normalize_phone(
    value: str,
) -> str:
    value = clean(
        value
    )

    value = re.sub(
        r"^(Phone|Call)\s*:?\s*",
        "",
        value,
        flags=re.I,
    )

    return value


def clean_google_redirect(
    href: str,
) -> str:
    href = clean(
        href
    )

    if not href:
        return ""

    if href.startswith(
        "/url?"
    ):
        match = re.search(
            r"[?&]q=([^&]+)",
            href,
        )

        if match:
            from urllib.parse import unquote

            return unquote(
                match.group(1)
            )

    return href


# ============================================================
# EMAIL EXTRACTION
# ============================================================

EMAIL_PATTERN = re.compile(
    r"[A-Z0-9._%+\-]+"
    r"@[A-Z0-9.\-]+"
    r"\.[A-Z]{2,}",
    re.I,
)


BLOCKED_EMAILS = {
    "example@example.com",
    "name@example.com",
    "email@example.com",
    "your@email.com",
    "you@example.com",
    "test@example.com",
}


def normalize_email(
    email: str,
) -> str:
    email = clean(
        email
    ).lower()

    email = email.strip(
        ".,;:()[]{}<>"
    )

    return email


def find_email_in_soup(
    soup: BeautifulSoup,
) -> str:
    # --------------------------------------------------------
    # MAILTO LINKS
    # --------------------------------------------------------

    for anchor in soup.select(
        'a[href^="mailto:"]'
    ):
        href = clean(
            anchor.get(
                "href",
                "",
            )
        )

        href = re.sub(
            r"^mailto:",
            "",
            href,
            flags=re.I,
        )

        href = href.split(
            "?",
            1,
        )[0]

        match = EMAIL_PATTERN.search(
            href
        )

        if not match:
            continue

        email = normalize_email(
            match.group(0)
        )

        if (
            email
            and email not in BLOCKED_EMAILS
        ):
            return email

    # --------------------------------------------------------
    # PAGE TEXT
    # --------------------------------------------------------

    page_text = soup.get_text(
        " ",
        strip=True,
    )

    emails = EMAIL_PATTERN.findall(
        page_text
    )

    for value in emails:
        email = normalize_email(
            value
        )

        if (
            email
            and email not in BLOCKED_EMAILS
        ):
            return email

    return ""


def extract_email_from_website(
    context,
    website: str,
) -> str:
    if not website:
        return ""

    page = context.new_page()

    try:
        logging.info(
            "Checking website for email: %s",
            website,
        )

        page.goto(
            website,
            wait_until="domcontentloaded",
            timeout=60000,
        )

        page.wait_for_timeout(
            random.randint(
                1400,
                2600,
            )
        )

        soup = BeautifulSoup(
            page.content(),
            "html.parser",
        )

        email = find_email_in_soup(
            soup
        )

        if email:
            return email

        # ----------------------------------------------------
        # FIND CONTACT / ABOUT PAGES
        # ----------------------------------------------------

        contact_urls: list[str] = []

        keywords = [
            "contact",
            "contact us",
            "about",
            "about us",
            "clinic",
            "hospital",
            "team",
            "location",
        ]

        for anchor in soup.select(
            "a[href]"
        ):
            label = clean(
                anchor.get_text(
                    " ",
                    strip=True,
                )
            ).casefold()

            href = clean(
                anchor.get(
                    "href",
                    "",
                )
            )

            if not href:
                continue

            if any(
                keyword in label
                for keyword in keywords
            ):
                absolute_url = urljoin(
                    page.url,
                    href,
                )

                if (
                    absolute_url
                    not in contact_urls
                ):
                    contact_urls.append(
                        absolute_url
                    )

        # ----------------------------------------------------
        # CHECK MAXIMUM 4 CONTACT-LIKE PAGES
        # ----------------------------------------------------

        for contact_url in contact_urls[:4]:
            try:
                logging.info(
                    "Checking contact page: %s",
                    contact_url,
                )

                page.goto(
                    contact_url,
                    wait_until="domcontentloaded",
                    timeout=45000,
                )

                page.wait_for_timeout(
                    random.randint(
                        1000,
                        1800,
                    )
                )

                contact_soup = BeautifulSoup(
                    page.content(),
                    "html.parser",
                )

                email = find_email_in_soup(
                    contact_soup
                )

                if email:
                    return email

            except Exception as exc:
                logging.debug(
                    "Contact page failed: %s | %s",
                    contact_url,
                    exc,
                )

                continue

        return ""

    except Exception as exc:
        logging.warning(
            "Website failed: %s | %s",
            website,
            exc,
        )

        return ""

    finally:
        try:
            page.close()
        except Exception:
            pass


# ============================================================
# GOOGLE MAPS PROFILE SCRAPER
# ============================================================

def scrape_company(
    context,
    profile_link: str,
    collect_email: bool,
    search_query: str,
    search_url: str,
) -> dict[str, str]:
    page = context.new_page()

    try:
        page.goto(
            profile_link,
            wait_until="domcontentloaded",
            timeout=90000,
        )

        page.wait_for_timeout(
            random.randint(
                1800,
                3200,
            )
        )

        deadline = (
            time.time()
            + 35
        )

        while time.time() < deadline:
            company_name = first_visible_text(
                page,
                [
                    "h1.DUwDvf",
                    "h1",
                    '[role="heading"][aria-level="1"]',
                ],
            )

            if company_name:
                break

            page.wait_for_timeout(
                800
            )

        company_name = first_visible_text(
            page,
            [
                "h1.DUwDvf",
                "h1",
                '[role="heading"][aria-level="1"]',
            ],
        )

        category = first_visible_text(
            page,
            [
                'button[jsaction*="category"]',
                "button.DkEaL",
                "div.DkEaL",
                ".skqShb",
            ],
        )

        rating_text = first_visible_text(
            page,
            [
                "div.F7nice span[aria-hidden='true']",
                'span[aria-label*="stars" i]',
                'span[aria-label*="Rated" i]',
            ],
        )

        rating = normalize_rating(
            rating_text
        )

        address = first_visible_text(
            page,
            [
                'button[data-item-id="address"]',
                'button[aria-label^="Address:" i]',
                'div[data-item-id="address"]',
            ],
        )

        telephone = first_visible_text(
            page,
            [
                'button[data-item-id^="phone:tel:"]',
                'button[aria-label^="Phone:" i]',
                'a[href^="tel:"]',
            ],
        )

        if not telephone:
            telephone = (
                first_visible_attribute(
                    page,
                    [
                        'a[href^="tel:"]',
                    ],
                    "href",
                )
                .replace(
                    "tel:",
                    "",
                )
            )

        telephone = normalize_phone(
            telephone
        )

        website = first_visible_attribute(
            page,
            [
                'a[data-item-id="authority"]',
                'a[aria-label*="Website" i]',
            ],
            "href",
        )

        website = clean_google_redirect(
            website
        )

        final_profile_link = (
            page.url
        )

        (
            latitude,
            longitude,
        ) = extract_coordinates(
            final_profile_link
        )

        email = ""

        if (
            collect_email
            and website
        ):
            email = extract_email_from_website(
                context,
                website,
            )

            human_delay(
                MIN_WEBSITE_DELAY,
                MAX_WEBSITE_DELAY,
                "Website",
            )

        return {
            "search_query": search_query,
            "company_name": company_name,
            "category": category,
            "rating": rating,
            "address": address,
            "telephone": telephone,
            "website": website,
            "email": email,
            "google_profile_link": final_profile_link,
            "latitude": latitude,
            "longitude": longitude,
            "search_url": search_url,
        }

    finally:
        try:
            page.close()
        except Exception:
            pass


# ============================================================
# CHECKPOINT
# ============================================================

def load_checkpoint() -> list[dict[str, str]]:
    if not CHECKPOINT_FILE.exists():
        return []

    dataframe = pd.read_csv(
        CHECKPOINT_FILE,
        dtype=str,
    ).fillna("")

    for column in COLUMNS:
        if column not in dataframe.columns:
            dataframe[
                column
            ] = ""

    return dataframe[
        COLUMNS
    ].to_dict(
        orient="records"
    )


# ============================================================
# DEDUPLICATION
# ============================================================

def normalize_profile_key(
    value: str,
) -> str:
    return clean(
        value
    ).lower().rstrip("/")


def normalize_company_key(
    value: str,
) -> str:
    return clean(
        value
    ).casefold()


def find_existing_record(
    records: list[dict[str, str]],
    profile_link: str,
    company_name: str = "",
) -> dict[str, str] | None:
    profile_key = normalize_profile_key(
        profile_link
    )

    company_key = normalize_company_key(
        company_name
    )

    for record in records:
        record_profile_key = normalize_profile_key(
            record.get(
                "google_profile_link",
                "",
            )
        )

        if (
            profile_key
            and record_profile_key
            and profile_key == record_profile_key
        ):
            return record

        if company_key:
            record_company_key = normalize_company_key(
                record.get(
                    "company_name",
                    "",
                )
            )

            if (
                record_company_key
                and company_key == record_company_key
            ):
                return record

    return None


def add_query_to_record(
    record: dict[str, str],
    query: str,
    search_url: str,
) -> None:
    existing_queries = [
        clean(item)
        for item in clean(
            record.get(
                "search_query",
                "",
            )
        ).split("|")
        if clean(item)
    ]

    existing_query_keys = {
        item.casefold()
        for item in existing_queries
    }

    if (
        query.casefold()
        not in existing_query_keys
    ):
        existing_queries.append(
            query
        )

    record[
        "search_query"
    ] = " | ".join(
        existing_queries
    )

    existing_urls = [
        clean(item)
        for item in clean(
            record.get(
                "search_url",
                "",
            )
        ).split("|")
        if clean(item)
    ]

    if (
        search_url
        and search_url not in existing_urls
    ):
        existing_urls.append(
            search_url
        )

    record[
        "search_url"
    ] = " | ".join(
        existing_urls
    )


# ============================================================
# SAVE OUTPUT
# ============================================================

def save_output(
    records: list[dict[str, str]],
) -> None:
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    dataframe = pd.DataFrame(
        records,
        columns=COLUMNS,
    ).fillna("")

    dataframe.to_csv(
        CHECKPOINT_FILE,
        index=False,
        encoding="utf-8-sig",
    )

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:
        dataframe.to_excel(
            writer,
            sheet_name="results",
            index=False,
        )

        worksheet = writer.sheets[
            "results"
        ]

        worksheet.freeze_panes = (
            "A2"
        )

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        widths = {
            "A": 70,
            "B": 44,
            "C": 32,
            "D": 12,
            "E": 65,
            "F": 24,
            "G": 55,
            "H": 42,
            "I": 80,
            "J": 16,
            "K": 16,
            "L": 90,
        }

        for (
            column,
            width,
        ) in widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        for row in worksheet.iter_rows(
            min_row=2
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    logging.info(
        "Excel saved: %s",
        OUTPUT_FILE.resolve(),
    )


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Scrape Google Maps results "
            "for multiple queries from Excel."
        )
    )

    parser.add_argument(
        "--input",
        type=str,
        default=None,
        help=(
            "Optional input Excel path. "
            "Default: queries.xlsx"
        ),
    )

    parser.add_argument(
        "--max-companies",
        type=int,
        default=None,
        help=(
            "Optional maximum companies "
            "per query for testing."
        ),
    )

    parser.add_argument(
        "--max-scrolls",
        type=int,
        default=MAX_SCROLLS,
    )

    parser.add_argument(
        "--skip-emails",
        action="store_true",
        help=(
            "Do not visit business websites "
            "to search for email addresses."
        ),
    )

    parser.add_argument(
        "--fresh",
        action="store_true",
        help=(
            "Delete existing checkpoint "
            "and scrape from the beginning."
        ),
    )

    args = parser.parse_args()

    # --------------------------------------------------------
    # OPTIONAL INPUT FILE OVERRIDE
    # --------------------------------------------------------

    global INPUT_FILE

    if args.input:
        INPUT_FILE = Path(
            args.input
        )

    # --------------------------------------------------------
    # LOGGING
    # --------------------------------------------------------

    logging.basicConfig(
        level=logging.INFO,
        format=(
            "%(asctime)s | "
            "%(levelname)s | "
            "%(message)s"
        ),
    )

    # --------------------------------------------------------
    # FRESH RUN
    # --------------------------------------------------------

    if args.fresh:
        if CHECKPOINT_FILE.exists():
            CHECKPOINT_FILE.unlink()

        if OUTPUT_FILE.exists():
            OUTPUT_FILE.unlink()

    # --------------------------------------------------------
    # LOAD EXCEL QUERIES
    # --------------------------------------------------------

    queries = load_queries()

    logging.info(
        "Queries loaded: %d",
        len(queries),
    )

    for index, query in enumerate(
        queries,
        start=1,
    ):
        logging.info(
            "Query %d: %s",
            index,
            query,
        )

    # --------------------------------------------------------
    # LOAD EXISTING CHECKPOINT
    # --------------------------------------------------------

    existing = load_checkpoint()

    logging.info(
        "Existing checkpoint records: %d",
        len(existing),
    )

    # --------------------------------------------------------
    # BROWSER PROFILE
    # --------------------------------------------------------

    PROFILE_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    with sync_playwright() as playwright:

        context = (
            playwright.chromium
            .launch_persistent_context(
                user_data_dir=str(
                    PROFILE_DIR.resolve()
                ),
                headless=False,
                channel="chrome",
                viewport=None,
                args=[
                    "--start-maximized",
                    (
                        "--disable-blink-features="
                        "AutomationControlled"
                    ),
                ],
            )
        )

        search_page = (
            context.pages[0]
            if context.pages
            else context.new_page()
        )

        # ====================================================
        # PROCESS EACH QUERY
        # ====================================================

        for (
            query_index,
            query,
        ) in enumerate(
            queries,
            start=1,
        ):

            logging.info(
                "=" * 80
            )

            logging.info(
                "QUERY [%d/%d]: %s",
                query_index,
                len(queries),
                query,
            )

            logging.info(
                "=" * 80
            )

            search_url = build_search_url(
                query
            )

            try:
                search_page.goto(
                    search_url,
                    wait_until="domcontentloaded",
                    timeout=90000,
                )

            except Exception as exc:
                logging.warning(
                    "Query page failed: %s | %s",
                    query,
                    exc,
                )

                continue

            print(
                "\n"
                f"Query: {query}\n"
                "Google Maps is open.\n"
                "Complete Google verification manually "
                "only if it appears.\n"
            )

            try:
                wait_for_maps_results(
                    search_page
                )

            except Exception as exc:
                logging.warning(
                    "Results failed for query: "
                    "%s | %s",
                    query,
                    exc,
                )

                continue

            # ------------------------------------------------
            # COLLECT ALL PROFILE LINKS
            # ------------------------------------------------

            profile_links = collect_profile_links(
                search_page,
                max_scrolls=args.max_scrolls,
            )

            if (
                args.max_companies
                is not None
            ):
                profile_links = (
                    profile_links[
                        :args.max_companies
                    ]
                )

            logging.info(
                "Query '%s' | "
                "company profiles collected: %d",
                query,
                len(profile_links),
            )

            # ------------------------------------------------
            # SCRAPE EACH COMPANY
            # ------------------------------------------------

            for (
                company_index,
                profile_link,
            ) in enumerate(
                profile_links,
                start=1,
            ):

                logging.info(
                    "[Query %d/%d] "
                    "[Company %d/%d] "
                    "Checking profile",
                    query_index,
                    len(queries),
                    company_index,
                    len(profile_links),
                )

                # --------------------------------------------
                # FIRST CHECK PROFILE LINK AGAINST EXISTING
                # --------------------------------------------

                existing_record = (
                    find_existing_record(
                        existing,
                        profile_link,
                    )
                )

                if existing_record:
                    add_query_to_record(
                        existing_record,
                        query,
                        search_url,
                    )

                    save_output(
                        existing
                    )

                    logging.info(
                        "Already known company. "
                        "Added query mapping: %s",
                        query,
                    )

                    continue

                # --------------------------------------------
                # SCRAPE COMPANY
                # --------------------------------------------

                try:
                    record = scrape_company(
                        context=context,
                        profile_link=profile_link,
                        collect_email=(
                            not args.skip_emails
                        ),
                        search_query=query,
                        search_url=search_url,
                    )

                except Exception as exc:
                    logging.warning(
                        "Company failed: %s | %s",
                        profile_link,
                        exc,
                    )

                    continue

                # --------------------------------------------
                # VALIDATE COMPANY NAME
                # --------------------------------------------

                if not record[
                    "company_name"
                ]:
                    logging.warning(
                        "No company name found: %s",
                        profile_link,
                    )

                # --------------------------------------------
                # SECOND DUPLICATE CHECK
                # --------------------------------------------

                duplicate = (
                    find_existing_record(
                        existing,
                        record.get(
                            "google_profile_link",
                            "",
                        ),
                        record.get(
                            "company_name",
                            "",
                        ),
                    )
                )

                if duplicate:
                    add_query_to_record(
                        duplicate,
                        query,
                        search_url,
                    )

                    # Fill missing data from new record
                    for column in [
                        "category",
                        "rating",
                        "address",
                        "telephone",
                        "website",
                        "email",
                        "latitude",
                        "longitude",
                    ]:
                        if (
                            not clean(
                                duplicate.get(
                                    column,
                                    "",
                                )
                            )
                            and clean(
                                record.get(
                                    column,
                                    "",
                                )
                            )
                        ):
                            duplicate[
                                column
                            ] = record[
                                column
                            ]

                    save_output(
                        existing
                    )

                    logging.info(
                        "Duplicate merged | %s | %s",
                        record.get(
                            "company_name",
                            "",
                        ),
                        query,
                    )

                    continue

                # --------------------------------------------
                # NEW COMPANY
                # --------------------------------------------

                existing.append(
                    record
                )

                save_output(
                    existing
                )

                logging.info(
                    "SAVED | Query=%s | "
                    "Company=%s | "
                    "Phone=%s | "
                    "Email=%s",
                    query,
                    record.get(
                        "company_name",
                        "",
                    ),
                    record.get(
                        "telephone",
                        "",
                    )
                    or "No phone",
                    record.get(
                        "email",
                        "",
                    )
                    or "No email",
                )

                human_delay(
                    MIN_PROFILE_DELAY,
                    MAX_PROFILE_DELAY,
                    "Profile",
                )

                # --------------------------------------------
                # LONG REST
                # --------------------------------------------

                if (
                    company_index
                    % LONG_BREAK_EVERY
                    == 0
                ):
                    human_delay(
                        LONG_BREAK_MIN,
                        LONG_BREAK_MAX,
                        "Long rest",
                    )

            # ------------------------------------------------
            # QUERY COMPLETED
            # ------------------------------------------------

            save_output(
                existing
            )

            logging.info(
                "COMPLETED QUERY [%d/%d]: %s",
                query_index,
                len(queries),
                query,
            )

            logging.info(
                "Total unique companies saved: %d",
                len(existing),
            )

            # ------------------------------------------------
            # DELAY BETWEEN QUERIES
            # ------------------------------------------------

            if (
                query_index
                < len(queries)
            ):
                human_delay(
                    MIN_QUERY_DELAY,
                    MAX_QUERY_DELAY,
                    "Between queries",
                )

        # ====================================================
        # FINAL SAVE
        # ====================================================

        save_output(
            existing
        )

        context.close()

    logging.info(
        "=" * 80
    )

    logging.info(
        "ALL QUERIES COMPLETED"
    )

    logging.info(
        "Final unique companies exported: %d",
        len(existing),
    )

    logging.info(
        "Output: %s",
        OUTPUT_FILE.resolve(),
    )

    logging.info(
        "=" * 80
    )


if __name__ == "__main__":
    main()