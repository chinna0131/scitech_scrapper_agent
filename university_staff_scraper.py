import asyncio
import json
import random
import re
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from urllib.parse import urljoin, urlparse, urldefrag, unquote

import requests
from bs4 import BeautifulSoup, Tag, NavigableString
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

try:
    import pycountry
except Exception:
    pycountry = None


# =============================================================================
# CONFIG
# =============================================================================

EMPLOYEE_NAME = "Thriveni"
EMPLOYEE_EMAIL = "dyavanapallythriveni2002@gmail.com"
FILE_NAME = "Thriveni_Nano_University_DATA_02-09-2026"
INPUT_FILE = Path("urls.xlsx")

USE_LLM = True
OLLAMA_MODEL = "qwen3:4b"
OLLAMA_URL = "http://127.0.0.1:11434/api/chat"
OLLAMA_TIMEOUT = 120
LLM_CONCURRENCY = 1
LLM_MAX_TEXT = 2500

HEADLESS = True
HTTP_TIMEOUT = 25
PLAYWRIGHT_TIMEOUT = 25_000
PROFILE_TIMEOUT = 25_000
PROFILE_CONCURRENCY = 2
FOLLOW_PAGINATION = True
MAX_PAGES_PER_URL = 20
MAX_SCROLL_ROUNDS = 10
MAX_LOAD_MORE_CLICKS = 15
MIN_DELAY = 0.25
MAX_DELAY = 0.75

# Email is the only mandatory export field. Name and country are optional.
MANDATORY_FIELDS = ("email",)


# =============================================================================
# OUTPUT
# =============================================================================

def safe_filename(value: str) -> str:
    value = str(value or "").strip()
    value = re.sub(r'[<>:"/\\|?*]', "_", value)
    value = re.sub(r"\s+", "_", value)
    return value.strip("._ ") or "output"


OUTPUT_DIR = Path("output") / safe_filename(EMPLOYEE_NAME) / safe_filename(FILE_NAME)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
XLSX_FILE = OUTPUT_DIR / f"{safe_filename(FILE_NAME)}.xlsx"
ERROR_FILE = OUTPUT_DIR / "errors.json"

COLUMNS = [
    "name",
    "email",
    "country",
    "alternate_emails",
    "email_type",
    "email_conflict",
    "employee_name",
    "employee_email",
    "page_type",
    "journal_name",
    "editorial_role",
    "academic_title",
    "academic_rank",
    "specialty",
    "affiliation",
    "university",
    "faculty",
    "school",
    "department",
    "institute",
    "division",
    "city",
    "address",
    "phone",
    "orcid",
    "google_scholar",
    "scopus_author_id",
    "researcher_id",
    "pubmed",
    "profile_url",
    "personal_homepage",
    "source_url",
    "country_source",
    "confidence",
    "extraction_method",
    "scrape_status",
]


# =============================================================================
# DATA MODEL
# =============================================================================

@dataclass
class PersonRecord:
    name: str = ""
    email: str = ""
    country: str = ""
    alternate_emails: str = ""
    email_type: str = "personal"
    email_conflict: str = "no"

    employee_name: str = EMPLOYEE_NAME
    employee_email: str = EMPLOYEE_EMAIL

    page_type: str = ""
    journal_name: str = ""
    editorial_role: str = ""
    academic_title: str = ""
    academic_rank: str = ""
    specialty: str = ""

    affiliation: str = ""
    university: str = ""
    faculty: str = ""
    school: str = ""
    department: str = ""
    institute: str = ""
    division: str = ""
    city: str = ""
    address: str = ""
    phone: str = ""

    orcid: str = ""
    google_scholar: str = ""
    scopus_author_id: str = ""
    researcher_id: str = ""
    pubmed: str = ""

    profile_url: str = ""
    personal_homepage: str = ""
    source_url: str = ""

    country_source: str = ""
    confidence: int = 0
    extraction_method: str = ""
    scrape_status: str = "accepted"


# =============================================================================
# REGEX / CONSTANTS
# =============================================================================

EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)
PHONE_RE = re.compile(r"(?:(?:\+|00)\d{1,4}[\s./()\-]*)?(?:\d[\s./()\-]*){6,16}")
ORCID_RE = re.compile(r"\b\d{4}-\d{4}-\d{4}-\d{3}[\dX]\b", re.I)
SCOPUS_RE = re.compile(r"(?:Scopus(?:\s+Author)?\s*ID\s*[:#]?\s*)(\d{8,14})", re.I)

ROLE_RE = re.compile(
    r"\b(?:"
    r"editor(?:[-\s]?in[-\s]?chief)?|chief editor|co[-\s]?editor(?:[-\s]?in[-\s]?chief)?|"
    r"associate editor(?:[-\s]?in[-\s]?chief)?|assistant editor|deputy editor(?:[-\s]?in[-\s]?chief)?|"
    r"managing editor|executive editor|section editor|statistics editor|statistical reviewer|"
    r"editorial board(?: members?)?|international advisory board|advisory board|lead editors?|"
    r"technical editor|business manager|editorial manager|editorial secretary|owner"
    r")\b",
    re.I,
)

ACADEMIC_RE = re.compile(
    r"\b(?:prof(?:essor)?\.?|associate professor|assistant professor|senior lecturer|lecturer|"
    r"consultant|director|head|chair|researcher|scientist|physician|doctor|dr\.?|md|phd|msc|mph|"
    r"mbbs|frcog|facog|rn|rm)\b",
    re.I,
)

BAD_NAME_EXACT = {
    "home", "about", "contact", "contact us", "sitemap", "email", "e-mail", "orcid",
    "editorial board", "editorial team", "editor", "editors", "editor in chief", "editor-in-chief",
    "associate editors", "managing editors", "editorial office", "technical editor",
    "aims and scope", "open access", "all issues", "author index", "most cited", "most view",
    "most download", "subscription information", "browse all articles", "journal information",
    "copyright", "copyright agreement", "copyright transfer agreement", "instructions for authors",
    "advertising policies", "funded articles", "ahead-of print", "ahead of print", "reviewers",
    "secretaria", "secretaria /contact", "secretaria/contact", "contact/secretaria",
    "urologie/urology", "urology", "poland", "revista brasileira de ginecologia e obstetrícia",
    "2026-present 2022-2025 2020–2021",
}

ORG_WORDS = {
    "university", "hospital", "college", "faculty", "school", "department", "institute", "institution",
    "center", "centre", "clinic", "laboratory", "lab", "society", "association", "foundation", "office",
    "ministry", "program", "programme", "medical center", "medical centre", "research center",
    "editorial office", "business office", "journal", "academy", "corporation", "company",
    "universidade", "universidad", "universität", "universitet", "université", "universita",
    "hospital universitário", "faculdade", "departamento", "instituto",
}

ROLE_SHARED_PREFIXES = (
    "info@", "editor@", "editorial@", "office@", "secretary@", "contact@", "admin@", "journal@",
    "support@", "webmaster@", "geschaeftsstelle@", "geschäftsstelle@", "bilgi@", "rektorozelkalem@",
)

COUNTRY_ALIASES = {
    "usa": "United States", "u.s.a.": "United States", "u.s.": "United States",
    "united states of america": "United States", "united states": "United States",
    "uk": "United Kingdom", "u.k.": "United Kingdom", "england": "United Kingdom",
    "scotland": "United Kingdom", "wales": "United Kingdom", "united kingdom": "United Kingdom",
    "the netherlands": "Netherlands", "netherlands": "Netherlands",
    "south korea": "South Korea", "republic of korea": "South Korea", "korea": "South Korea",
    "p.r. china": "China", "pr china": "China", "people's republic of china": "China",
    "taiwan, province of china": "Taiwan", "taiwan": "Taiwan",
    "turkiye": "Turkey", "türkiye": "Turkey", "turkey": "Turkey",
    "uae": "United Arab Emirates", "u.a.e.": "United Arab Emirates",
    "hong kong": "Hong Kong",
    "brasil": "Brazil", "brazil": "Brazil",
    "canadá": "Canada", "canada": "Canada",
    "suécia": "Sweden", "suecia": "Sweden", "sweden": "Sweden",
    "alemanha": "Germany", "deutschland": "Germany",
    "españa": "Spain", "espana": "Spain",
    "itália": "Italy", "italia": "Italy",
    "áustria": "Austria", "austria": "Austria",
    "suíça": "Switzerland", "suica": "Switzerland",
}

TLD_COUNTRY = {
    ".au": "Australia", ".at": "Austria", ".bd": "Bangladesh", ".be": "Belgium", ".br": "Brazil",
    ".ca": "Canada", ".ch": "Switzerland", ".cn": "China", ".de": "Germany", ".dk": "Denmark",
    ".eg": "Egypt", ".es": "Spain", ".et": "Ethiopia", ".fi": "Finland", ".fr": "France",
    ".gr": "Greece", ".hk": "Hong Kong", ".id": "Indonesia", ".ie": "Ireland", ".il": "Israel",
    ".in": "India", ".ir": "Iran", ".it": "Italy", ".jp": "Japan", ".ke": "Kenya",
    ".kr": "South Korea", ".lb": "Lebanon", ".lk": "Sri Lanka", ".my": "Malaysia",
    ".ng": "Nigeria", ".nl": "Netherlands", ".no": "Norway", ".np": "Nepal", ".nz": "New Zealand",
    ".ph": "Philippines", ".pk": "Pakistan", ".pl": "Poland", ".pt": "Portugal", ".sa": "Saudi Arabia",
    ".se": "Sweden", ".sg": "Singapore", ".th": "Thailand", ".tn": "Tunisia", ".tr": "Turkey",
    ".tw": "Taiwan", ".ug": "Uganda", ".uk": "United Kingdom", ".za": "South Africa",
}

INSTITUTION_COUNTRY_HINTS = {
    "university of the philippines": "Philippines",
    "philippine general hospital": "Philippines",
    "southern philippines medical center": "Philippines",
    "baguio general hospital": "Philippines",
    "university of santo tomas": "Philippines",
    "manila doctors hospital": "Philippines",
    "st. luke's medical center": "Philippines",
    "st. luke’s medical center": "Philippines",
    "cebu doctors’ university hospital": "Philippines",
    "cebu doctors' university hospital": "Philippines",
    "veterans memorial medical center": "Philippines",
    "university of sri jayewardenepura": "Sri Lanka",
    "university of colombo": "Sri Lanka",
    "university of kelaniya": "Sri Lanka",
    "university of peradeniya": "Sri Lanka",
    "university of ruhuna": "Sri Lanka",
    "chulalongkorn university": "Thailand",
    "mahidol university": "Thailand",
    "chiang mai university": "Thailand",
    "national university of singapore": "Singapore",
    "university of tokyo": "Japan",
    "osaka university": "Japan",
    "university of athens": "Greece",
    "yale university": "United States",
    "emory university": "United States",
    "university of miami": "United States",
    "oxford university hospital": "United Kingdom",
    "imperial college": "United Kingdom",
    "st george's university of london": "United Kingdom",
    "st george’s university of london": "United Kingdom",
}


# =============================================================================
# BASIC HELPERS
# =============================================================================

def clean_text(value) -> str:
    if value is None:
        return ""
    value = str(value)
    value = (value.replace("\xa0", " ").replace("\u200b", "").replace("\ufeff", "")
             .replace("\u00ad", "").replace("\r", " ").replace("\n", " ").replace("\t", " "))
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_url(url: str) -> str:
    if not url:
        return ""
    url = clean_text(url)
    url, _ = urldefrag(url)
    return url.strip()


def unique(values: List[str]) -> List[str]:
    out = []
    seen = set()
    for v in values:
        v = clean_text(v)
        if not v:
            continue
        k = v.lower()
        if k not in seen:
            seen.add(k)
            out.append(v)
    return out


async def polite_delay():
    await asyncio.sleep(random.uniform(MIN_DELAY, MAX_DELAY))


# =============================================================================
# EMAIL NORMALIZATION
# =============================================================================

def decode_email_obfuscation(value: str) -> str:
    if not value:
        return ""
    value = unquote(str(value))
    value = value.replace("\xa0", " ").replace("\u200b", "").replace("\ufeff", "")
    value = re.sub(r"(?i)^mailto:", "", value).strip()
    value = value.rstrip("/;,)")

    for token in ["remove-this", "remove_this", "removethis", "nospam", "no-spam", "spamprotect"]:
        value = re.sub(re.escape(token), "", value, flags=re.I)

    value = re.sub(r"(?i)\s*\[\s*at\s*\]\s*", "@", value)
    value = re.sub(r"(?i)\s*\(\s*at\s*\)\s*", "@", value)
    value = re.sub(r"(?i)\s+at\s+", "@", value)
    value = re.sub(r"(?i)\s*\[\s*dot\s*\]\s*", ".", value)
    value = re.sub(r"(?i)\s*\(\s*dot\s*\)\s*", ".", value)
    value = re.sub(r"(?i)\s+dot\s+", ".", value)
    value = re.sub(r"\s*@\s*", "@", value)
    value = re.sub(r"\s*\.\s*", ".", value)
    return value.strip()


def normalize_email(value: str) -> str:
    if not value:
        return ""
    value = decode_email_obfuscation(value)
    m = EMAIL_RE.search(value)
    if not m:
        return ""
    return m.group(0).lower().strip(".,;:/")


def extract_text_emails(text: str) -> List[str]:
    text = decode_email_obfuscation(text or "")
    return unique([normalize_email(x) for x in EMAIL_RE.findall(text) if normalize_email(x)])


def extract_email_candidates(block: Tag) -> Tuple[List[str], List[str], bool]:
    """Returns primary candidates, alternates, conflict flag.
    Visible anchor text is preferred over href when both are valid and disagree.
    """
    primary = []
    alternates = []
    conflict = False

    for a in block.select('a[href^="mailto:"]'):
        href_email = normalize_email(a.get("href", ""))
        visible_email = normalize_email(a.get_text(" ", strip=True))

        if visible_email:
            primary.append(visible_email)
            if href_email and href_email != visible_email:
                alternates.append(href_email)
                conflict = True
        elif href_email:
            primary.append(href_email)

    text = clean_text(block.get_text(" ", strip=True))
    for email in extract_text_emails(text):
        if email not in primary:
            primary.append(email)

    primary = unique(primary)
    alternates = [e for e in unique(alternates) if e not in primary]
    return primary, alternates, conflict


def email_type(email: str) -> str:
    e = normalize_email(email)
    return "shared/role" if e.startswith(ROLE_SHARED_PREFIXES) else "personal"


# =============================================================================
# COUNTRY NORMALIZATION
# =============================================================================

def build_country_lookup() -> Dict[str, str]:
    lookup = dict(COUNTRY_ALIASES)
    if pycountry:
        for c in pycountry.countries:
            lookup[c.name.lower()] = c.name
            if hasattr(c, "official_name"):
                lookup[c.official_name.lower()] = c.name
            if hasattr(c, "common_name"):
                lookup[c.common_name.lower()] = c.name
            lookup[c.alpha_2.lower()] = c.name
            lookup[c.alpha_3.lower()] = c.name
    # Preferred canonical forms
    lookup["korea, republic of"] = "South Korea"
    lookup["iran, islamic republic of"] = "Iran"
    lookup["russian federation"] = "Russia"
    lookup["viet nam"] = "Vietnam"
    lookup["bolivia, plurinational state of"] = "Bolivia"
    lookup["tanzania, united republic of"] = "Tanzania"
    return lookup


COUNTRY_LOOKUP = build_country_lookup()


def normalize_country(value: str) -> str:
    value = clean_text(value).strip(" ,.;:-")
    if not value:
        return ""
    lower = value.lower()
    if lower in COUNTRY_LOOKUP:
        return COUNTRY_LOOKUP[lower]
    if lower in COUNTRY_ALIASES:
        return COUNTRY_ALIASES[lower]
    return ""


def extract_country_from_text(text: str) -> str:
    text = clean_text(text)
    if not text:
        return ""
    lower = text.lower()

    matches = []
    # Aliases first, including phrases such as "Taiwan, Province of China"
    for alias, canonical in COUNTRY_ALIASES.items():
        for m in re.finditer(rf"(?<![\w]){re.escape(alias)}(?![\w])", lower, flags=re.I):
            matches.append((m.start(), len(alias), canonical))

    if pycountry:
        for country in pycountry.countries:
            names = [country.name]
            if hasattr(country, "official_name"):
                names.append(country.official_name)
            if hasattr(country, "common_name"):
                names.append(country.common_name)
            for name in names:
                if len(name) < 4:
                    continue
                for m in re.finditer(rf"(?<![\w]){re.escape(name.lower())}(?![\w])", lower, flags=re.I):
                    canon = normalize_country(name) or name
                    matches.append((m.start(), len(name), canon))

    if not matches:
        return ""
    # Prefer longest mention; if tie, the last occurrence (often affiliation suffix)
    matches.sort(key=lambda x: (x[1], x[0]), reverse=True)
    return matches[0][2]


def country_from_tld(email: str) -> str:
    e = normalize_email(email)
    if not e:
        return ""
    domain = e.split("@", 1)[1].lower()
    for suffix, country in TLD_COUNTRY.items():
        if domain.endswith(suffix):
            return country
    return ""


def country_from_institution(text: str) -> str:
    low = clean_text(text).lower()
    for phrase, country in INSTITUTION_COUNTRY_HINTS.items():
        if phrase in low:
            return country
    return ""


# =============================================================================
# NAME VALIDATION
# =============================================================================

def strip_credentials(text: str) -> str:
    text = clean_text(text)
    text = re.sub(r"(?i)\s*:\s*h\s*index\s*\d+.*$", "", text)
    text = re.sub(r"(?i)\s*\(\s*(?:md|phd|msc|mph|drph|facog|frcog|mbbs|bsc|ma|mba)(?:[^)]*)\)\s*$", "", text)
    # Keep Prof/Dr prefixes, drop common trailing degrees.
    text = re.sub(r"(?i)(?:,|\s)+(?:MD|M\.D\.|PhD|Ph\.D\.|MSc|M\.Sc\.|MPH|MBA|FACOG|FRCOG|MBBS|BSc|MA)(?:[,\s].*)?$", "", text)
    return clean_text(text.strip(" ,;:-"))


def clean_person_name_candidate(text: str) -> str:
    """
    Clean only obvious affiliation/metadata contamination from a proposed name.

    Important: this never tries to manufacture a person name. It only shortens a
    candidate when the discarded suffix contains clear organisation/country data.
    """
    text = strip_credentials(text)
    text = text.replace("\u200b", "").replace("\ufeff", "")
    text = clean_text(text)
    if not text:
        return ""

    # Mojibake/date/navigation/category strings are never names.
    low = text.lower().strip(" :;,.|")
    if re.search(r"\b(?:19|20)\d{2}\s*[-–]\s*(?:present|(?:19|20)\d{2})", low, re.I):
        return ""
    if re.fullmatch(r"[a-zà-öø-ÿā-ž]+\s*/\s*[a-zà-öø-ÿā-ž]+", low, re.I):
        return ""

    # Remove obvious trailing affiliation after an en/em dash.
    for sep in (" – ", " — ", " - "):
        if sep in text:
            head, tail = text.split(sep, 1)
            tail_low = tail.lower()
            if (
                any(w in tail_low for w in ORG_WORDS)
                or bool(extract_country_from_text(tail))
                or re.search(r"\b(?:faculty|universidade|universidad|universität|universite|hospital|department|institute|centre|center)\b", tail_low, re.I)
            ):
                text = head.strip(" ,;:-")
                break

    # Comma-separated contamination: Diletta Marcolin, AULSS 7, ..., Italy
    # Keep surname-first names such as 'Aflatoonian, Abbas' by truncating only
    # when the suffix clearly looks like affiliation/geography.
    parts = [clean_text(x) for x in text.split(",")]
    if len(parts) >= 2:
        for i in range(1, len(parts)):
            tail = ", ".join(parts[i:])
            tail_low = tail.lower()
            if (
                any(w in tail_low for w in ORG_WORDS)
                or bool(extract_country_from_text(tail))
                or bool(re.search(r"\b(?:universidade|universidad|universität|university|hospital|faculty|department|institute|centre|center|aulss|clinic)\b", tail_low, re.I))
            ):
                head = ", ".join(parts[:i]).strip(" ,;:-")
                if len(head.split()) >= 2:
                    text = head
                break

    text = strip_credentials(text)
    return clean_text(text.strip(" ,;:-"))


def name_from_email(email: str) -> str:
    """
    Conservative last-resort display name derived only from the email local-part.

    Returns blank for role/shared/generic/numeric mailboxes. This is intentional:
    name is optional, and a blank name is safer than a false identity.
    """
    email = normalize_email(email)
    if not email:
        return ""

    local = email.split("@", 1)[0].strip().lower()
    local = re.sub(r"\+.*$", "", local)

    generic_tokens = {
        "info", "contact", "office", "editor", "editorial", "journal", "secretary",
        "secretaria", "admin", "support", "webmaster", "submission", "submissions",
        "help", "mail", "email", "enquiries", "inquiries", "business", "manager",
        "reception", "head", "hq", "brz", "geschaeftsstelle", "geschäftsstelle",
        "bilgi", "rektorozelkalem", "objhead",
    }
    if local in generic_tokens or any(local.startswith(x + ".") or local.startswith(x + "_") for x in generic_tokens):
        return ""
    if re.fullmatch(r"\d+", local):
        return ""

    # Human-readable only when separators provide credible token boundaries.
    if re.search(r"[._-]", local):
        tokens = [t for t in re.split(r"[._-]+", local) if t]
    else:
        # A single alphabetic mailbox can be a first name but is weak evidence.
        tokens = [local]

    cleaned = []
    for token in tokens:
        token = re.sub(r"[^a-zà-öø-ÿā-ž]", "", token, flags=re.I)
        if not token or token in generic_tokens:
            continue
        if len(token) == 1:
            cleaned.append(token.upper() + ".")
        elif len(token) >= 2:
            cleaned.append(token[:1].upper() + token[1:])

    if not cleaned:
        return ""

    candidate = " ".join(cleaned)
    # For one token require a reasonable personal-looking mailbox.
    if len(cleaned) == 1 and len(re.sub(r"\W", "", cleaned[0])) < 4:
        return ""
    if not plausible_name(candidate):
        return ""
    return candidate


def looks_like_url(text: str) -> bool:
    t = clean_text(text).lower()
    return t.startswith(("http://", "https://", "www.")) or "://" in t


def is_role_heading(text: str) -> bool:
    t = clean_text(text)
    return bool(t and ROLE_RE.search(t) and len(t.split()) <= 10)


def plausible_name(text: str) -> bool:
    raw = clean_text(text)
    if not raw or len(raw) < 2 or len(raw) > 130:
        return False
    if looks_like_url(raw) or "@" in raw:
        return False
    low = raw.lower().strip(" :")
    if low in BAD_NAME_EXACT:
        return False
    if re.search(r"\b(?:19|20)\d{2}\s*[-–]\s*(?:present|(?:19|20)\d{2})", low, re.I):
        return False
    if re.fullmatch(r"[a-zà-öø-ÿā-ž]+\s*/\s*[a-zà-öø-ÿā-ž]+", low, re.I):
        return False
    if any(k in low for k in ("contact us", "secretaria /contact", "revista brasileira de ginecologia", "journal of", "editorial office")):
        return False
    if is_role_heading(raw):
        return False
    if normalize_country(raw):
        return False
    if re.fullmatch(r"[\d\W_]+", raw):
        return False
    words = raw.split()
    if len(words) > 16:
        return False

    # Organization-heavy phrases are not people unless they contain a clear title and several name-like tokens
    org_hits = sum(1 for w in ORG_WORDS if w in low)
    if org_hits >= 1 and not re.search(r"\b(?:prof|dr|md|phd)\b", low):
        return False

    # Must contain letters
    return bool(re.search(r"[A-Za-zÀ-ÖØ-öø-ÿĀ-žİıŞşĞğÇçÖöÜüÉéÁáÍíÓóÚúÑñ]", raw))


def normalize_name_key(name: str) -> str:
    name = strip_credentials(name).lower()
    name = re.sub(r"\b(?:prof(?:essor)?|dr|md|phd|msc|mph|mbbs|frcog|facog)\b", " ", name, flags=re.I)
    name = re.sub(r"[^a-zà-öø-ÿā-ž]+", " ", name)
    return " ".join(sorted(x for x in name.split() if len(x) > 1))


def email_name_score(name: str, email: str) -> int:
    if not name or not email:
        return 0
    local = email.split("@", 1)[0].lower()
    local_norm = re.sub(r"[^a-z]", "", local)
    tokens = re.findall(r"[a-zà-öø-ÿā-ž]+", strip_credentials(name).lower())
    tokens = [re.sub(r"[^a-z]", "", t) for t in tokens]
    tokens = [t for t in tokens if len(t) >= 3]
    if not tokens:
        return 0
    score = 0
    for t in tokens:
        if t and t in local_norm:
            score += 20
    return min(score, 60)


def collect_name_candidates(block: Tag, email: str = "") -> List[Tuple[int, str]]:
    candidates = []
    selector_scores = [
        ("h1,h2,h3,h4,h5,h6", 65),
        ("strong,b", 60),
        ('a[href^="mailto:"]', 70),
        ('[class*="name"]', 60),
        ("cite", 20),
    ]
    for selector, base in selector_scores:
        for el in block.select(selector):
            text = clean_person_name_candidate(el.get_text(" ", strip=True))
            if plausible_name(text):
                s = base + email_name_score(text, email)
                candidates.append((s, text))

    # Lines in local text are useful for flow/table cases
    raw = block.get_text("\n", strip=True)
    for line in raw.splitlines():
        text = clean_person_name_candidate(line)
        if plausible_name(text) and len(text.split()) <= 10:
            candidates.append((35 + email_name_score(text, email), text))

    best = {}
    for score, text in candidates:
        key = text.lower()
        if score > best.get(key, -1):
            best[key] = score
    return sorted([(s, k) for k, s in best.items()], reverse=True)


# =============================================================================
# LINKS / EXTRA FIELDS
# =============================================================================

def extract_links(block: Tag, base_url: str) -> Dict[str, str]:
    out = {"orcid": "", "google_scholar": "", "pubmed": "", "personal_homepage": ""}
    for a in block.select("a[href]"):
        href = clean_text(a.get("href", ""))
        if not href or href.startswith(("mailto:", "javascript:", "#")):
            continue
        url = normalize_url(urljoin(base_url, href))
        low = url.lower()
        text = clean_text(a.get_text(" ", strip=True)).lower()
        if "orcid.org" in low and not out["orcid"]:
            m = ORCID_RE.search(url) or ORCID_RE.search(text)
            out["orcid"] = f"https://orcid.org/{m.group(0)}" if m else url
        elif "scholar.google." in low and not out["google_scholar"]:
            out["google_scholar"] = url
        elif "pubmed" in low and not out["pubmed"]:
            out["pubmed"] = url
        elif any(k in text for k in ("homepage", "personal website", "personal page")) and not out["personal_homepage"]:
            out["personal_homepage"] = url
    return out


def extract_phone(block: Tag) -> str:
    for a in block.select('a[href^="tel:"]'):
        t = clean_text(a.get_text(" ", strip=True)) or clean_text(a.get("href", "").replace("tel:", ""))
        if t:
            return t
    text = clean_text(block.get_text(" ", strip=True))
    for m in PHONE_RE.findall(text):
        digits = re.sub(r"\D", "", m)
        if 7 <= len(digits) <= 16:
            return clean_text(m)
    return ""


def extract_scopus(text: str) -> str:
    m = SCOPUS_RE.search(clean_text(text))
    return m.group(1) if m else ""


def extract_orcid_text(text: str) -> str:
    m = ORCID_RE.search(text or "")
    return f"https://orcid.org/{m.group(0)}" if m else ""


def extract_js_profile_url(href: str, base_url: str) -> str:
    if not href:
        return ""
    # javascript:openRTWindow('https://...')
    m = re.search(r"['\"](https?://[^'\"]+)['\"]", href)
    if m:
        return normalize_url(m.group(1))
    m = re.search(r"openRTWindow\(['\"]([^'\"]+)['\"]\)", href, re.I)
    if m:
        return normalize_url(urljoin(base_url, m.group(1)))
    return ""


# =============================================================================
# DOM NORMALIZATION
# =============================================================================

def normalize_dom(soup: BeautifulSoup):
    """Normalize visual @ signs and remove high-noise elements."""
    # @ icons/images
    for img in list(soup.find_all("img")):
        attrs = " ".join([
            clean_text(img.get("src", "")), clean_text(img.get("alt", "")), clean_text(img.get("title", "")),
            " ".join(img.get("class", [])),
        ]).lower()
        if any(k in attrs for k in ("em_sign", "email_sign", "at_sign", "at-sign", "arroba")):
            img.replace_with(NavigableString("@"))

    for i in list(soup.find_all(["i", "span"])):
        classes = " ".join(i.get("class", [])).lower()
        aria = clean_text(i.get("aria-label", "")).lower()
        if "fa-at" in classes or "icon-at" in classes or aria == "@":
            i.replace_with(NavigableString("@"))

    # Noise
    for selector in [
        "script", "style", "noscript", "svg", "footer", "nav", "aside",
        ".sidebar", ".breadcrumb", ".breadcrumbs", ".advertisement", ".ads", ".cookie",
        ".cookie-banner", ".social", ".social-media",
    ]:
        for el in soup.select(selector):
            el.decompose()


# =============================================================================
# HTTP / PLAYWRIGHT FETCH
# =============================================================================

def create_http_session() -> requests.Session:
    s = requests.Session()
    retry = Retry(total=2, connect=2, read=2, backoff_factor=0.6,
                  status_forcelist=[429, 500, 502, 503, 504], allowed_methods=["GET"])
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return s


HTTP_SESSION = create_http_session()


def fetch_http_sync(url: str) -> Dict:
    try:
        r = HTTP_SESSION.get(url, timeout=(10, HTTP_TIMEOUT), allow_redirects=True)
        if r.status_code >= 400:
            return {"ok": False, "error": f"HTTP {r.status_code}", "html": "", "url": url}
        html = r.text or ""
        if len(html) < 300:
            return {"ok": False, "error": "HTML too small", "html": html, "url": r.url}
        low = html.lower()
        if any(x in low for x in ["verify you are human", "checking your browser", "captcha", "access denied"]):
            return {"ok": False, "error": "possible anti-bot page", "html": html, "url": r.url}
        return {"ok": True, "error": "", "html": html, "url": r.url}
    except Exception as exc:
        return {"ok": False, "error": str(exc), "html": "", "url": url}


async def fetch_http(url: str) -> Dict:
    return await asyncio.to_thread(fetch_http_sync, url)


async def playwright_fetch(context, url: str) -> Dict:
    page = await context.new_page()
    try:
        print(f"   [PLAYWRIGHT] {url}", flush=True)
        try:
            resp = await page.goto(url, wait_until="commit", timeout=PLAYWRIGHT_TIMEOUT)
        except PlaywrightTimeoutError:
            return {"ok": False, "error": "Playwright timeout", "html": "", "url": url}
        try:
            await page.wait_for_selector("body", timeout=8000)
        except Exception:
            pass
        await dismiss_popups(page)
        await auto_scroll(page)
        await click_load_more(page)
        await page.wait_for_timeout(800)
        html = await page.content()
        if resp and resp.status and resp.status >= 400:
            return {"ok": False, "error": f"HTTP {resp.status}", "html": html, "url": page.url}
        return {"ok": True, "error": "", "html": html, "url": page.url}
    except Exception as exc:
        return {"ok": False, "error": str(exc), "html": "", "url": url}
    finally:
        await page.close()


async def dismiss_popups(page):
    for label in ["Accept all", "Accept All", "Accept cookies", "Allow all", "I agree", "Agree", "OK", "Akzeptieren"]:
        try:
            b = page.get_by_role("button", name=label, exact=True).first
            if await b.count() and await b.is_visible():
                await b.click(timeout=1200)
                return
        except Exception:
            pass


async def auto_scroll(page):
    previous = -1
    stable = 0
    for _ in range(MAX_SCROLL_ROUNDS):
        try:
            height = await page.evaluate("document.body.scrollHeight")
            if height == previous:
                stable += 1
            else:
                stable = 0
            if stable >= 2:
                break
            previous = height
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(450)
        except Exception:
            break


async def click_load_more(page):
    for _ in range(MAX_LOAD_MORE_CLICKS):
        clicked = False
        for label in ["Load more", "Load More", "Show more", "Show More", "More results", "View more"]:
            try:
                x = page.get_by_text(label, exact=True).first
                if await x.count() and await x.is_visible():
                    await x.click(timeout=1500)
                    await page.wait_for_timeout(600)
                    clicked = True
                    break
            except Exception:
                pass
        if not clicked:
            break


# =============================================================================
# OLLAMA - ONLY RESOLVES EXISTING CANDIDATES
# =============================================================================

def ollama_available_sync() -> bool:
    """Return True only when Ollama is reachable and the configured local model exists."""
    if not USE_LLM:
        return False
    try:
        r = requests.get("http://127.0.0.1:11434/api/tags", timeout=5)
        if r.status_code != 200:
            return False
        models = r.json().get("models", [])
        names = {
            str(m.get("name", "")).casefold()
            for m in models
            if isinstance(m, dict)
        }
        wanted = OLLAMA_MODEL.casefold()
        return wanted in names or any(x.startswith(wanted + ":") for x in names)
    except Exception:
        return False


async def ollama_available() -> bool:
    return await asyncio.to_thread(ollama_available_sync)


def extract_json_object(text: str):
    if not text:
        return None
    text = re.sub(r"^```(?:json)?", "", text.strip(), flags=re.I)
    text = re.sub(r"```$", "", text).strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    s, e = text.find("{"), text.rfind("}")
    if s >= 0 and e > s:
        try:
            return json.loads(text[s:e+1])
        except Exception:
            return None
    return None


def llm_choose_name_sync(text: str, names: List[str], email: str) -> Optional[int]:
    if not names:
        return None
    prompt = f"""
You resolve a university/editorial-board person record.
You MAY ONLY select an index from names. Do not invent anything.

Email: {email}
Name candidates: {json.dumps(names, ensure_ascii=False)}
Local text: {clean_text(text)[:LLM_MAX_TEXT]}

Return JSON only:
{{"valid_person": true, "name_index": 0, "confidence": 0}}

Rules:
- Reject hospitals, universities, departments, countries, editorial role headings, navigation labels, awards and URLs.
- The selected name must be the human associated with the email.
- If uncertain return valid_person=false and name_index=-1.
""".strip()
    try:
        r = requests.post(OLLAMA_URL, json={
            "model": OLLAMA_MODEL,
            "stream": False,
            "options": {"temperature": 0, "num_predict": 120},
            "messages": [{"role": "user", "content": prompt}],
        }, timeout=OLLAMA_TIMEOUT)
        r.raise_for_status()
        data = extract_json_object(r.json().get("message", {}).get("content", ""))
        if not data or not data.get("valid_person"):
            return None
        idx = int(data.get("name_index", -1))
        return idx if 0 <= idx < len(names) else None
    except Exception:
        return None


async def llm_choose_name(semaphore, text: str, names: List[str], email: str) -> Optional[int]:
    async with semaphore:
        return await asyncio.to_thread(llm_choose_name_sync, text, names, email)


def llm_country_sync(affiliation: str) -> str:
    prompt = f"""
Identify the country explicitly or unambiguously associated with this institution/affiliation.
Return JSON only: {{"country":"Country Name","confidence":0}}
If uncertain use an empty country.
Affiliation: {clean_text(affiliation)[:1200]}
""".strip()
    try:
        r = requests.post(OLLAMA_URL, json={
            "model": OLLAMA_MODEL, "stream": False,
            "options": {"temperature": 0, "num_predict": 80},
            "messages": [{"role": "user", "content": prompt}],
        }, timeout=OLLAMA_TIMEOUT)
        r.raise_for_status()
        data = extract_json_object(r.json().get("message", {}).get("content", ""))
        if not data:
            return ""
        country = normalize_country(data.get("country", ""))
        conf = int(data.get("confidence", 0) or 0)
        return country if country and conf >= 85 else ""
    except Exception:
        return ""


async def llm_country(semaphore, affiliation: str) -> str:
    async with semaphore:
        return await asyncio.to_thread(llm_country_sync, affiliation)


# =============================================================================
# PAGE META
# =============================================================================

def detect_page_type(soup: BeautifulSoup, url: str) -> str:
    low_url = url.lower()
    text = clean_text(soup.get_text(" ", strip=True)).lower()[:12000]
    if any(k in low_url for k in ["editorial", "editorialteam", "editorial-board", "editorial.board", "/board"]):
        return "EDITORIAL_BOARD"
    if any(k in text for k in ["editorial board", "editor-in-chief", "editorial team", "associate editors"]):
        return "EDITORIAL_BOARD"
    return "UNIVERSITY_DIRECTORY"


def journal_name(soup: BeautifulSoup) -> str:
    for sel in [".journal-title", ".journal-name", "header h1", "h1"]:
        el = soup.select_one(sel)
        if el:
            t = clean_text(el.get_text(" ", strip=True))
            if t and "editorial board" not in t.lower() and "editorial team" not in t.lower() and len(t) < 180:
                return t
    if soup.title:
        return clean_text(soup.title.get_text(" ", strip=True))[:180]
    return ""


def current_role_for(node: Tag) -> str:
    # nearest previous heading or role paragraph
    for prev in node.find_all_previous(["h1", "h2", "h3", "h4", "h5", "h6", "p", "th"], limit=25):
        t = clean_text(prev.get_text(" ", strip=True))
        if is_role_heading(t):
            return t
    return ""


def current_specialty_for(node: Tag) -> str:
    for prev in node.find_all_previous(["h2", "h3", "h4", "h5", "p", "b", "strong"], limit=15):
        t = clean_text(prev.get_text(" ", strip=True))
        if not t or len(t) > 100 or is_role_heading(t):
            continue
        if "/" in t and not plausible_name(t):
            return t
    return ""


# =============================================================================
# RECORD CONSTRUCTION
# =============================================================================

def split_affiliation_from_block(block: Tag, name: str, emails: List[str], role: str) -> str:
    text = clean_text(block.get_text(" ", strip=True))
    for x in [name, role] + emails:
        if x:
            text = re.sub(re.escape(x), " ", text, flags=re.I)
    text = re.sub(r"(?i)\b(?:email(?: address)?|e-mail|mail)\s*:?", " ", text)
    text = EMAIL_RE.sub(" ", text)
    text = clean_text(text)
    return text[:1000]


def resolve_country(record: PersonRecord, local_text: str = "") -> PersonRecord:
    # explicit already set
    if record.country:
        record.country = normalize_country(record.country) or record.country
        if not record.country_source:
            record.country_source = "explicit"
        return record

    for text, source in [
        (record.affiliation, "affiliation"),
        (record.address, "address"),
        (local_text, "local_text"),
    ]:
        c = extract_country_from_text(text)
        if c:
            record.country = c
            record.country_source = source
            return record

    c = country_from_institution(record.affiliation or local_text)
    if c:
        record.country = c
        record.country_source = "institution_hint"
        return record

    c = country_from_tld(record.email)
    if c:
        record.country = c
        record.country_source = "email_tld"
        return record
    return record


def make_record_from_block(
    block: Tag,
    source_url: str,
    page_type: str,
    journal: str,
    method: str,
    inherited_role: str = "",
    explicit_country: str = "",
    explicit_name: str = "",
    explicit_affiliation: str = "",
    explicit_email: str = "",
) -> Optional[PersonRecord]:
    text = clean_text(block.get_text(" ", strip=True))
    primary_emails, alternates, conflict = extract_email_candidates(block)
    if explicit_email:
        e = normalize_email(explicit_email)
        if e and e not in primary_emails:
            primary_emails.insert(0, e)
    if not primary_emails:
        return None
    email = primary_emails[0]

    explicit_clean = clean_person_name_candidate(explicit_name) if explicit_name else ""
    if explicit_clean and plausible_name(explicit_clean):
        name = explicit_clean
        name_score = 95
    else:
        candidates = collect_name_candidates(block, email)
        if candidates:
            name_score, name = candidates[0]
            name = clean_person_name_candidate(name)
        else:
            name_score, name = 0, ""

    # Name is OPTIONAL. If DOM evidence is not trustworthy, derive a conservative
    # display name from the email. For generic/shared mailboxes keep name blank.
    if not name or not plausible_name(name):
        name = name_from_email(email)
        name_score = 25 if name else 0

    role = inherited_role or current_role_for(block)
    links = extract_links(block, source_url)
    affiliation = explicit_affiliation or split_affiliation_from_block(block, name, primary_emails + alternates, role)

    rec = PersonRecord(
        name=name,
        email=email,
        alternate_emails=" | ".join(unique(alternates + primary_emails[1:])),
        email_type=email_type(email),
        email_conflict="yes" if conflict else "no",
        page_type=page_type,
        journal_name=journal,
        editorial_role=role if page_type == "EDITORIAL_BOARD" else "",
        affiliation=affiliation,
        country=normalize_country(explicit_country),
        phone=extract_phone(block),
        orcid=links["orcid"] or extract_orcid_text(text),
        google_scholar=links["google_scholar"],
        scopus_author_id=extract_scopus(text),
        pubmed=links["pubmed"],
        personal_homepage=links["personal_homepage"],
        source_url=source_url,
        confidence=min(100, 55 + name_score // 3 + (20 if explicit_country else 0) + (10 if role else 0)),
        extraction_method=method,
    )
    rec = resolve_country(rec, text)
    return rec


# =============================================================================
# ADAPTER 1: REPEATED CARDS / OJS GROUP MEMBERS / KNE / IJFS-like
# =============================================================================

def adapter_cards(soup: BeautifulSoup, source_url: str, page_type: str, journal: str) -> List[PersonRecord]:
    records = []
    selectors = [
        ".editor-profile", ".member-item", "li[data-uid]", ".editorialTeam .member", ".editorialTeam li",
        '[id^="edb"]', ".board-member", ".editorial-board-member", ".editor-item", ".person-card",
    ]
    seen_nodes = set()
    for sel in selectors:
        for block in soup.select(sel):
            if id(block) in seen_nodes:
                continue
            seen_nodes.add(id(block))

            # Placentum / OJS JS profile link
            js_profile = ""
            explicit_name = ""
            for a in block.select("a[href]"):
                href = a.get("href", "")
                if "openRTWindow" in href:
                    js_profile = extract_js_profile_url(href, source_url)
                    explicit_name = clean_text(a.get_text(" ", strip=True))
                    break

            explicit_country = ""
            country_el = block.select_one(".country")
            if country_el:
                explicit_country = extract_country_from_text(country_el.get_text(" ", strip=True)) or clean_text(country_el.get_text(" ", strip=True))

            explicit_affiliation = ""
            aff = block.select_one(".afiliacja_button .Data, .affiliation, cite")
            if aff:
                explicit_affiliation = clean_text(aff.get_text(" ", strip=True) or aff.get("title", ""))

            rec = make_record_from_block(
                block, source_url, page_type, journal, "card",
                inherited_role=current_role_for(block),
                explicit_country=explicit_country,
                explicit_name=explicit_name,
                explicit_affiliation=explicit_affiliation,
            )
            if rec:
                if js_profile:
                    rec.profile_url = js_profile
                records.append(rec)
    return records


# =============================================================================
# ADAPTER 2: LIST ITEMS / SCIELO
# =============================================================================

def adapter_list_items(soup: BeautifulSoup, source_url: str, page_type: str, journal: str) -> List[PersonRecord]:
    records = []
    for li in soup.select("li"):
        text = clean_text(li.get_text(" ", strip=True))
        if "@" not in decode_email_obfuscation(text) and not li.select_one('a[href^="mailto:"]'):
            continue
        if len(text) > 1800:
            continue
        rec = make_record_from_block(li, source_url, page_type, journal, "list_item", inherited_role=current_role_for(li))
        if rec:
            records.append(rec)
    return records


# =============================================================================
# ADAPTER 3: TABLE SAME ROW / GRS / ECERM
# =============================================================================

def adapter_table_rows(soup: BeautifulSoup, source_url: str, page_type: str, journal: str) -> List[PersonRecord]:
    records = []
    for tr in soup.select("tr"):
        text = clean_text(tr.get_text(" ", strip=True))
        if not text or len(text) > 2200:
            continue
        if tr.select_one('a[href^="mailto:"]') or extract_text_emails(text):
            # If row has multiple independent cells/persons, parse td chunks first
            email_cells = [td for td in tr.find_all(["td", "th"], recursive=False)
                           if td.select_one('a[href^="mailto:"]') or extract_text_emails(td.get_text(" ", strip=True))]
            if len(email_cells) > 1:
                for td in email_cells:
                    rec = make_record_from_block(td, source_url, page_type, journal, "table_cell", inherited_role=current_role_for(tr))
                    if rec:
                        records.append(rec)
                continue
            rec = make_record_from_block(tr, source_url, page_type, journal, "table_row", inherited_role=current_role_for(tr))
            if rec:
                records.append(rec)
    return records


# =============================================================================
# ADAPTER 4: TABLE MULTI-ROW (THAI example)
# =============================================================================

def adapter_table_multirow(soup: BeautifulSoup, source_url: str, page_type: str, journal: str) -> List[PersonRecord]:
    records = []
    for table in soup.select("table"):
        rows = table.find_all("tr", recursive=False)
        if not rows:
            rows = table.select("tr")
        current_role = ""
        i = 0
        while i < len(rows):
            row = rows[i]
            cells = row.find_all(["td", "th"], recursive=False)
            vals = [clean_text(c.get_text(" ", strip=True)) for c in cells]
            row_text = clean_text(" | ".join(vals))
            if row_text and is_role_heading(row_text):
                current_role = row_text
                i += 1
                continue

            # Candidate data row: name | institution | country
            if len(cells) >= 3:
                name_text = strip_credentials(vals[0])
                country = extract_country_from_text(vals[-1])
                if plausible_name(name_text) and (country or vals[1]):
                    # next one/two rows may contain email
                    collected = BeautifulSoup("<div></div>", "html.parser").div
                    for c in cells:
                        collected.append(BeautifulSoup(str(c), "html.parser"))
                    explicit_email = ""
                    if i + 1 < len(rows):
                        nxt = rows[i + 1]
                        p, a, conf = extract_email_candidates(nxt)
                        if p:
                            explicit_email = p[0]
                            collected.append(BeautifulSoup(str(nxt), "html.parser"))
                            i += 1
                    if explicit_email:
                        rec = make_record_from_block(
                            collected, source_url, page_type, journal, "table_multirow",
                            inherited_role=current_role,
                            explicit_country=country,
                            explicit_name=name_text,
                            explicit_affiliation=vals[1],
                            explicit_email=explicit_email,
                        )
                        if rec:
                            # preserve visible/href email conflict from next row via fresh extraction
                            if i < len(rows):
                                prim, alts, conflict = extract_email_candidates(rows[i])
                                if prim:
                                    rec.email = prim[0]
                                    rec.alternate_emails = " | ".join(unique(alts + prim[1:]))
                                    rec.email_conflict = "yes" if conflict else "no"
                            records.append(rec)
            i += 1
    return records


# =============================================================================
# ADAPTER 5: ONE-PER-PARAGRAPH
# =============================================================================

def adapter_paragraphs(soup: BeautifulSoup, source_url: str, page_type: str, journal: str) -> List[PersonRecord]:
    records = []
    for p in soup.select("p"):
        text = clean_text(p.get_text(" ", strip=True))
        if not text or len(text) > 1800:
            continue
        if not (p.select_one('a[href^="mailto:"]') or extract_text_emails(text)):
            continue
        rec = make_record_from_block(p, source_url, page_type, journal, "paragraph", inherited_role=current_role_for(p))
        if rec:
            records.append(rec)
    return records


# =============================================================================
# ADAPTER 6: SEQUENTIAL FLOW (name p -> affiliation p -> email p -> ORCID p)
# =============================================================================

def adapter_sequential_flow(soup: BeautifulSoup, source_url: str, page_type: str, journal: str) -> List[PersonRecord]:
    records = []
    roots = [soup.select_one("main"), soup.select_one("#main-content"), soup.select_one("#content"), soup.body]
    root = next((x for x in roots if x), soup)

    children = [x for x in root.find_all(["h1", "h2", "h3", "h4", "h5", "h6", "p", "div"], recursive=True)]
    current_role = ""
    i = 0
    while i < len(children):
        node = children[i]
        text = clean_text(node.get_text(" ", strip=True))
        if not text or len(text) > 300:
            i += 1
            continue
        if is_role_heading(text) and node.name in ["h1", "h2", "h3", "h4", "h5", "h6", "p"]:
            current_role = text
            i += 1
            continue

        # Strong-only or short name paragraph starts a person
        strong = node.find(["strong", "b"], recursive=True)
        name_candidate = strip_credentials(strong.get_text(" ", strip=True)) if strong else strip_credentials(text)
        if not plausible_name(name_candidate):
            i += 1
            continue

        # gather next siblings/items until next person/role, max 6 items
        fragments = [node]
        for j in range(i + 1, min(len(children), i + 7)):
            nxt = children[j]
            nt = clean_text(nxt.get_text(" ", strip=True))
            if is_role_heading(nt) and nxt.name in ["h1", "h2", "h3", "h4", "h5", "h6"]:
                break
            nstrong = nxt.find(["strong", "b"], recursive=True)
            nname = strip_credentials(nstrong.get_text(" ", strip=True)) if nstrong else ""
            if nname and plausible_name(nname) and (nxt.select_one('a[href^="mailto:"]') is None):
                break
            fragments.append(nxt)
            if nxt.select_one('a[href^="mailto:"]') or extract_text_emails(nt):
                # include one more line for ORCID then stop
                if j + 1 < len(children):
                    extra = children[j + 1]
                    et = clean_text(extra.get_text(" ", strip=True))
                    if "orcid" in et.lower():
                        fragments.append(extra)
                break

        html = "<div>" + "".join(str(x) for x in fragments) + "</div>"
        block = BeautifulSoup(html, "html.parser").div
        if block.select_one('a[href^="mailto:"]') or extract_text_emails(block.get_text(" ", strip=True)):
            rec = make_record_from_block(
                block, source_url, page_type, journal, "sequential_flow",
                inherited_role=current_role, explicit_name=name_candidate,
            )
            if rec:
                records.append(rec)
        i += 1
    return records


# =============================================================================
# ADAPTER 7: GENERIC EMAIL-ANCHORED FALLBACK
# =============================================================================

def smallest_email_block(anchor: Tag) -> Tag:
    node = anchor
    for _ in range(6):
        parent = node.parent if isinstance(node, Tag) else None
        if not parent or parent.name in ["body", "html"]:
            break
        text = clean_text(parent.get_text(" ", strip=True))
        emails = extract_text_emails(text)
        if len(text) <= 1800 and len(emails) <= 3:
            node = parent
        else:
            break
    return node if isinstance(node, Tag) else anchor


def adapter_email_anchors(soup: BeautifulSoup, source_url: str, page_type: str, journal: str) -> List[PersonRecord]:
    records = []
    seen = set()
    anchors = soup.select('a[href^="mailto:"]')
    # also wrap plain-text email elements by scanning paragraphs/divs
    for a in anchors:
        block = smallest_email_block(a)
        sig = clean_text(block.get_text(" ", strip=True))[:1000]
        if sig in seen:
            continue
        seen.add(sig)
        rec = make_record_from_block(block, source_url, page_type, journal, "email_anchor", inherited_role=current_role_for(block))
        if rec:
            records.append(rec)

    for block in soup.select("p, li, td, div"):
        text = clean_text(block.get_text(" ", strip=True))
        if "@" not in decode_email_obfuscation(text):
            continue
        if block.select_one('a[href^="mailto:"]'):
            continue
        if len(text) > 1200:
            continue
        sig = text[:1000]
        if sig in seen:
            continue
        seen.add(sig)
        rec = make_record_from_block(block, source_url, page_type, journal, "plain_email_fallback", inherited_role=current_role_for(block))
        if rec:
            records.append(rec)
    return records


# =============================================================================
# UNIVERSITY DIRECTORY ADAPTER (profiles + cards)
# =============================================================================

def looks_like_profile_url(url: str) -> bool:
    if not url:
        return False
    p = urlparse(url)
    path = p.path.lower()
    query = p.query.lower()
    if any(x in path for x in ["/disciplines", "/find-an-expert", "/search", "/category", "/editorial-board"]):
        return False
    if any(x in query for x in ["discipline=", "search=", "category="]):
        return False
    return bool(re.search(r"/(?:staff|people|person|team|researchers?|experts?|faculty|profile)/[^/]+/?$", path))


def adapter_university_profiles(soup: BeautifulSoup, source_url: str, journal: str) -> List[PersonRecord]:
    records = []
    for block in soup.select("article, .person, .person-card, .staff-member, .staff-item, .faculty-member, .team-member, .researcher, .expert, .card, tr"):
        text = clean_text(block.get_text(" ", strip=True))
        if not text or len(text) > 2200:
            continue
        emails, alts, conflict = extract_email_candidates(block)
        profile = ""
        for a in block.select("a[href]"):
            href = normalize_url(urljoin(source_url, a.get("href", "")))
            if looks_like_profile_url(href):
                profile = href
                break
        if not emails and not profile:
            continue
        if emails:
            rec = make_record_from_block(block, source_url, "UNIVERSITY_DIRECTORY", journal, "university_card")
            if rec:
                rec.profile_url = profile
                records.append(rec)
        else:
            names = collect_name_candidates(block)
            if names:
                _, name = names[0]
                rec = PersonRecord(name=name, page_type="UNIVERSITY_DIRECTORY", source_url=source_url,
                                   profile_url=profile, extraction_method="university_profile_link", confidence=55)
                records.append(rec)
    return records


# =============================================================================
# PROFILE FETCH / ENRICHMENT
# =============================================================================

async def fetch_profile_html(context, url: str) -> Optional[str]:
    r = await fetch_http(url)
    if r["ok"]:
        return r["html"]
    r = await playwright_fetch(context, url)
    return r["html"] if r["ok"] else None


def extract_profile_record(html: str, url: str, name_hint: str, page_type: str) -> Optional[PersonRecord]:
    soup = BeautifulSoup(html, "html.parser")
    normalize_dom(soup)
    root = soup.select_one("main, article, #main-content, #content") or soup.body or soup
    text = clean_text(root.get_text(" ", strip=True))

    emails, alts, conflict = extract_email_candidates(root)
    if not emails:
        emails = extract_text_emails(text)
    if not emails:
        return None

    names = collect_name_candidates(root, emails[0])
    if name_hint:
        names.append((100, strip_credentials(name_hint)))
        names.sort(reverse=True)
    if not names:
        return None
    _, name = names[0]
    if name_hint and normalize_name_key(name_hint) and normalize_name_key(name):
        # require overlap in profile enrichment, but allow exact hint fallback
        a = set(normalize_name_key(name_hint).split())
        b = set(normalize_name_key(name).split())
        if a and b and not (a & b):
            name = strip_credentials(name_hint)

    links = extract_links(root, url)
    rec = PersonRecord(
        name=name,
        email=emails[0],
        alternate_emails=" | ".join(unique(alts + emails[1:])),
        email_conflict="yes" if conflict else "no",
        email_type=email_type(emails[0]),
        page_type=page_type,
        affiliation=split_affiliation_from_block(root, name, emails + alts, current_role_for(root)),
        country=extract_country_from_text(text),
        country_source="profile_text",
        phone=extract_phone(root),
        orcid=links["orcid"] or extract_orcid_text(text),
        google_scholar=links["google_scholar"],
        scopus_author_id=extract_scopus(text),
        pubmed=links["pubmed"],
        personal_homepage=links["personal_homepage"],
        profile_url=url,
        confidence=92,
        extraction_method="profile",
    )
    return resolve_country(rec, text)


async def enrich_missing_profiles(context, records: List[PersonRecord], sem: asyncio.Semaphore) -> List[PersonRecord]:
    async def one(rec: PersonRecord):
        if not rec.profile_url:
            return rec
        # Only profile-enrich when mandatory fields missing or for university pages
        if all(getattr(rec, f) for f in MANDATORY_FIELDS) and rec.page_type == "EDITORIAL_BOARD":
            return rec
        async with sem:
            try:
                html = await fetch_profile_html(context, rec.profile_url)
                if not html:
                    return rec
                p = extract_profile_record(html, rec.profile_url, rec.name, rec.page_type)
                if not p:
                    return rec
                for field_name in COLUMNS:
                    if field_name in ["employee_name", "employee_email", "source_url", "page_type", "journal_name"]:
                        continue
                    old = getattr(rec, field_name, "") if hasattr(rec, field_name) else ""
                    new = getattr(p, field_name, "") if hasattr(p, field_name) else ""
                    if not old and new:
                        setattr(rec, field_name, new)
                rec.confidence = max(rec.confidence, p.confidence)
                rec.extraction_method = f"{rec.extraction_method}+profile"
                return rec
            except Exception:
                return rec

    return await asyncio.gather(*(one(r) for r in records))


# =============================================================================
# OJS JAVASCRIPT PROFILE DISCOVERY
# =============================================================================

def discover_js_profile_placeholders(soup: BeautifulSoup, source_url: str, page_type: str, journal: str) -> List[PersonRecord]:
    records = []
    current_role = ""
    for group in soup.select("#group, .group"):
        h = group.find(["h2", "h3", "h4", "h5"])
        if h:
            current_role = clean_text(h.get_text(" ", strip=True))
        for a in group.select("a[href*='openRTWindow']"):
            name = strip_credentials(a.get_text(" ", strip=True))
            url = extract_js_profile_url(a.get("href", ""), source_url)
            li = a.find_parent("li") or a.parent
            text = clean_text(li.get_text(" ", strip=True)) if li else name
            country = extract_country_from_text(text)
            if plausible_name(name) and url:
                rec = PersonRecord(
                    name=name,
                    country=country,
                    country_source="listing" if country else "",
                    page_type=page_type,
                    journal_name=journal,
                    editorial_role=current_role,
                    affiliation=text.replace(name, "", 1).strip(" ,"),
                    profile_url=url,
                    source_url=source_url,
                    confidence=65,
                    extraction_method="javascript_profile_listing",
                )
                records.append(rec)
    return records


# =============================================================================
# PAGE PARSE ORCHESTRATOR
# =============================================================================

def dedupe_records(records: List[PersonRecord]) -> List[PersonRecord]:
    merged: Dict[Tuple, PersonRecord] = {}
    for r in records:
        email = normalize_email(r.email)
        if email:
            key = ("email", email)
        elif r.profile_url:
            key = ("profile", normalize_url(r.profile_url).lower())
        else:
            key = ("name", normalize_name_key(r.name), urlparse(r.source_url).netloc.lower())

        if key not in merged:
            merged[key] = r
            continue
        cur = merged[key]
        for field_name in COLUMNS:
            if not hasattr(cur, field_name):
                continue
            old = getattr(cur, field_name)
            new = getattr(r, field_name)
            if field_name == "confidence":
                cur.confidence = max(cur.confidence, r.confidence)
            elif field_name == "alternate_emails":
                vals = []
                for part in [old, new]:
                    vals.extend([x.strip() for x in str(part or "").split("|") if x.strip()])
                setattr(cur, field_name, " | ".join(unique(vals)))
            elif not old and new:
                setattr(cur, field_name, new)
    return list(merged.values())


async def parse_html(html: str, source_url: str, llm_enabled: bool, llm_sem: asyncio.Semaphore) -> Tuple[str, List[PersonRecord]]:
    soup = BeautifulSoup(html, "html.parser")
    page_type = detect_page_type(soup, source_url)
    journal = journal_name(soup) if page_type == "EDITORIAL_BOARD" else ""
    normalize_dom(soup)

    records: List[PersonRecord] = []

    if page_type == "EDITORIAL_BOARD":
        # Most structured -> least structured
        records += adapter_cards(soup, source_url, page_type, journal)
        records += adapter_table_multirow(soup, source_url, page_type, journal)
        records += adapter_table_rows(soup, source_url, page_type, journal)
        records += adapter_list_items(soup, source_url, page_type, journal)
        records += adapter_paragraphs(soup, source_url, page_type, journal)
        records += adapter_sequential_flow(soup, source_url, page_type, journal)
        records += discover_js_profile_placeholders(soup, source_url, page_type, journal)
        records += adapter_email_anchors(soup, source_url, page_type, journal)
    else:
        records += adapter_university_profiles(soup, source_url, journal)
        records += adapter_cards(soup, source_url, page_type, journal)
        records += adapter_table_rows(soup, source_url, page_type, journal)
        records += adapter_list_items(soup, source_url, page_type, journal)
        records += adapter_email_anchors(soup, source_url, page_type, journal)

    # Ensure source URL / journal / page type filled
    for r in records:
        r.source_url = r.source_url or source_url
        r.page_type = r.page_type or page_type
        r.journal_name = r.journal_name or journal
        r = resolve_country(r, r.affiliation)

    records = dedupe_records(records)

    # LLM only fixes ambiguous/missing name among existing local candidates is handled sparingly here.
    # We do NOT let LLM invent emails or names.
    if llm_enabled:
        for r in records:
            if r.name and plausible_name(r.name):
                continue
            # no block retained here, so skip unsafe global LLM lookup

    return page_type, records


# =============================================================================
# COUNTRY LLM LAST RESORT
# =============================================================================

async def fill_missing_countries(records: List[PersonRecord], llm_enabled: bool, llm_sem: asyncio.Semaphore):
    # Country is optional. Do not spend Qwen calls guessing it.
    # resolve_country() already handles explicit/local/TLD evidence.
    return records




# =============================================================================
# QWEN3 AMBIGUOUS-ONLY RECORD VALIDATOR
# =============================================================================
# Designed for CPU-only / 16 GB RAM machines:
#   * qwen3:4b
#   * one LLM request at a time
#   * deterministic rules accept strong records directly
#   * Qwen is used only when identity mapping is ambiguous
#   * if Ollama/Qwen fails, records are NOT blindly deleted

LLM_VALIDATE_ALL_RECORDS = False
LLM_ONLY_AMBIGUOUS = True
LLM_FAIL_OPEN = True

LLM_ACCEPT_THRESHOLD = 90
LLM_AMBIGUOUS_THRESHOLD = 72
LLM_STRONG_THRESHOLD = 88

GENERIC_EMAIL_PREFIXES_STRICT = (
    "helpdesk@", "info@", "contact@", "admin@", "office@", "media@", "hr@",
    "support@", "feedback@", "faculty@", "department@", "dept@", "webmaster@",
    "communications@", "communication@", "admissions@", "admission@",
    "dentalfacultypractice@", "asod_pbs@", "resident@", "residents@",
    "team@", "research@", "researchteam@", "research-team@",
)

BAD_NAME_LABELS_STRICT = {
    "view full profile", "view profile", "read bio", "read biography",
    "personal website", "personal webpage", "website", "homepage",
    "appt", "appointment", "phone", "fax", "people", "team", "students",
    "alumni", "faculty", "faculty members", "clinical faculty",
    "research interests", "education", "contact", "contact information",
    "employment opportunities", "rti quarterly returns report",
    "adult autism spectrum cohort project lead", "asd family support manager",
    "adult autism study qualitative researcher", "thanks to this experience",
}


def is_strict_generic_email(email: str) -> bool:
    e = normalize_email(email)
    if not e:
        return True
    low = e.lower()
    return any(low.startswith(prefix) for prefix in GENERIC_EMAIL_PREFIXES_STRICT)


def is_bad_name_label_strict(name: str) -> bool:
    n = clean_text(name).strip()
    if not n:
        return False
    low = n.casefold().strip(" :;,.|-")
    if low in BAD_NAME_LABELS_STRICT:
        return True
    if low.startswith(("appt ", "appointment ", "phone ", "fax ")):
        return True
    if re.fullmatch(r"(?:\+?\d[\d\s().xX-]{6,})", n):
        return True
    if re.search(r"\b(?:19|20)\d{2}\s*[-–—]\s*(?:present|(?:19|20)\d{2})\b", low):
        return True
    return False


def _safe_llm_text(value: str, limit: int = LLM_MAX_TEXT) -> str:
    return clean_text(value)[:limit]


def _strict_json_from_model(text: str) -> Optional[Dict]:
    data = extract_json_object(text)
    return data if isinstance(data, dict) else None


def deterministic_identity_score(rec: PersonRecord) -> Tuple[int, List[str]]:
    """
    Score how strongly the current DOM/profile evidence supports the mapping.

    Strong profile + human name + personal email can skip Qwen.
    Ambiguous rows are sent to Qwen.
    Obvious junk is rejected without spending LLM time.
    """
    score = 0
    reasons: List[str] = []

    email = normalize_email(rec.email)
    name = clean_person_name_candidate(rec.name)
    method = (rec.extraction_method or "").lower()
    profile = normalize_url(rec.profile_url)

    if not email:
        return 0, ["missing_email"]

    if is_strict_generic_email(email) or is_generic_email(email):
        score -= 100
        reasons.append("generic_email")
    else:
        score += 35
        reasons.append("personal_email")

    if name and plausible_name(name) and not is_bad_name_label_strict(name):
        score += 25
        reasons.append("credible_name")
    elif rec.name:
        score -= 35
        reasons.append("bad_name")
    else:
        reasons.append("blank_name")

    match = email_name_score(name, email) if name else 0
    if match >= 40:
        score += 30
        reasons.append("strong_name_email_match")
    elif match >= 20:
        score += 18
        reasons.append("some_name_email_match")

    if profile:
        score += 18
        reasons.append("profile_url")

    if "profile_v7" in method or method.startswith("profile"):
        score += 22
        reasons.append("profile_enriched")
    elif "university_profile" in method:
        score += 15
        reasons.append("profile_discovery")
    elif "university_local_email_v7" in method:
        score += 16
        reasons.append("local_email_block")
    elif "table_row" in method or "table_cell" in method or "card" in method:
        score += 10
        reasons.append("structured_block")
    elif "plain_email" in method or "email_anchor" in method:
        score -= 5
        reasons.append("weak_generic_extractor")

    # Exact strong same-domain profile patterns.
    if profile and email:
        try:
            p_host = urlparse(profile).netloc.casefold().replace("www.", "")
            e_host = email.rsplit("@", 1)[1].casefold()
            if p_host and (
                e_host.endswith(p_host)
                or p_host.endswith(e_host)
                or e_host.split(".")[-2:] == p_host.split(".")[-2:]
            ):
                score += 8
                reasons.append("profile_email_domain_related")
        except Exception:
            pass

    # Shared team/project mailbox attached to multiple named people is unsafe.
    local = email.split("@", 1)[0].casefold()
    if any(x in local for x in ("autismspectrum", "researchteam", "research-team")):
        score -= 80
        reasons.append("project_shared_mailbox")

    return max(0, min(100, score)), reasons


def llm_validate_record_sync(
    record: PersonRecord,
    local_text: str = "",
    candidate_names: Optional[List[str]] = None,
    candidate_emails: Optional[List[str]] = None,
) -> Dict:
    candidate_names = unique(candidate_names or ([record.name] if record.name else []))
    candidate_emails = unique(candidate_emails or ([record.email] if record.email else []))

    payload = {
        "source_url": record.source_url,
        "profile_url": record.profile_url,
        "candidate_name": record.name,
        "candidate_email": record.email,
        "candidate_country": record.country,
        "candidate_affiliation": _safe_llm_text(record.affiliation, 900),
        "candidate_names": candidate_names[:12],
        "candidate_emails": candidate_emails[:12],
        "local_person_block": _safe_llm_text(local_text or record.affiliation, LLM_MAX_TEXT),
    }

    system_prompt = """
You validate scientific/university directory identity mappings.

ACCURACY IS MORE IMPORTANT THAN RECALL.

You MUST NOT invent a name or email.
You may only validate the supplied candidate_email and select/clean a name
already present in candidate_names/candidate_name.

Return JSON only:
{
  "is_person": true,
  "clean_name": "",
  "email_matches_person": true,
  "is_generic_email": false,
  "country": "",
  "confidence": 0,
  "reason": ""
}

Reject if:
- candidate_name is navigation, title-only, phone, "Personal website", category,
  department, team label, or unrelated text
- candidate_email is shared/generic/project-wide
- email appears to belong to a different nearby person
- evidence does not support the exact person/email pairing

If uncertain, confidence must be below 80.
Never manufacture missing data.
""".strip()

    try:
        r = requests.post(
            OLLAMA_URL,
            json={
                "model": OLLAMA_MODEL,
                "stream": False,
                "options": {
                    "temperature": 0,
                    "num_predict": 180,
                },
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
                ],
            },
            timeout=OLLAMA_TIMEOUT,
        )
        r.raise_for_status()

        raw = r.json().get("message", {}).get("content", "")
        data = _strict_json_from_model(raw)
        if not data:
            return {
                "llm_available": False,
                "llm_error": "invalid_llm_json",
            }

        result = {
            "llm_available": True,
            "is_person": bool(data.get("is_person", False)),
            "email_matches_person": bool(data.get("email_matches_person", False)),
            "is_generic_email": bool(data.get("is_generic_email", False)),
            "clean_name": "",
            "country": normalize_country(data.get("country", "")),
            "confidence": 0,
            "reason": clean_text(data.get("reason", ""))[:400],
        }

        try:
            result["confidence"] = max(0, min(100, int(data.get("confidence", 0) or 0)))
        except Exception:
            result["confidence"] = 0

        proposed_name = clean_person_name_candidate(data.get("clean_name", ""))
        allowed = {}
        for n in candidate_names + ([record.name] if record.name else []):
            cn = clean_person_name_candidate(n)
            if cn:
                allowed[normalize_name_key(cn)] = cn

        if proposed_name:
            key = normalize_name_key(proposed_name)
            if key in allowed and plausible_name(proposed_name):
                result["clean_name"] = proposed_name

        # Deterministic rules override Qwen.
        if is_strict_generic_email(record.email) or is_generic_email(record.email):
            result["is_generic_email"] = True

        return result

    except Exception as exc:
        return {
            "llm_available": False,
            "llm_error": clean_text(str(exc))[:300],
        }


async def llm_validate_record(
    semaphore: asyncio.Semaphore,
    record: PersonRecord,
    local_text: str = "",
    candidate_names: Optional[List[str]] = None,
    candidate_emails: Optional[List[str]] = None,
) -> Dict:
    async with semaphore:
        return await asyncio.to_thread(
            llm_validate_record_sync,
            record,
            local_text,
            candidate_names,
            candidate_emails,
        )


async def llm_filter_records(
    records: List[PersonRecord],
    llm_enabled: bool,
    llm_sem: asyncio.Semaphore,
    errors: List[Dict],
    source_url: str,
) -> List[PersonRecord]:
    """
    Three-tier identity decision:

      STRONG:
        deterministic score >= LLM_STRONG_THRESHOLD
        -> accept without Qwen

      AMBIGUOUS:
        score >= LLM_AMBIGUOUS_THRESHOLD
        -> Qwen validation when available
        -> if Qwen fails, keep only if deterministic score >= 82

      WEAK/JUNK:
        score < LLM_AMBIGUOUS_THRESHOLD
        -> reject without Qwen

    This prevents a stopped Ollama service from producing an empty workbook.
    """
    output: List[PersonRecord] = []

    async def check_one(rec: PersonRecord):
        score, reasons = deterministic_identity_score(rec)
        rec.confidence = max(int(rec.confidence or 0), score)

        # Immediate hard rejects.
        if not normalize_email(rec.email):
            return None, {
                "type": "deterministic_rejected_record",
                "reason": "missing_or_invalid_email",
                "score": score,
            }

        if is_strict_generic_email(rec.email) or is_generic_email(rec.email):
            return None, {
                "type": "deterministic_rejected_record",
                "reason": "generic_or_shared_email",
                "score": score,
            }

        if is_bad_name_label_strict(rec.name):
            # Keep name blank only when the rest of the record is very strong.
            if score >= LLM_STRONG_THRESHOLD and rec.profile_url:
                rec.name = ""
            else:
                return None, {
                    "type": "deterministic_rejected_record",
                    "reason": "non_person_name_label",
                    "score": score,
                }

        # Strong deterministic mapping: do not waste LLM.
        if score >= LLM_STRONG_THRESHOLD:
            rec.extraction_method = (
                f"{rec.extraction_method}+deterministic_strong"
                if rec.extraction_method else "deterministic_strong"
            )
            return rec, None

        # Too weak to justify an LLM call.
        if score < LLM_AMBIGUOUS_THRESHOLD:
            return None, {
                "type": "deterministic_rejected_record",
                "reason": "weak_identity_mapping",
                "score": score,
                "evidence": reasons,
            }

        # Ambiguous record: use Qwen if Ollama is available.
        if not llm_enabled:
            if LLM_FAIL_OPEN and score >= 82:
                rec.extraction_method = (
                    f"{rec.extraction_method}+deterministic_fallback"
                    if rec.extraction_method else "deterministic_fallback"
                )
                return rec, None
            return None, {
                "type": "deterministic_rejected_record",
                "reason": "ambiguous_and_llm_unavailable",
                "score": score,
            }

        result = await llm_validate_record(
            llm_sem,
            rec,
            local_text=rec.affiliation,
            candidate_names=[rec.name] if rec.name else [],
            candidate_emails=[rec.email] if rec.email else [],
        )

        if not result.get("llm_available", False):
            # Important: model failure is not an automatic record rejection.
            if LLM_FAIL_OPEN and score >= 82:
                rec.extraction_method = (
                    f"{rec.extraction_method}+qwen_unavailable_fallback"
                    if rec.extraction_method else "qwen_unavailable_fallback"
                )
                return rec, {
                    "type": "llm_warning",
                    "reason": "qwen_unavailable_used_deterministic_fallback",
                    "score": score,
                    "llm_error": result.get("llm_error", ""),
                }
            return None, {
                "type": "llm_warning_rejected",
                "reason": "qwen_unavailable_and_not_strong_enough",
                "score": score,
                "llm_error": result.get("llm_error", ""),
            }

        accepted = (
            result.get("is_person") is True
            and result.get("email_matches_person") is True
            and result.get("is_generic_email") is False
            and int(result.get("confidence", 0) or 0) >= LLM_ACCEPT_THRESHOLD
        )

        if not accepted:
            return None, {
                "type": "qwen_rejected_record",
                "reason": result.get("reason", ""),
                "score": score,
                "llm_confidence": int(result.get("confidence", 0) or 0),
            }

        clean_name = clean_person_name_candidate(result.get("clean_name", ""))
        if clean_name and plausible_name(clean_name):
            rec.name = clean_name

        if not rec.country:
            c = normalize_country(result.get("country", ""))
            if c:
                rec.country = c
                rec.country_source = "qwen3_local_validation"

        rec.confidence = max(
            int(rec.confidence or 0),
            int(result.get("confidence", 0) or 0),
        )
        rec.extraction_method = (
            f"{rec.extraction_method}+qwen3_validation"
            if rec.extraction_method else "qwen3_validation"
        )
        return rec, None

    checked = await asyncio.gather(*(check_one(r) for r in records))

    for rec, diagnostic in checked:
        if diagnostic:
            item = {
                "source_url": source_url,
                "name": excel_safe(getattr(rec, "name", "") if rec else ""),
                "email": excel_safe(getattr(rec, "email", "") if rec else ""),
            }
            item.update(diagnostic)
            errors.append(item)
        if rec is not None:
            output.append(rec)

    return dedupe_records(output)

# =============================================================================
# FINAL VALIDATION
# =============================================================================

def validate_record(r: PersonRecord) -> Tuple[bool, str]:
    # EMAIL is the ONLY mandatory identity field.
    r.email = normalize_email(r.email)
    if not r.email:
        return False, "missing_or_invalid_email"

    # NAME is optional, but it must never contain affiliation/navigation/category
    # garbage. If the scraped candidate is unsafe, discard it and try a
    # conservative email-derived display name. Blank is acceptable.
    scraped_name = clean_person_name_candidate(r.name)
    if scraped_name and plausible_name(scraped_name):
        r.name = scraped_name
    else:
        r.name = name_from_email(r.email)

    # COUNTRY is optional. Keep blank if it cannot be resolved reliably.
    country = normalize_country(r.country)
    r.country = country if country else ""

    return True, ""


# =============================================================================
# INPUT / OUTPUT
# =============================================================================

def read_urls() -> List[str]:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE.resolve()}")
    wb = load_workbook(INPUT_FILE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = clean_text(ws.cell(1, 1).value).lower()
    if header != "url":
        wb.close()
        raise RuntimeError("Cell A1 must be exactly: url")

    # Strict: reject other populated columns
    for col in range(2, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            if clean_text(ws.cell(row, col).value):
                wb.close()
                raise RuntimeError("urls.xlsx must contain only one populated column: url")

    urls = []
    for row in range(2, ws.max_row + 1):
        value = clean_text(ws.cell(row, 1).value)
        if not value:
            continue
        m = re.search(r"https?://[^\s)]+", value)
        if m:
            url = m.group(0).rstrip(".,;)>]}")
            if url not in urls:
                urls.append(url)
    wb.close()
    return urls


def save_excel(records: List[PersonRecord]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"
    ws.append(COLUMNS)
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in records:
        data = asdict(r)
        ws.append([data.get(c, "") for c in COLUMNS])

    widths = {
        "A": 42, "B": 42, "C": 22, "D": 55, "E": 18, "F": 14,
        "G": 18, "H": 38, "I": 24, "J": 48, "K": 32, "L": 32, "M": 26,
        "N": 28, "O": 75, "P": 50, "Q": 45, "R": 45, "S": 50, "T": 50,
        "U": 42, "V": 24, "W": 65, "X": 24, "Y": 65, "Z": 65, "AA": 24,
        "AB": 24, "AC": 65, "AD": 65, "AE": 85, "AF": 25, "AG": 14, "AH": 28, "AI": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(XLSX_FILE)


def save_errors(errors: List[Dict]):
    tmp = ERROR_FILE.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(errors, f, ensure_ascii=False, indent=2)
    tmp.replace(ERROR_FILE)


# =============================================================================
# PAGINATION
# =============================================================================

def next_page_url(html: str, current_url: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for sel in ['a[rel="next"]', '.pagination-next a', '.pager-next a', 'a.next']:
        a = soup.select_one(sel)
        if a and a.get("href"):
            return normalize_url(urljoin(current_url, a.get("href")))
    for a in soup.select("a[href]"):
        t = clean_text(a.get_text(" ", strip=True)).lower()
        if t in {"next", "next page", ">", "›", "»", "weiter", "suivant"}:
            return normalize_url(urljoin(current_url, a.get("href")))
    return ""


# =============================================================================
# SOURCE SCRAPE
# =============================================================================

async def scrape_source(context, source_url: str, llm_enabled: bool, llm_sem: asyncio.Semaphore,
                        profile_sem: asyncio.Semaphore, errors: List[Dict]) -> List[PersonRecord]:
    all_records = []
    current_url = source_url
    visited = set()

    for page_num in range(1, MAX_PAGES_PER_URL + 1):
        if not current_url or current_url in visited:
            break
        visited.add(current_url)
        print(f"\n   [PAGE {page_num}] {current_url}", flush=True)
        print("   [FETCH] HTTP first...", flush=True)
        r = await fetch_http(current_url)
        method = "http"
        if not r["ok"]:
            print(f"   [HTTP FAILED] {r['error']}", flush=True)
            print("   [FALLBACK] Playwright...", flush=True)
            r = await playwright_fetch(context, current_url)
            method = "playwright"
        else:
            print(f"   [HTTP OK] {len(r['html']):,} chars", flush=True)

        if not r["ok"]:
            errors.append({"type": "fetch_failed", "source_url": source_url, "page_url": current_url, "error": r["error"]})
            break

        page_type, records = await parse_html(r["html"], source_url, llm_enabled, llm_sem)
        print(f"   [PARSE] type={page_type} raw_records={len(records)} method={method}", flush=True)

        # Profile enrichment helps OJS JS bios and university directory profile pages
        records = await enrich_missing_profiles(context, records, profile_sem)
        records = await fill_missing_countries(records, llm_enabled, llm_sem)

        for rec in records:
            valid, reason = validate_record(rec)
            if valid:
                print(f"      {rec.name[:38]:38} | {rec.email[:36]:36} | {rec.country[:18]:18} | {rec.editorial_role[:24]}", flush=True)
                all_records.append(rec)
            else:
                errors.append({
                    "type": "rejected_record",
                    "reason": reason,
                    "name": rec.name,
                    "email": rec.email,
                    "country": rec.country,
                    "profile_url": rec.profile_url,
                    "source_url": source_url,
                    "method": rec.extraction_method,
                })

        if not FOLLOW_PAGINATION:
            break
        nxt = next_page_url(r["html"], r["url"])
        if not nxt or nxt in visited:
            break
            # same-site only
        if urlparse(nxt).netloc.lower().replace("www.", "") != urlparse(source_url).netloc.lower().replace("www.", ""):
            break
        current_url = nxt
        await polite_delay()

    return dedupe_records(all_records)


# =============================================================================
# MAIN
# =============================================================================

async def main():
    urls = read_urls()
    if not urls:
        raise RuntimeError("No URLs found in urls.xlsx")

    llm_enabled = await ollama_available()
    print("=" * 110)
    print("UNIVERSAL UNIVERSITY + JOURNAL EDITORIAL BOARD SCRAPER")
    print("=" * 110)
    print(f"Employee : {EMPLOYEE_NAME}")
    print(f"Email    : {EMPLOYEE_EMAIL}")
    print(f"URLs     : {len(urls)}")
    print(f"Ollama   : {'ENABLED' if llm_enabled else 'DISABLED'}")
    print(f"LLM Mode : ambiguous-only validation; deterministic fail-safe")
    if llm_enabled:
        print(f"Model    : {OLLAMA_MODEL}")
    print(f"Output   : {XLSX_FILE.resolve()}")

    errors: List[Dict] = []
    master: List[PersonRecord] = []
    llm_sem = asyncio.Semaphore(LLM_CONCURRENCY)
    profile_sem = asyncio.Semaphore(PROFILE_CONCURRENCY)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context(
            viewport={"width": 1440, "height": 1000},
            locale="en-US",
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0 Safari/537.36",
        )

        for idx, url in enumerate(urls, start=1):
            print("\n" + "#" * 110)
            print(f"URL {idx}/{len(urls)}")
            print(url)
            print("#" * 110)
            try:
                records = await scrape_source(context, url, llm_enabled, llm_sem, profile_sem, errors)
                master.extend(records)
                master = dedupe_records(master)
                save_excel(master)
                save_errors(errors)
                print(f"   [SOURCE DONE] accepted={len(records)} master={len(master)}", flush=True)
            except Exception as exc:
                errors.append({"type": "source_exception", "source_url": url, "error": str(exc)})
                save_errors(errors)
            await polite_delay()

        await browser.close()

    # Final clean validation + sort
    final_records = []
    for rec in dedupe_records(master):
        ok, reason = validate_record(rec)
        if ok:
            final_records.append(rec)
        else:
            errors.append({"type": "final_rejected", "reason": reason, "name": rec.name, "email": rec.email,
                           "country": rec.country, "source_url": rec.source_url})

    final_records.sort(key=lambda r: (r.page_type, r.journal_name.lower(), r.country.lower(), r.name.lower()))
    save_excel(final_records)
    save_errors(errors)

    print("\n" + "=" * 110)
    print("SCRAPING COMPLETE")
    print("=" * 110)
    print(f"URLs processed : {len(urls)}")
    print(f"Final records  : {len(final_records)}")
    print(f"Errors/rejects : {len(errors)}")
    print(f"Excel          : {XLSX_FILE.resolve()}")
    print(f"Errors         : {ERROR_FILE.resolve()}")



# =============================================================================
# V7 PRODUCTION-SAFETY OVERRIDES
# =============================================================================
# These definitions intentionally override the earlier generic implementations.
# Core policy:
#   * email = mandatory
#   * name = optional, but never guess a compact username as a person name
#   * country = optional
#   * a person's email/name must come from the same local card/profile
#   * generic site mailboxes are rejected unless explicitly attached to a person

import unicodedata

try:
    from email_validator import validate_email as _validate_email_address
except Exception:
    _validate_email_address = None


ILLEGAL_EXCEL_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]")
FAKE_TLDS = {
    "the", "with", "from", "this", "that", "where", "when", "page", "html",
    "http", "https", "email", "mailto", "none", "null", "undefined"
}
GENERIC_LOCALPARTS = {
    "info", "contact", "contacts", "office", "editor", "editorial", "journal",
    "secretary", "secretaria", "admin", "support", "webmaster", "submission",
    "submissions", "help", "mail", "email", "enquiries", "inquiries",
    "business", "manager", "reception", "communications", "communication",
    "media", "press", "news", "employment", "jobs", "careers", "admissions",
    "admission", "parking", "health", "web", "website", "general", "hello",
    "team", "faculty", "department", "dept", "school", "college", "center",
    "centre", "program", "programme", "events", "event", "equity", "privacy"
}
BAD_NAME_PHRASES_V7 = {
    "view profile", "read bio", "read biography", "contact info", "contact information",
    "research interests", "education", "footer navigation", "links of interest",
    "strategic engagement", "workforce development", "employment opportunities",
    "view information for parking", "diversity, equity", "administrative contact",
    "faculty members", "clinical faculty", "primary faculty", "secondary faculty",
    "faculty and staff", "faculty & staff", "our faculty", "all faculty",
    "click here", "learn more", "more information", "email address", "e-mail address",
}

PROFILE_PATH_HINTS_V7 = (
    "/people/", "/person/", "/faculty/", "/faculty-profiles/", "/profile/",
    "/profiles/", "/staff/", "/team/", "/researcher/", "/researchers/",
    "/investigator/", "/investigators/", "/directory/", "/expert/", "/experts/",
    "/pi/", "/bios/", "/bio/"
)

AUTHORITATIVE_PROFILE_HOST_HINTS_V7 = (
    "profiles.stanford.edu",
    "med.stanford.edu",
    "directory.stanford.edu",
    "directory.campbell.edu",
    "profiles.umich.edu",
    "faculty.wustl.edu",
    "profiles.wustl.edu",
    "directory.wustl.edu",
    "profiles.case.edu",
    "directory.case.edu",
    "irp.nih.gov",
)


def excel_safe(value) -> str:
    if value is None:
        return ""
    s = str(value)
    s = ILLEGAL_EXCEL_RE.sub("", s)
    # Excel maximum cell text length
    if len(s) > 32767:
        s = s[:32767]
    return s


def clean_text(value) -> str:
    if value is None:
        return ""
    value = excel_safe(value)
    value = (
        value.replace("\xa0", " ")
        .replace("\u200b", "")
        .replace("\ufeff", "")
        .replace("\u00ad", "")
        .replace("\r", " ")
        .replace("\n", " ")
        .replace("\t", " ")
    )
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_email(value: str) -> str:
    if not value:
        return ""
    value = decode_email_obfuscation(value)
    m = EMAIL_RE.search(value)
    if not m:
        return ""
    email = m.group(0).lower().strip(".,;:/()[]{}<>")
    if email.count("@") != 1:
        return ""
    local, domain = email.rsplit("@", 1)
    if not local or not domain or "." not in domain:
        return ""
    if ".." in local or ".." in domain:
        return ""
    if domain.startswith("-") or domain.endswith("-"):
        return ""
    labels = domain.split(".")
    if any(not x or x.startswith("-") or x.endswith("-") for x in labels):
        return ""
    tld = labels[-1].lower()
    if tld in FAKE_TLDS or len(tld) < 2 or not re.fullmatch(r"[a-z]{2,24}", tld):
        return ""

    if _validate_email_address is not None:
        try:
            # Do not make DNS/network checks during scraping.
            info = _validate_email_address(email, check_deliverability=False)
            email = info.normalized.lower()
        except Exception:
            return ""
    return email


def is_generic_email(email: str) -> bool:
    email = normalize_email(email)
    if not email:
        return True
    local = email.split("@", 1)[0].lower()
    root = re.split(r"[._+\-]", local)[0]
    if local in GENERIC_LOCALPARTS or root in GENERIC_LOCALPARTS:
        return True
    return any(
        local.startswith(prefix)
        for prefix in (
            "info.", "info_", "contact.", "contact_", "office.", "office_",
            "admin.", "admin_", "media.", "media_", "health.", "health_",
            "employment.", "employment_", "communications.", "communications_",
        )
    )


def _has_unicode_letter(value: str) -> bool:
    return any(unicodedata.category(ch).startswith("L") for ch in value)


def _unicode_words(value: str) -> List[str]:
    words = []
    current = []
    for ch in clean_text(value):
        if unicodedata.category(ch).startswith("L") or ch in "'’.-":
            current.append(ch)
        else:
            if current:
                token = "".join(current).strip(".'’-")
                if token:
                    words.append(token)
                current = []
    if current:
        token = "".join(current).strip(".'’-")
        if token:
            words.append(token)
    return words


def plausible_name(text: str) -> bool:
    raw = clean_text(text)
    if not raw or len(raw) < 2 or len(raw) > 120:
        return False
    if looks_like_url(raw) or "@" in raw:
        return False

    low = raw.casefold().strip(" :;,.|")
    if low in {x.casefold() for x in BAD_NAME_EXACT}:
        return False
    if normalize_country(raw):
        return False
    if is_role_heading(raw):
        return False
    if any(p in low for p in BAD_NAME_PHRASES_V7):
        return False
    if re.search(r"\b(?:19|20)\d{2}\s*[-–—]\s*(?:present|(?:19|20)\d{2})", low, re.I):
        return False
    if "/" in raw and len(_unicode_words(raw)) <= 4:
        return False
    if re.fullmatch(r"[\d\W_]+", raw):
        return False
    if not _has_unicode_letter(raw):
        return False

    words = _unicode_words(raw)
    if not words or len(words) > 10:
        return False

    # Reject obvious organisation/department phrases.
    org_hits = sum(1 for w in ORG_WORDS if w.casefold() in low)
    if org_hits:
        return False

    # Labels that happen to be alphabetic.
    if any(
        x in low
        for x in (
            "faculty", "department", "division", "program", "programme", "research center",
            "research centre", "hospital", "clinic", "university", "school of", "college of",
            "contact", "navigation", "copyright", "privacy", "policy", "opportunities",
        )
    ):
        return False

    return True


def normalize_name_key(name: str) -> str:
    name = clean_person_name_candidate(name).casefold()
    drop = {
        "prof", "professor", "dr", "md", "phd", "msc", "mph", "mbbs", "frcog",
        "facog", "mba", "bsc", "ma"
    }
    tokens = []
    for token in _unicode_words(name):
        key = token.casefold().strip(".'’-")
        if len(key) > 1 and key not in drop:
            tokens.append(key)
    return " ".join(sorted(tokens))


def name_from_email(email: str) -> str:
    """
    Extremely conservative fallback.

    Only addresses with explicit token separators can create a name:
        john.smith@x.edu -> John Smith
        john_smith@x.edu -> John Smith
        john-smith@x.edu -> John Smith

    Compact institutional usernames such as dbarbour, joannaa, kmbennett,
    danielgridley etc. return blank.
    """
    email = normalize_email(email)
    if not email or is_generic_email(email):
        return ""
    local = email.split("@", 1)[0].split("+", 1)[0].strip().lower()

    if not re.search(r"[._-]", local):
        return ""

    raw_tokens = [x for x in re.split(r"[._-]+", local) if x]
    if len(raw_tokens) < 2:
        return ""

    cleaned = []
    for token in raw_tokens:
        token = re.sub(r"\d+$", "", token)
        token = "".join(ch for ch in token if unicodedata.category(ch).startswith("L"))
        if not token:
            continue
        if token.casefold() in GENERIC_LOCALPARTS:
            return ""
        if len(token) == 1:
            cleaned.append(token.upper() + ".")
        else:
            cleaned.append(token[:1].upper() + token[1:])

    if len(cleaned) < 2:
        return ""

    candidate = " ".join(cleaned)
    return candidate if plausible_name(candidate) else ""


def clean_person_name_candidate(text: str) -> str:
    text = strip_credentials(text)
    text = clean_text(text)
    if not text:
        return ""

    low = text.casefold()
    if any(p in low for p in BAD_NAME_PHRASES_V7):
        return ""

    # Remove role suffixes joined with vertical bar.
    if "|" in text:
        pieces = [clean_text(x) for x in text.split("|") if clean_text(x)]
        if pieces and plausible_name(pieces[0]):
            text = pieces[0]

    # Strip clear affiliation after dash.
    for sep in (" – ", " — ", " - "):
        if sep in text:
            head, tail = text.split(sep, 1)
            tail_low = tail.casefold()
            if (
                any(w.casefold() in tail_low for w in ORG_WORDS)
                or bool(extract_country_from_text(tail))
                or re.search(
                    r"\b(?:faculty|universidade|universidad|universität|universite|"
                    r"hospital|department|institute|centre|center|school|college)\b",
                    tail_low,
                    re.I,
                )
            ):
                text = head.strip(" ,;:-")
                break

    # Strip affiliation after comma while preserving surname-first names.
    parts = [clean_text(x) for x in text.split(",")]
    if len(parts) >= 2:
        for i in range(1, len(parts)):
            tail = ", ".join(parts[i:])
            tl = tail.casefold()
            if (
                any(w.casefold() in tl for w in ORG_WORDS)
                or bool(extract_country_from_text(tail))
                or re.search(
                    r"\b(?:university|universidade|universidad|universität|hospital|"
                    r"faculty|department|institute|centre|center|clinic|school|college)\b",
                    tl,
                    re.I,
                )
            ):
                head = ", ".join(parts[:i]).strip(" ,;:-")
                if plausible_name(head):
                    text = head
                break

    text = strip_credentials(text).strip(" ,;:-|")
    return clean_text(text)


def collect_name_candidates(block: Tag, email: str = "") -> List[Tuple[int, str]]:
    """
    Keep the ORIGINAL spelling/case. The previous implementation stored the
    lowercase dedupe key as the returned name.
    """
    candidates: List[Tuple[int, str]] = []
    selector_scores = [
        ("h1", 115),
        ("h2,h3,h4,h5,h6", 90),
        ('[itemprop="name"]', 105),
        ('[class*="name"]', 95),
        ("strong,b", 78),
        ("a[href]", 62),
        ("cite", 30),
    ]

    for selector, base in selector_scores:
        for el in block.select(selector):
            text = clean_person_name_candidate(el.get_text(" ", strip=True))
            if not plausible_name(text):
                continue
            # Do not treat utility link labels as names.
            if text.casefold() in {"view profile", "read bio", "email", "contact", "learn more"}:
                continue
            score = base + email_name_score(text, email)
            if el.name == "a":
                href = clean_text(el.get("href", ""))
                if href.lower().startswith("mailto:"):
                    score -= 20
                elif looks_like_profile_url(urljoin("https://example.invalid", href)):
                    score += 20
            candidates.append((score, text))

    # Only short, line-local strings; never use the entire card/page as a name.
    raw = block.get_text("\n", strip=True)
    for line in raw.splitlines():
        text = clean_person_name_candidate(line)
        if plausible_name(text) and len(_unicode_words(text)) <= 6:
            candidates.append((50 + email_name_score(text, email), text))

    best: Dict[str, Tuple[int, str]] = {}
    for score, text in candidates:
        key = normalize_name_key(text) or text.casefold()
        old = best.get(key)
        if old is None or score > old[0]:
            best[key] = (score, text)

    return sorted(best.values(), key=lambda x: x[0], reverse=True)


def extract_email_candidates(block: Tag) -> Tuple[List[str], List[str], bool]:
    """
    Keep emails local to this DOM block.

    Important change: multiple unrelated emails discovered in a broad parent are
    NOT automatically treated as alternate emails.
    """
    primaries: List[str] = []
    alternates: List[str] = []
    conflict = False

    for a in block.select('a[href^="mailto:"], a[href^="MAILTO:"]'):
        href_email = normalize_email(a.get("href", ""))
        visible_email = normalize_email(a.get_text(" ", strip=True))
        if visible_email:
            primaries.append(visible_email)
            if href_email and href_email != visible_email:
                alternates.append(href_email)
                conflict = True
        elif href_email:
            primaries.append(href_email)

    # Plain-text email is useful only when the block is already small/local.
    text = clean_text(block.get_text(" ", strip=True))
    if len(text) <= 900:
        for email in extract_text_emails(text):
            if email not in primaries:
                primaries.append(email)

    primaries = unique([x for x in primaries if normalize_email(x)])
    alternates = unique([x for x in alternates if normalize_email(x) and x not in primaries])
    return primaries, alternates, conflict


def _profile_link_score(anchor: Tag, source_url: str) -> Tuple[int, str, str]:
    href = clean_text(anchor.get("href", ""))
    if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
        return 0, "", ""
    url = normalize_url(urljoin(source_url, href))
    if not url.startswith(("http://", "https://")):
        return 0, "", ""

    text = clean_person_name_candidate(anchor.get_text(" ", strip=True))
    if not plausible_name(text):
        # Sometimes name is in a child heading but anchor text contains credentials.
        text = clean_person_name_candidate(anchor.get("aria-label", "") or anchor.get("title", ""))
    if not plausible_name(text):
        return 0, "", ""

    p = urlparse(url)
    path = p.path.casefold()
    score = 0
    if any(h in path for h in PROFILE_PATH_HINTS_V7):
        score += 50
    if re.search(r"/[^/]+-[^/]+/?$", path):
        score += 15
    if p.netloc.casefold().replace("www.", "") == urlparse(source_url).netloc.casefold().replace("www.", ""):
        score += 15
    if any(h in p.netloc.casefold() for h in AUTHORITATIVE_PROFILE_HOST_HINTS_V7):
        score += 35
    return score, url, text


def looks_like_profile_url(url: str) -> bool:
    if not url:
        return False
    p = urlparse(url)
    path = p.path.casefold()
    query = p.query.casefold()

    if any(x in path for x in (
        "/disciplines", "/find-an-expert", "/search", "/category",
        "/editorial-board", "/news/", "/events/", "/event/", "/tag/"
    )):
        return False
    if any(x in query for x in ("discipline=", "search=", "category=")):
        return False

    if any(h in path for h in PROFILE_PATH_HINTS_V7):
        # Require a leaf-ish URL, except known faculty-profile systems.
        tail = [x for x in path.split("/") if x]
        return len(tail) >= 2
    if any(h in p.netloc.casefold() for h in AUTHORITATIVE_PROFILE_HOST_HINTS_V7):
        return True
    return False


def _nearest_person_container(node: Tag, email: str = "", max_chars: int = 1800) -> Tag:
    """
    Walk upward and choose the best container containing at most one real email
    (or two when one is a visible/href conflict) plus a credible human name.
    """
    best = node
    best_score = -999
    cur = node
    for depth in range(7):
        if not isinstance(cur, Tag):
            break
        text = clean_text(cur.get_text(" ", strip=True))
        if len(text) > max_chars:
            break

        emails = unique(extract_text_emails(text) + [
            normalize_email(a.get("href", ""))
            for a in cur.select('a[href^="mailto:"], a[href^="MAILTO:"]')
            if normalize_email(a.get("href", ""))
        ])
        if len(emails) > 2:
            cur = cur.parent
            continue

        names = collect_name_candidates(cur, email)
        score = 0
        if names:
            score += names[0][0]
        if len(emails) == 1:
            score += 40
        if cur.name in {"article", "li", "tr"}:
            score += 15
        cls = " ".join(cur.get("class", [])).casefold()
        if any(x in cls for x in ("person", "faculty", "staff", "profile", "card", "member")):
            score += 25
        score -= depth * 2

        if score > best_score:
            best_score = score
            best = cur

        if score >= 100 and len(emails) <= 1:
            break
        cur = cur.parent

    return best if isinstance(best, Tag) else node


def _email_associated_name(block: Tag, email: str) -> Tuple[str, int]:
    names = collect_name_candidates(block, email)
    if not names:
        return "", 0
    score, name = names[0]
    return name, score


def _same_person_email_ok(name: str, email: str, block: Tag) -> bool:
    email = normalize_email(email)
    if not email:
        return False
    if not is_generic_email(email):
        return True

    # Generic mailbox is only allowed if the block explicitly presents a human
    # name AND contains only this single email. This is intentionally strict.
    if not name or not plausible_name(name):
        return False
    emails, _, _ = extract_email_candidates(block)
    return len(unique(emails)) == 1 and email_name_score(name, email) >= 20


def make_record_from_block(
    block: Tag,
    source_url: str,
    page_type: str,
    journal: str,
    method: str,
    inherited_role: str = "",
    explicit_country: str = "",
    explicit_name: str = "",
    explicit_affiliation: str = "",
    explicit_email: str = "",
) -> Optional[PersonRecord]:
    text = clean_text(block.get_text(" ", strip=True))
    primary_emails, conflict_alts, conflict = extract_email_candidates(block)

    if explicit_email:
        e = normalize_email(explicit_email)
        if e:
            primary_emails = [e] + [x for x in primary_emails if x != e]

    if not primary_emails:
        return None

    email = normalize_email(primary_emails[0])
    if not email:
        return None

    explicit_clean = clean_person_name_candidate(explicit_name) if explicit_name else ""
    if explicit_clean and plausible_name(explicit_clean):
        name = explicit_clean
        name_score = 120
    else:
        candidates = collect_name_candidates(block, email)
        if candidates:
            name_score, name = candidates[0]
            name = clean_person_name_candidate(name)
        else:
            name_score, name = 0, ""

    if not name or not plausible_name(name):
        name = name_from_email(email)
        name_score = 35 if name else 0

    if is_generic_email(email) and not _same_person_email_ok(name, email, block):
        return None

    role = inherited_role or current_role_for(block)
    links = extract_links(block, source_url)

    # IMPORTANT: only mailto visible/href conflicts become alternate emails.
    # Neighboring addresses are never alternates merely because they occur in
    # the same parent node.
    alternate_emails = " | ".join(
        e for e in unique(conflict_alts) if normalize_email(e) and e != email
    )

    affiliation = explicit_affiliation or split_affiliation_from_block(
        block, name, [email] + ([alternate_emails] if alternate_emails else []), role
    )

    # Keep affiliation local and bounded.
    affiliation = clean_text(affiliation)[:700]

    rec = PersonRecord(
        name=name,
        email=email,
        alternate_emails=alternate_emails,
        email_type=email_type(email),
        email_conflict="yes" if conflict else "no",
        page_type=page_type,
        journal_name=journal,
        editorial_role=role if page_type == "EDITORIAL_BOARD" else "",
        affiliation=affiliation,
        country=normalize_country(explicit_country),
        phone=extract_phone(block),
        orcid=links["orcid"] or extract_orcid_text(text),
        google_scholar=links["google_scholar"],
        scopus_author_id=extract_scopus(text),
        pubmed=links["pubmed"],
        personal_homepage=links["personal_homepage"],
        source_url=source_url,
        confidence=min(100, 45 + min(name_score, 120) // 2 + (10 if role else 0)),
        extraction_method=method,
    )
    return resolve_country(rec, affiliation)


def adapter_university_v7(soup: BeautifulSoup, source_url: str) -> List[PersonRecord]:
    """
    University-specific structural pass used before the old generic adapters.

    1. Parse every local mailto in its nearest single-person container.
    2. Parse plain-text emails from small leaf containers.
    3. Discover human-name profile links even when no email is on the listing.
    """
    records: List[PersonRecord] = []
    seen_email_nodes = set()

    # A) mailto anchored records
    for a in soup.select('a[href^="mailto:"], a[href^="MAILTO:"]'):
        email = normalize_email(a.get("href", "")) or normalize_email(a.get_text(" ", strip=True))
        if not email:
            continue
        block = _nearest_person_container(a, email)
        key = (email, clean_text(block.get_text(" ", strip=True))[:500])
        if key in seen_email_nodes:
            continue
        seen_email_nodes.add(key)

        # Best explicit profile/name inside the same block.
        profile = ""
        explicit_name = ""
        best_profile_score = 0
        for link in block.select("a[href]"):
            sc, u, n = _profile_link_score(link, source_url)
            if sc > best_profile_score:
                best_profile_score, profile, explicit_name = sc, u, n

        rec = make_record_from_block(
            block,
            source_url,
            "UNIVERSITY_DIRECTORY",
            "",
            "university_local_email_v7",
            explicit_name=explicit_name,
            explicit_email=email,
        )
        if rec:
            rec.profile_url = profile
            records.append(rec)

    # B) plain-text email nodes with no mailto
    # Keep only leaf-ish/short elements so we never absorb a full faculty page.
    for tag in soup.select("p, li, td, address, .email, [class*='contact']"):
        if tag.select_one('a[href^="mailto:"], a[href^="MAILTO:"]'):
            continue
        text = clean_text(tag.get_text(" ", strip=True))
        if not text or len(text) > 550:
            continue
        emails = extract_text_emails(text)
        if len(emails) != 1:
            continue
        email = emails[0]
        block = _nearest_person_container(tag, email, max_chars=1200)
        rec = make_record_from_block(
            block, source_url, "UNIVERSITY_DIRECTORY", "",
            "university_plain_email_v7", explicit_email=email
        )
        if rec:
            # Find a profile URL in the same local container.
            best = (0, "", "")
            for link in block.select("a[href]"):
                candidate = _profile_link_score(link, source_url)
                if candidate[0] > best[0]:
                    best = candidate
            if best[1]:
                rec.profile_url = best[1]
                if not rec.name and best[2]:
                    rec.name = best[2]
            records.append(rec)

    # C) profile placeholders. These are enriched later and exported only if an
    # email is eventually found.
    seen_profiles = set()
    for a in soup.select("a[href]"):
        score, profile_url, person_name = _profile_link_score(a, source_url)
        if score < 45 or not profile_url or not person_name:
            continue
        if profile_url in seen_profiles:
            continue
        seen_profiles.add(profile_url)

        # Reject the directory page itself.
        if normalize_url(profile_url).rstrip("/") == normalize_url(source_url).rstrip("/"):
            continue

        records.append(PersonRecord(
            name=person_name,
            page_type="UNIVERSITY_DIRECTORY",
            profile_url=profile_url,
            source_url=source_url,
            confidence=min(90, 50 + score // 2),
            extraction_method="university_profile_discovery_v7",
        ))

    return dedupe_records(records)


def _jsonld_people(soup: BeautifulSoup) -> List[Dict]:
    out = []

    def walk(obj):
        if isinstance(obj, dict):
            typ = obj.get("@type")
            types = typ if isinstance(typ, list) else [typ]
            if any(str(x).casefold() == "person" for x in types if x):
                out.append(obj)
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for v in obj:
                walk(v)

    for script in soup.select('script[type="application/ld+json"]'):
        raw = script.string or script.get_text("", strip=True)
        if not raw:
            continue
        try:
            walk(json.loads(raw))
        except Exception:
            continue
    return out


def _profile_name_candidates(soup: BeautifulSoup, root: Tag, name_hint: str = "") -> List[Tuple[int, str]]:
    candidates: List[Tuple[int, str]] = []
    if name_hint:
        n = clean_person_name_candidate(name_hint)
        if plausible_name(n):
            candidates.append((150, n))

    for person in _jsonld_people(soup):
        n = clean_person_name_candidate(person.get("name", ""))
        if plausible_name(n):
            candidates.append((145, n))

    for selector, score in (
        ('[itemprop="name"]', 135),
        ("main h1", 130),
        ("article h1", 130),
        ("h1", 125),
        (".person-name", 120),
        (".faculty-name", 120),
        (".profile-name", 120),
    ):
        for el in soup.select(selector):
            n = clean_person_name_candidate(el.get_text(" ", strip=True))
            if plausible_name(n):
                candidates.append((score, n))

    # og:title only if it looks like a person, not a page title.
    og = soup.select_one('meta[property="og:title"], meta[name="twitter:title"]')
    if og:
        n = clean_person_name_candidate(og.get("content", ""))
        if plausible_name(n):
            candidates.append((90, n))

    seen = {}
    for score, name in candidates:
        key = normalize_name_key(name) or name.casefold()
        if key not in seen or score > seen[key][0]:
            seen[key] = (score, name)
    return sorted(seen.values(), key=lambda x: x[0], reverse=True)


def _profile_email_candidates(soup: BeautifulSoup, root: Tag) -> List[Tuple[int, str, Tag]]:
    candidates: List[Tuple[int, str, Tag]] = []
    seen = set()

    for person in _jsonld_people(soup):
        value = person.get("email", "")
        if isinstance(value, list):
            vals = value
        else:
            vals = [value]
        for v in vals:
            e = normalize_email(v)
            if e and e not in seen:
                # JSON-LD Person email is strongest evidence.
                candidates.append((180, e, root))
                seen.add(e)

    for a in root.select('a[href^="mailto:"], a[href^="MAILTO:"]'):
        e = normalize_email(a.get("href", "")) or normalize_email(a.get_text(" ", strip=True))
        if not e or e in seen:
            continue
        local = _nearest_person_container(a, e, max_chars=1100)
        score = 120
        label_context = clean_text(local.get_text(" ", strip=True)).casefold()
        if any(x in label_context for x in ("administrative contact", "administrator", "assistant to", "media contact")):
            score -= 80
        if is_generic_email(e):
            score -= 70
        candidates.append((score, e, local))
        seen.add(e)

    # Plain text only from short contact-like elements, never whole biography.
    for el in root.select("p, li, address, .contact, .contact-info, [class*='email']"):
        txt = clean_text(el.get_text(" ", strip=True))
        if not txt or len(txt) > 500:
            continue
        for e in extract_text_emails(txt):
            if e in seen:
                continue
            score = 75 - (60 if is_generic_email(e) else 0)
            candidates.append((score, e, el))
            seen.add(e)

    return sorted(candidates, key=lambda x: x[0], reverse=True)


def _choose_profile_email(
    candidates: List[Tuple[int, str, Tag]],
    person_name: str,
) -> Tuple[str, List[str], bool, Optional[Tag]]:
    if not candidates:
        return "", [], False, None

    ranked = []
    for base, email, block in candidates:
        score = base + email_name_score(person_name, email)
        if is_generic_email(email):
            score -= 80
        ranked.append((score, email, block))
    ranked.sort(key=lambda x: x[0], reverse=True)

    best_score, best_email, best_block = ranked[0]

    # Strong anti-contamination rule:
    # - accept strong structured email
    # - or a non-generic email when only one exists
    non_generic = [x for x in ranked if not is_generic_email(x[1])]
    if is_generic_email(best_email):
        return "", [], False, None
    if best_score < 70 and len(non_generic) != 1:
        return "", [], False, None

    # Do not label unrelated page emails as alternates.
    return best_email, [], False, best_block


def _secondary_profile_urls(soup: BeautifulSoup, current_url: str, name_hint: str) -> List[str]:
    current_host = urlparse(current_url).netloc.casefold().replace("www.", "")
    scored = []
    for a in soup.select("a[href]"):
        href = clean_text(a.get("href", ""))
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            continue
        url = normalize_url(urljoin(current_url, href))
        if not url.startswith(("http://", "https://")):
            continue
        if normalize_url(url).rstrip("/") == normalize_url(current_url).rstrip("/"):
            continue

        host = urlparse(url).netloc.casefold().replace("www.", "")
        path = urlparse(url).path.casefold()
        anchor_text = clean_person_name_candidate(a.get_text(" ", strip=True))
        score = 0

        if any(h in host for h in AUTHORITATIVE_PROFILE_HOST_HINTS_V7):
            score += 80
        if "profiles.stanford.edu" in host:
            score += 100
        if any(h in path for h in PROFILE_PATH_HINTS_V7):
            score += 25
        if name_hint and plausible_name(anchor_text):
            akey = set(normalize_name_key(anchor_text).split())
            hkey = set(normalize_name_key(name_hint).split())
            if akey and hkey and akey & hkey:
                score += 45
        if host != current_host:
            score += 5

        # External social/publication links are not authoritative staff profiles.
        if any(x in host for x in (
            "linkedin.com", "twitter.com", "x.com", "facebook.com", "instagram.com",
            "orcid.org", "scholar.google", "researchgate.net", "youtube.com"
        )):
            continue

        if score >= 70:
            scored.append((score, url))

    seen = set()
    out = []
    for _, url in sorted(scored, reverse=True):
        if url not in seen:
            seen.add(url)
            out.append(url)
    return out[:3]


def extract_profile_record(html: str, url: str, name_hint: str, page_type: str) -> Optional[PersonRecord]:
    soup = BeautifulSoup(html, "html.parser")
    normalize_dom(soup)
    root = soup.select_one("main") or soup.select_one("article") or soup.select_one("#main-content") or soup.select_one("#content") or soup.body or soup

    names = _profile_name_candidates(soup, root, name_hint)
    name = names[0][1] if names else ""
    if not name and name_hint and plausible_name(clean_person_name_candidate(name_hint)):
        name = clean_person_name_candidate(name_hint)

    email_candidates = _profile_email_candidates(soup, root)
    email, alternates, conflict, email_block = _choose_profile_email(email_candidates, name)
    if not email:
        return None

    local_block = email_block or root
    local_text = clean_text(local_block.get_text(" ", strip=True))
    links = extract_links(root, url)

    affiliation = split_affiliation_from_block(
        local_block, name, [email] + alternates, current_role_for(local_block)
    )
    affiliation = clean_text(affiliation)[:700]

    rec = PersonRecord(
        name=name if plausible_name(name) else "",
        email=email,
        alternate_emails=" | ".join(alternates),
        email_conflict="yes" if conflict else "no",
        email_type=email_type(email),
        page_type=page_type,
        affiliation=affiliation,
        country="",
        country_source="",
        phone=extract_phone(local_block),
        orcid=links["orcid"] or extract_orcid_text(clean_text(root.get_text(" ", strip=True))[:5000]),
        google_scholar=links["google_scholar"],
        scopus_author_id=extract_scopus(clean_text(root.get_text(" ", strip=True))[:5000]),
        pubmed=links["pubmed"],
        personal_homepage=links["personal_homepage"],
        profile_url=url,
        confidence=96 if name else 82,
        extraction_method="profile_v7",
    )

    # Country from same-person/local affiliation only. Never scan a full biography.
    return resolve_country(rec, affiliation)


async def enrich_missing_profiles(context, records: List[PersonRecord], sem: asyncio.Semaphore) -> List[PersonRecord]:
    async def fetch_one(url: str):
        r = await fetch_http(url)
        if r["ok"]:
            return r["html"], r.get("url", url)
        r = await playwright_fetch(context, url)
        if r["ok"]:
            return r["html"], r.get("url", url)
        return None, url

    async def one(rec: PersonRecord):
        if not rec.profile_url:
            return rec
        if rec.email and rec.page_type == "EDITORIAL_BOARD":
            return rec

        async with sem:
            try:
                html, final_url = await fetch_one(rec.profile_url)
                if not html:
                    return rec

                p = extract_profile_record(html, final_url, rec.name, rec.page_type)

                # Stanford AIMI/AI Health and similar microsites frequently link
                # to an authoritative institutional profile that has the email.
                if p is None or not p.email:
                    soup = BeautifulSoup(html, "html.parser")
                    secondary = _secondary_profile_urls(soup, final_url, rec.name)
                    for second_url in secondary:
                        html2, final2 = await fetch_one(second_url)
                        if not html2:
                            continue
                        p = extract_profile_record(html2, final2, rec.name, rec.page_type)
                        if p and p.email:
                            # Preserve the listing/source profile while recording
                            # the authoritative contact page as personal homepage.
                            p.personal_homepage = p.personal_homepage or final2
                            break

                if not p or not p.email:
                    return rec

                # Preserve listing identity if it is trustworthy.
                if rec.name and plausible_name(rec.name):
                    p.name = rec.name

                for field_name in COLUMNS:
                    if field_name in {
                        "employee_name", "employee_email", "source_url",
                        "page_type", "journal_name"
                    }:
                        continue
                    old = getattr(rec, field_name, "") if hasattr(rec, field_name) else ""
                    new = getattr(p, field_name, "") if hasattr(p, field_name) else ""
                    if not old and new:
                        setattr(rec, field_name, new)

                rec.email = rec.email or p.email
                if not rec.name and p.name:
                    rec.name = p.name
                rec.confidence = max(rec.confidence, p.confidence)
                rec.extraction_method = (
                    f"{rec.extraction_method}+profile_v7"
                    if rec.extraction_method else "profile_v7"
                )
                return rec
            except Exception:
                return rec

    # Unique profile URL to avoid fetching the same profile many times.
    unique_records = []
    duplicates = []
    seen = {}
    for r in records:
        key = normalize_url(r.profile_url).casefold() if r.profile_url else ""
        if key and key in seen:
            duplicates.append((r, seen[key]))
        else:
            if key:
                seen[key] = r
            unique_records.append(r)

    enriched = list(await asyncio.gather(*(one(r) for r in unique_records)))

    # We deliberately do not re-add duplicate profile placeholders. Records with
    # direct distinct emails are already represented separately by dedupe_records.
    return enriched


def resolve_country(record: PersonRecord, local_text: str = "") -> PersonRecord:
    if record.country:
        c = normalize_country(record.country)
        record.country = c
        if c and not record.country_source:
            record.country_source = "explicit"
        return record

    # Strict locality: affiliation/address first.
    for text, source in (
        (record.affiliation, "affiliation"),
        (record.address, "address"),
    ):
        c = extract_country_from_text(text)
        if c:
            record.country = c
            record.country_source = source
            return record

    # local_text is allowed only if it is genuinely local, never a long bio/page.
    local_text = clean_text(local_text)
    if local_text and len(local_text) <= 500:
        c = extract_country_from_text(local_text)
        if c:
            record.country = c
            record.country_source = "local_person_block"
            return record

    c = country_from_institution(record.affiliation)
    if c:
        record.country = c
        record.country_source = "institution_hint"
        return record

    c = country_from_tld(record.email)
    if c:
        record.country = c
        record.country_source = "email_tld"
        return record

    record.country = ""
    return record


def dedupe_records(records: List[PersonRecord]) -> List[PersonRecord]:
    """
    Preserve shared editorial mailboxes across different named people while
    merging true duplicate observations of the same identity.
    """
    merged: Dict[Tuple, PersonRecord] = {}

    for r in records:
        r.email = normalize_email(r.email)
        nkey = normalize_name_key(r.name)
        purl = normalize_url(r.profile_url).casefold()

        if purl and not r.email:
            key = ("profile", purl)
        elif r.email and nkey:
            key = ("person_email", nkey, r.email)
        elif r.email:
            key = ("email_only", r.email)
        elif purl:
            key = ("profile", purl)
        else:
            key = ("name_source", nkey, normalize_url(r.source_url).casefold())

        if key not in merged:
            merged[key] = r
            continue

        cur = merged[key]
        # Prefer a trustworthy visible/profile name over blank/email-derived.
        if (not cur.name or not plausible_name(cur.name)) and plausible_name(r.name):
            cur.name = r.name
        if not cur.email and r.email:
            cur.email = r.email
        if not cur.country and r.country:
            cur.country = r.country
            cur.country_source = r.country_source
        if not cur.profile_url and r.profile_url:
            cur.profile_url = r.profile_url

        for field_name in COLUMNS:
            if not hasattr(cur, field_name):
                continue
            old = getattr(cur, field_name)
            new = getattr(r, field_name)
            if field_name == "confidence":
                cur.confidence = max(int(cur.confidence or 0), int(r.confidence or 0))
            elif field_name == "alternate_emails":
                # Only retain explicit conflict alternates already produced by a
                # same-anchor visible/href mismatch.
                vals = []
                for part in (old, new):
                    vals.extend([x.strip() for x in str(part or "").split("|") if x.strip()])
                vals = [
                    e for e in unique(vals)
                    if normalize_email(e) and normalize_email(e) != normalize_email(cur.email)
                ]
                cur.alternate_emails = " | ".join(vals)
            elif not old and new:
                setattr(cur, field_name, new)

    return list(merged.values())


def validate_record(r: PersonRecord) -> Tuple[bool, str]:
    r.email = normalize_email(r.email)
    if not r.email:
        return False, "missing_or_invalid_email"

    # Generic website mailbox with no verified human association is not a
    # university-person record.
    if is_generic_email(r.email):
        if not r.name or not plausible_name(r.name) or email_name_score(r.name, r.email) < 20:
            return False, "generic_site_email_not_person"

    scraped_name = clean_person_name_candidate(r.name)
    if scraped_name and plausible_name(scraped_name):
        r.name = scraped_name
    else:
        # Compact usernames remain blank.
        r.name = name_from_email(r.email)

    country = normalize_country(r.country)
    r.country = country if country else ""

    # Clean/validate alternates. Never keep primary or invalid values.
    alts = []
    for raw in str(r.alternate_emails or "").split("|"):
        e = normalize_email(raw.strip())
        if e and e != r.email:
            alts.append(e)
    r.alternate_emails = " | ".join(unique(alts))

    r.affiliation = clean_text(r.affiliation)[:700]
    return True, ""


def _university_parse_v7(html: str, page_url: str) -> Tuple[str, List[PersonRecord]]:
    soup = BeautifulSoup(html, "html.parser")
    normalize_dom(soup)
    records = adapter_university_v7(soup, page_url)

    # Keep the old table/card parsing only as secondary support, then run the
    # strict deduper/validator later. It remains useful for Campbell/Case and
    # some table-based university directories.
    records += adapter_table_rows(soup, page_url, "UNIVERSITY_DIRECTORY", "")
    records += adapter_cards(soup, page_url, "UNIVERSITY_DIRECTORY", "")

    # Generic broad email fallback is intentionally NOT used for university
    # pages because it caused UH/Creighton/WashU cross-person contamination.
    return "UNIVERSITY_DIRECTORY", dedupe_records(records)


async def parse_html(html: str, source_url: str, llm_enabled: bool, llm_sem: asyncio.Semaphore) -> Tuple[str, List[PersonRecord]]:
    soup0 = BeautifulSoup(html, "html.parser")
    page_type = detect_page_type(soup0, source_url)

    if page_type != "EDITORIAL_BOARD":
        return _university_parse_v7(html, source_url)

    # Editorial-board logic keeps the existing structured adapters.
    journal = journal_name(soup0)
    normalize_dom(soup0)
    records: List[PersonRecord] = []
    records += adapter_cards(soup0, source_url, page_type, journal)
    records += adapter_table_multirow(soup0, source_url, page_type, journal)
    records += adapter_table_rows(soup0, source_url, page_type, journal)
    records += adapter_list_items(soup0, source_url, page_type, journal)
    records += adapter_paragraphs(soup0, source_url, page_type, journal)
    records += adapter_sequential_flow(soup0, source_url, page_type, journal)
    records += discover_js_profile_placeholders(soup0, source_url, page_type, journal)
    records += adapter_email_anchors(soup0, source_url, page_type, journal)
    for r in records:
        r.source_url = r.source_url or source_url
        r.page_type = r.page_type or page_type
        r.journal_name = r.journal_name or journal
    return page_type, dedupe_records(records)


def save_excel(records: List[PersonRecord]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"
    ws.append([excel_safe(c) for c in COLUMNS])

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in records:
        data = asdict(r)
        ws.append([excel_safe(data.get(c, "")) for c in COLUMNS])

    widths = {
        "A": 42, "B": 42, "C": 22, "D": 55, "E": 18, "F": 14,
        "G": 18, "H": 38, "I": 24, "J": 48, "K": 32, "L": 32, "M": 26,
        "N": 28, "O": 75, "P": 50, "Q": 45, "R": 45, "S": 50, "T": 50,
        "U": 42, "V": 24, "W": 65, "X": 24, "Y": 65, "Z": 65, "AA": 24,
        "AB": 24, "AC": 65, "AD": 65, "AE": 85, "AF": 25, "AG": 14, "AH": 28,
        "AI": 20, "AJ": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(XLSX_FILE)


def _safe_json_value(value):
    if isinstance(value, dict):
        return {excel_safe(k): _safe_json_value(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_safe_json_value(v) for v in value]
    if isinstance(value, tuple):
        return [_safe_json_value(v) for v in value]
    if isinstance(value, str):
        return excel_safe(value)
    return value


def save_errors(errors: List[Dict]):
    tmp = ERROR_FILE.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(_safe_json_value(errors), f, ensure_ascii=False, indent=2)
    tmp.replace(ERROR_FILE)


async def scrape_source(context, source_url: str, llm_enabled: bool, llm_sem: asyncio.Semaphore,
                        profile_sem: asyncio.Semaphore, errors: List[Dict]) -> List[PersonRecord]:
    """
    HTTP first. If HTTP succeeds but yields no accepted email records on a
    university directory, automatically render with Playwright and retry.
    """
    all_records: List[PersonRecord] = []
    current_url = source_url
    visited = set()

    for page_num in range(1, MAX_PAGES_PER_URL + 1):
        if not current_url or current_url in visited:
            break
        visited.add(current_url)

        print(f"\n   [PAGE {page_num}] {current_url}", flush=True)
        print("   [FETCH] HTTP first...", flush=True)

        r = await fetch_http(current_url)
        method = "http"
        if not r["ok"]:
            print(f"   [HTTP FAILED] {r['error']}", flush=True)
            print("   [FALLBACK] Playwright...", flush=True)
            r = await playwright_fetch(context, current_url)
            method = "playwright"
        else:
            print(f"   [HTTP OK] {len(r['html']):,} chars", flush=True)

        if not r["ok"]:
            errors.append({
                "type": "fetch_failed",
                "source_url": source_url,
                "page_url": current_url,
                "error": excel_safe(r["error"]),
            })
            break

        actual_page_url = r.get("url") or current_url
        page_type, records = await parse_html(r["html"], actual_page_url, llm_enabled, llm_sem)
        print(f"   [PARSE] type={page_type} raw_records={len(records)} method={method}", flush=True)

        records = await enrich_missing_profiles(context, records, profile_sem)
        records = await fill_missing_countries(records, llm_enabled, llm_sem)

        # Ambiguous-only semantic identity validation with local Qwen3.
        records = await llm_filter_records(
            records,
            llm_enabled,
            llm_sem,
            errors,
            source_url,
        )

        accepted_page: List[PersonRecord] = []
        rejected_page = []
        for rec in records:
            # Preserve original source URL even when current page redirects.
            rec.source_url = source_url
            valid, reason = validate_record(rec)
            if valid:
                accepted_page.append(rec)
            else:
                rejected_page.append((rec, reason))

        # HTTP can return a JS shell or incomplete markup with 200 OK. If it
        # produced no usable university emails, render the same page and retry.
        if (
            page_type == "UNIVERSITY_DIRECTORY"
            and method == "http"
            and len(accepted_page) == 0
        ):
            print("   [HTTP PARSE EMPTY] Retrying rendered DOM with Playwright...", flush=True)
            br = await playwright_fetch(context, current_url)
            if br["ok"]:
                btype, brecords = await parse_html(
                    br["html"], br.get("url") or current_url, llm_enabled, llm_sem
                )
                brecords = await enrich_missing_profiles(context, brecords, profile_sem)
                brecords = await fill_missing_countries(brecords, llm_enabled, llm_sem)
                brecords = await llm_filter_records(
                    brecords,
                    llm_enabled,
                    llm_sem,
                    errors,
                    source_url,
                )

                accepted_page = []
                rejected_page = []
                for rec in brecords:
                    rec.source_url = source_url
                    valid, reason = validate_record(rec)
                    if valid:
                        accepted_page.append(rec)
                    else:
                        rejected_page.append((rec, reason))
                method = "playwright_retry"

        for rec in accepted_page:
            print(
                f"      {(rec.name or '')[:38]:38} | {rec.email[:36]:36} | "
                f"{rec.country[:18]:18} | {rec.editorial_role[:24]}",
                flush=True,
            )
            all_records.append(rec)

        for rec, reason in rejected_page:
            errors.append({
                "type": "rejected_record",
                "reason": reason,
                "name": excel_safe(rec.name),
                "email": excel_safe(rec.email),
                "country": excel_safe(rec.country),
                "profile_url": excel_safe(rec.profile_url),
                "source_url": source_url,
                "method": rec.extraction_method,
            })

        print(
            f"   [QUALITY] discovered={len(records)} accepted={len(accepted_page)} "
            f"rejected={len(rejected_page)} blank_names="
            f"{sum(1 for x in accepted_page if not x.name)}",
            flush=True,
        )

        if page_num == 1 and len(accepted_page) == 0:
            errors.append({
                "type": "no_verified_people",
                "source_url": source_url,
                "page_url": current_url,
                "message": "No records with a verified valid email were found after HTTP/browser/profile checks.",
            })

        if not FOLLOW_PAGINATION:
            break

        # Editorial pages are not generically paginated; generic 'Next' links can
        # leave the board and create false positives. University directory paging
        # is allowed.
        if page_type == "EDITORIAL_BOARD":
            break

        nxt = next_page_url(r["html"], actual_page_url)
        if not nxt or nxt in visited:
            break
        if urlparse(nxt).netloc.casefold().replace("www.", "") != urlparse(source_url).netloc.casefold().replace("www.", ""):
            break

        current_url = nxt
        await polite_delay()

    return dedupe_records(all_records)


# =============================================================================
# V10 CANCER-SITE PROFILE CLICK + MAILTO OVERRIDES
# =============================================================================
# Targeted support:
#   1) clinicalresearch.radonc.jhmi.edu
#   2) uhcancercenter.org clinical faculty
#   3) mayo.edu cancer research faculty
#
# Strategy:
#   LISTING -> discover visible person name + profile URL
#           -> open profile (HTTP first, browser fallback)
#           -> extract email from:
#                mailto:
#                visible email text
#                data-email / data-mail / data-address attributes
#                JSON-LD Person.email
#                Cloudflare data-cfemail
#           -> require same-person profile evidence
#           -> export only records with verified email
#
# Name/country remain optional; email remains mandatory.

from urllib.parse import parse_qs


CANCER_PROFILE_MAX_CONCURRENCY = 2
MAYO_MAX_PAGES = 60

CANCER_BAD_PROFILE_TEXT = {
    "email", "publications", "learn more", "read more", "view profile", "profile",
    "clinical trials", "find a doctor", "contact us", "research", "faculty",
    "staff", "directory", "cancer care team",
}

CANCER_NONPERSON_PATHS = (
    "/about", "/contact", "/clinical-trials", "/patients", "/research-program",
    "/events", "/news", "/careers", "/donate", "/privacy", "/disclaimer",
)


def _host(url: str) -> str:
    return urlparse(url).netloc.casefold().replace("www.", "").replace("ww.", "").replace("wws.", "")


def _is_jhmi_cancer(url: str) -> bool:
    return _host(url) == "clinicalresearch.radonc.jhmi.edu"


def _is_uh_cancer(url: str) -> bool:
    return "uhcancercenter.org" in _host(url)


def _is_mayo_cancer(url: str) -> bool:
    return _host(url) == "mayo.edu" and "/cancer-research/faculty-staff" in urlparse(url).path.casefold()


def _decode_cfemail(hex_string: str) -> str:
    """Decode Cloudflare email-protection data-cfemail values."""
    try:
        data = bytes.fromhex(clean_text(hex_string))
        if len(data) < 2:
            return ""
        key = data[0]
        decoded = "".join(chr(b ^ key) for b in data[1:])
        return normalize_email(decoded)
    except Exception:
        return ""


def _extract_all_emails_from_dom(root: Tag) -> List[Tuple[str, int, Optional[Tag], str]]:
    """
    Return tuples: (email, evidence_score, local_node, evidence_type).

    Scores:
      JSON-LD Person email      180
      direct mailto             150
      data-email attributes     140
      Cloudflare data-cfemail   135
      visible local text        95
    """
    found: Dict[str, Tuple[int, Optional[Tag], str]] = {}

    def add(email, score, node, kind):
        e = normalize_email(email)
        if not e:
            return
        old = found.get(e)
        if old is None or score > old[0]:
            found[e] = (score, node, kind)

    soup = root if isinstance(root, BeautifulSoup) else root.find_parent() or root

    # JSON-LD Person objects.
    try:
        search_root = root if isinstance(root, BeautifulSoup) else root
        for script in search_root.select('script[type="application/ld+json"]'):
            raw = script.string or script.get_text("", strip=True)
            if not raw:
                continue
            try:
                payload = json.loads(raw)
            except Exception:
                continue

            stack = [payload]
            while stack:
                obj = stack.pop()
                if isinstance(obj, dict):
                    typ = obj.get("@type")
                    types = typ if isinstance(typ, list) else [typ]
                    is_person = any(str(x).casefold() == "person" for x in types if x)
                    if is_person:
                        vals = obj.get("email", [])
                        vals = vals if isinstance(vals, list) else [vals]
                        for v in vals:
                            add(v, 180, root, "jsonld_person_email")
                    stack.extend(obj.values())
                elif isinstance(obj, list):
                    stack.extend(obj)
    except Exception:
        pass

    # mailto
    for a in root.select('a[href]'):
        href = clean_text(a.get("href", ""))
        if href.casefold().startswith("mailto:"):
            add(href, 150, a, "mailto")
            visible = normalize_email(a.get_text(" ", strip=True))
            if visible:
                add(visible, 155, a, "visible_mailto")

    # common data attributes / onclick scripts
    for el in root.select("*"):
        for attr in ("data-email", "data-mail", "data-address", "data-email-address"):
            if el.has_attr(attr):
                add(el.get(attr, ""), 140, el, attr)

        if el.has_attr("data-cfemail"):
            e = _decode_cfemail(el.get("data-cfemail", ""))
            if e:
                add(e, 135, el, "cloudflare_cfemail")

        for attr in ("onclick", "href"):
            val = clean_text(el.get(attr, ""))
            if "@" in val or "mailto" in val.casefold():
                for e in extract_text_emails(val):
                    add(e, 125, el, f"{attr}_embedded_email")

    # Visible text only from small local elements, never the whole profile.
    for el in root.select(
        "p, li, dd, dt, address, span, div.email, div.contact, "
        "[class*='email'], [class*='contact'], [id*='email'], [id*='contact']"
    ):
        txt = clean_text(el.get_text(" ", strip=True))
        if not txt or len(txt) > 600:
            continue
        for e in extract_text_emails(txt):
            add(e, 95, el, "visible_local_text")

    return [
        (email, score, node, kind)
        for email, (score, node, kind) in found.items()
    ]


def _name_tokens(name: str) -> List[str]:
    name = clean_person_name_candidate(name)
    words = re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿĀ-ž]+", name.casefold())
    drop = {
        "md", "phd", "mph", "do", "mbbs", "facp", "facs", "dnp", "aprn",
        "professor", "associate", "assistant", "doctor", "dr",
    }
    return [x for x in words if len(x) >= 2 and x not in drop]


def _profile_identity_matches(name_hint: str, profile_name: str) -> bool:
    """Require name overlap when a listing supplied a person name."""
    hint = set(_name_tokens(name_hint))
    prof = set(_name_tokens(profile_name))
    if not hint:
        return True
    if not prof:
        return False

    # surname/meaningful token overlap
    if hint & prof:
        return True

    hk = normalize_name_key(name_hint)
    pk = normalize_name_key(profile_name)
    return bool(hk and pk and (hk == pk or hk in pk or pk in hk))


def _pick_profile_root(soup: BeautifulSoup) -> Tag:
    return (
        soup.select_one("main")
        or soup.select_one("article")
        or soup.select_one("#main-content")
        or soup.select_one("#content")
        or soup.select_one(".main-content")
        or soup.body
        or soup
    )


def _profile_display_name(soup: BeautifulSoup, root: Tag, name_hint: str = "") -> str:
    candidates: List[Tuple[int, str]] = []

    # JSON-LD
    for person in _jsonld_people(soup):
        n = clean_person_name_candidate(person.get("name", ""))
        if plausible_name(n):
            candidates.append((190, n))

    selectors = (
        ('[itemprop="name"]', 180),
        ("main h1", 175),
        ("main h2", 165),
        ("article h1", 175),
        ("article h2", 165),
        (".profile-name", 170),
        (".person-name", 170),
        (".faculty-name", 170),
        ("h1", 150),
        ("h2", 135),
    )
    for sel, score in selectors:
        for el in soup.select(sel):
            n = clean_person_name_candidate(el.get_text(" ", strip=True))
            if plausible_name(n) and n.casefold() not in CANCER_BAD_PROFILE_TEXT:
                candidates.append((score, n))

    # Listing name is valuable but must not override a clear profile heading.
    if name_hint:
        n = clean_person_name_candidate(name_hint)
        if plausible_name(n):
            candidates.append((160, n))

    if not candidates:
        return ""

    candidates.sort(key=lambda x: x[0], reverse=True)

    if name_hint:
        for _, n in candidates:
            if _profile_identity_matches(name_hint, n):
                return n
        # Preserve trustworthy listing identity if profile heading is noisy.
        n = clean_person_name_candidate(name_hint)
        return n if plausible_name(n) else ""

    return candidates[0][1]


def _email_person_score(name: str, email: str, base: int, local_node: Optional[Tag]) -> int:
    score = base
    if is_generic_email(email) or is_strict_generic_email(email):
        return -999

    score += email_name_score(name, email)

    if local_node is not None:
        try:
            local = _nearest_person_container(local_node, email, max_chars=1000)
            context = clean_text(local.get_text(" ", strip=True))
            low = context.casefold()

            if any(x in low for x in (
                "media inquiries", "general information", "employment opportunities",
                "administrative contact", "development office", "webmaster",
            )):
                score -= 120

            tokens = _name_tokens(name)
            if tokens and any(t in low for t in tokens):
                score += 30
        except Exception:
            pass

    return score


def extract_cancer_profile_record(
    html: str,
    profile_url: str,
    source_url: str,
    name_hint: str,
) -> Optional[PersonRecord]:
    """
    Strict profile-page extraction.

    Critical safeguard:
    Footer/global emails cannot be assigned to the profile person unless there
    is local person evidence or strong email/name agreement.
    """
    soup = BeautifulSoup(html, "html.parser")

    # Preserve JSON-LD before normalize_dom() deletes script tags.
    jsonld_people = _jsonld_people(soup)
    root = _pick_profile_root(soup)

    profile_name = _profile_display_name(soup, root, name_hint)
    if name_hint and profile_name and not _profile_identity_matches(name_hint, profile_name):
        return None
    name = profile_name or clean_person_name_candidate(name_hint)

    # Collect emails before removing noisy elements.
    candidates = _extract_all_emails_from_dom(root)

    # Include JSON-LD captured from complete soup.
    for person in jsonld_people:
        person_name = clean_person_name_candidate(person.get("name", ""))
        if name_hint and person_name and not _profile_identity_matches(name_hint, person_name):
            continue
        vals = person.get("email", [])
        vals = vals if isinstance(vals, list) else [vals]
        for value in vals:
            e = normalize_email(value)
            if e:
                candidates.append((e, 190, root, "jsonld_person_email"))

    ranked = []
    seen = set()
    for email, base, node, kind in candidates:
        email = normalize_email(email)
        if not email or email in seen:
            continue
        seen.add(email)

        score = _email_person_score(name, email, base, node)
        if score <= 0:
            continue

        # When name exists and an email comes only from weak visible text, require
        # some name/email or local name-context association.
        if name and base < 120:
            direct_match = email_name_score(name, email)
            local_has_name = False
            if node is not None:
                try:
                    ctx = clean_text(
                        _nearest_person_container(node, email, max_chars=900)
                        .get_text(" ", strip=True)
                    ).casefold()
                    local_has_name = any(t in ctx for t in _name_tokens(name))
                except Exception:
                    pass
            if direct_match == 0 and not local_has_name:
                continue

        ranked.append((score, email, node, kind))

    if not ranked:
        return None

    ranked.sort(key=lambda x: x[0], reverse=True)
    best_score, email, email_node, email_kind = ranked[0]

    # Same profile must support the email strongly.
    if best_score < 120:
        return None

    normalized_soup = BeautifulSoup(html, "html.parser")
    normalize_dom(normalized_soup)
    normalized_root = _pick_profile_root(normalized_soup)

    links = extract_links(normalized_root, profile_url)
    local_block = normalized_root
    if email_node is not None:
        # email_node belongs to original soup, so derive compact text separately.
        try:
            local_text = clean_text(
                _nearest_person_container(email_node, email, max_chars=1200)
                .get_text(" ", strip=True)
            )
        except Exception:
            local_text = ""
    else:
        local_text = ""

    affiliation = ""
    if local_text:
        temp = BeautifulSoup(f"<div>{local_text}</div>", "html.parser").div
        affiliation = split_affiliation_from_block(temp, name, [email], "")
    affiliation = clean_text(affiliation)[:700]

    rec = PersonRecord(
        name=name if plausible_name(name) else "",
        email=email,
        alternate_emails="",
        email_type=email_type(email),
        email_conflict="no",
        page_type="UNIVERSITY_DIRECTORY",
        affiliation=affiliation,
        profile_url=profile_url,
        personal_homepage=links.get("personal_homepage", ""),
        source_url=source_url,
        country="",
        country_source="",
        confidence=min(100, max(90, best_score)),
        extraction_method=f"cancer_profile:{email_kind}",
    )
    return resolve_country(rec, affiliation)


def _person_anchor_name(anchor: Tag) -> str:
    text = clean_person_name_candidate(anchor.get_text(" ", strip=True))
    if not plausible_name(text):
        text = clean_person_name_candidate(
            anchor.get("aria-label", "") or anchor.get("title", "")
        )
    if not plausible_name(text):
        return ""
    if text.casefold() in CANCER_BAD_PROFILE_TEXT:
        return ""
    return text


def discover_jhmi_profiles(html: str, source_url: str) -> List[PersonRecord]:
    soup = BeautifulSoup(html, "html.parser")
    out = []
    seen = set()

    for a in soup.select("a[href]"):
        href = normalize_url(urljoin(source_url, a.get("href", "")))
        if not href:
            continue
        parsed = urlparse(href)
        qs = parse_qs(parsed.query)

        # Actual JHMI research physician pages use /research?faculty=<id>.
        is_faculty_query = (
            _host(href) == "clinicalresearch.radonc.jhmi.edu"
            and parsed.path.rstrip("/").casefold().endswith("/research")
            and bool(qs.get("faculty"))
        )
        if not is_faculty_query:
            continue

        name = _person_anchor_name(a)
        if not name:
            continue

        key = href.casefold()
        if key in seen:
            continue
        seen.add(key)

        out.append(PersonRecord(
            name=name,
            profile_url=href,
            source_url=source_url,
            page_type="UNIVERSITY_DIRECTORY",
            confidence=90,
            extraction_method="jhmi_faculty_profile_listing",
        ))

    return out


def discover_uh_profiles(html: str, source_url: str) -> List[PersonRecord]:
    """
    UH clinical faculty contains a mixture:
      - linked names -> profile pages
      - plain-text names -> no profile URL

    Only linked people are profile-enriched. We NEVER assign footer info/media/hr
    addresses to unlinked clinicians.
    """
    soup = BeautifulSoup(html, "html.parser")
    root = (
        soup.select_one("main")
        or soup.select_one("#content")
        or soup.select_one(".item-page")
        or soup.body
        or soup
    )
    out = []
    seen = set()

    for a in root.select("a[href]"):
        name = _person_anchor_name(a)
        if not name:
            continue

        href = normalize_url(urljoin(source_url, a.get("href", "")))
        if not href.startswith(("http://", "https://")):
            continue
        if "uhcancercenter.org" not in _host(href):
            continue

        path = urlparse(href).path.casefold()
        if any(x in path for x in CANCER_NONPERSON_PATHS):
            continue

        # UH faculty profiles commonly look like /surname-firstname or
        # /ueno-naoto. Reject long navigation paths.
        segments = [x for x in path.split("/") if x]
        if not segments or len(segments) > 3:
            continue

        # Require human credentials or 2+ person-name words.
        low_name = name.casefold()
        looks_credentialed = bool(
            re.search(r"\b(?:md|phd|mph|dnp|aprn|fnp|facp|faap|pharmd)\b", low_name, re.I)
        )
        if not looks_credentialed and len(_name_tokens(name)) < 2:
            continue

        key = href.casefold()
        if key in seen:
            continue
        seen.add(key)

        out.append(PersonRecord(
            name=name,
            profile_url=href,
            source_url=source_url,
            page_type="UNIVERSITY_DIRECTORY",
            confidence=88,
            extraction_method="uh_clinical_faculty_profile_listing",
        ))

    return out


def _mayo_card_records(html: str, source_url: str) -> List[PersonRecord]:
    """
    Mayo listing: each person block normally includes
      name/profile link + academic titles + Email mailto/button.

    Parse each faculty unit locally to prevent cross-person email assignment.
    """
    soup = BeautifulSoup(html, "html.parser")
    records = []

    # Find person-looking headings/name links.
    candidate_name_nodes = []
    for h in soup.select("h2, h3, h4"):
        name = clean_person_name_candidate(h.get_text(" ", strip=True))
        if plausible_name(name) and name.casefold() not in {"faculty", "adjunct faculty"}:
            candidate_name_nodes.append((h, name))

    # Also profile links if headings are not wrapping the name.
    for a in soup.select("a[href*='/research/faculty/']"):
        name = _person_anchor_name(a)
        if name:
            candidate_name_nodes.append((a, name))

    seen_profiles = set()
    seen_email = set()

    for node, name in candidate_name_nodes:
        # Find smallest parent that contains this person but not another person email.
        block = node
        chosen = node
        for _ in range(7):
            parent = block.parent if isinstance(block, Tag) else None
            if not parent or parent.name in {"body", "html"}:
                break
            txt = clean_text(parent.get_text(" ", strip=True))
            emails = _extract_all_emails_from_dom(parent)
            person_links = parent.select("a[href*='/research/faculty/']")
            if len(txt) <= 1800 and len(emails) <= 2 and len(person_links) <= 2:
                chosen = parent
                block = parent
            else:
                break

        profile_url = ""
        for a in chosen.select("a[href]"):
            href = normalize_url(urljoin(source_url, a.get("href", "")))
            if "/research/faculty/" in urlparse(href).path.casefold():
                profile_url = href
                break

        emails = _extract_all_emails_from_dom(chosen)
        ranked = []
        for email, base, email_node, kind in emails:
            email = normalize_email(email)
            if not email or is_generic_email(email):
                continue
            ranked.append((
                _email_person_score(name, email, base, email_node),
                email,
                kind,
            ))
        ranked.sort(reverse=True)

        if ranked and ranked[0][0] >= 120:
            score, email, kind = ranked[0]
            if email in seen_email:
                continue
            seen_email.add(email)

            text = clean_text(chosen.get_text(" ", strip=True))
            affiliation = split_affiliation_from_block(chosen, name, [email], "")
            rec = PersonRecord(
                name=name,
                email=email,
                page_type="UNIVERSITY_DIRECTORY",
                affiliation=clean_text(affiliation)[:700],
                profile_url=profile_url,
                source_url=source_url,
                country="United States",
                country_source="trusted_source_site",
                confidence=min(100, max(92, score)),
                extraction_method=f"mayo_listing_local:{kind}",
            )
            records.append(rec)
        elif profile_url and profile_url.casefold() not in seen_profiles:
            seen_profiles.add(profile_url.casefold())
            records.append(PersonRecord(
                name=name,
                profile_url=profile_url,
                source_url=source_url,
                page_type="UNIVERSITY_DIRECTORY",
                confidence=88,
                extraction_method="mayo_profile_listing",
            ))

    # Adjunct faculty can be plain text with explicit email.
    for email_el in soup.select("a[href^='mailto:'], a[href^='MAILTO:']"):
        email = normalize_email(email_el.get("href", ""))
        if not email or email in seen_email or is_generic_email(email):
            continue
        block = _nearest_person_container(email_el, email, max_chars=800)
        names = collect_name_candidates(block, email)
        if not names:
            continue
        name = names[0][1]
        if not plausible_name(name):
            continue
        seen_email.add(email)
        rec = make_record_from_block(
            block, source_url, "UNIVERSITY_DIRECTORY", "",
            "mayo_mailto_local", explicit_name=name, explicit_email=email
        )
        if rec:
            rec.country = "United States"
            rec.country_source = "trusted_source_site"
            records.append(rec)

    return dedupe_records(records)


async def _fetch_profile_record_strict(
    context,
    placeholder: PersonRecord,
    profile_sem: asyncio.Semaphore,
) -> Optional[PersonRecord]:
    async with profile_sem:
        r = await fetch_http(placeholder.profile_url)
        if not r["ok"]:
            r = await playwright_fetch(context, placeholder.profile_url)
        if not r["ok"]:
            return None

        return extract_cancer_profile_record(
            r["html"],
            r.get("url") or placeholder.profile_url,
            placeholder.source_url,
            placeholder.name,
        )


async def _enrich_cancer_placeholders(
    context,
    records: List[PersonRecord],
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
    source_url: str,
) -> List[PersonRecord]:
    direct = [r for r in records if normalize_email(r.email)]
    placeholders = [
        r for r in records
        if not normalize_email(r.email) and normalize_url(r.profile_url)
    ]

    # Unique profile URLs.
    unique_placeholders = []
    seen = set()
    for r in placeholders:
        key = normalize_url(r.profile_url).casefold()
        if key and key not in seen:
            seen.add(key)
            unique_placeholders.append(r)

    results = await asyncio.gather(*(
        _fetch_profile_record_strict(context, r, profile_sem)
        for r in unique_placeholders
    ))

    for placeholder, rec in zip(unique_placeholders, results):
        if rec and rec.email:
            direct.append(rec)
        else:
            errors.append({
                "type": "profile_no_verified_email",
                "source_url": source_url,
                "profile_url": placeholder.profile_url,
                "name": placeholder.name,
            })

    return dedupe_records(direct)


async def scrape_jhmi_cancer(
    context, source_url: str, profile_sem: asyncio.Semaphore, errors: List[Dict]
) -> List[PersonRecord]:
    r = await fetch_http(source_url)
    if not r["ok"]:
        r = await playwright_fetch(context, source_url)
    if not r["ok"]:
        errors.append({
            "type": "fetch_failed",
            "source_url": source_url,
            "error": r["error"],
        })
        return []

    # Direct manager/contact emails are valid people only when the page explicitly
    # names them in a local block. Capture them separately.
    soup = BeautifulSoup(r["html"], "html.parser")
    direct = []
    for a in soup.select('a[href^="mailto:"], a[href^="MAILTO:"]'):
        email = normalize_email(a.get("href", ""))
        if not email or is_generic_email(email):
            continue
        block = _nearest_person_container(a, email, max_chars=900)
        candidates = collect_name_candidates(block, email)
        name = candidates[0][1] if candidates else ""
        if not name or not plausible_name(name):
            continue
        rec = make_record_from_block(
            block, source_url, "UNIVERSITY_DIRECTORY", "",
            "jhmi_listing_mailto", explicit_name=name, explicit_email=email
        )
        if rec:
            rec.country = "United States"
            rec.country_source = "trusted_source_site"
            direct.append(rec)

    placeholders = discover_jhmi_profiles(r["html"], source_url)
    enriched = await _enrich_cancer_placeholders(
        context, placeholders, profile_sem, errors, source_url
    )

    for rec in enriched:
        rec.country = rec.country or "United States"
        rec.country_source = rec.country_source or "trusted_source_site"

    return dedupe_records(direct + enriched)


async def scrape_uh_cancer(
    context, source_url: str, profile_sem: asyncio.Semaphore, errors: List[Dict]
) -> List[PersonRecord]:
    r = await fetch_http(source_url)
    if not r["ok"]:
        r = await playwright_fetch(context, source_url)
    if not r["ok"]:
        errors.append({
            "type": "fetch_failed",
            "source_url": source_url,
            "error": r["error"],
        })
        return []

    placeholders = discover_uh_profiles(r["html"], source_url)
    enriched = await _enrich_cancer_placeholders(
        context, placeholders, profile_sem, errors, source_url
    )

    # DO NOT scrape info@, media@, hr@ or arbitrary footer/development addresses.
    for rec in enriched:
        rec.country = rec.country or "United States"
        rec.country_source = rec.country_source or "trusted_source_site"

    return dedupe_records(enriched)


async def _mayo_rendered_pages(context, source_url: str) -> List[str]:
    """
    Render Mayo and iterate numeric pagination when it is JS-driven.

    We stop when:
      - no new faculty signature appears,
      - next numeric button is missing/disabled,
      - or MAYO_MAX_PAGES is reached.
    """
    page = await context.new_page()
    html_pages = []
    try:
        await page.goto(source_url, wait_until="domcontentloaded", timeout=PLAYWRIGHT_TIMEOUT)
        await dismiss_popups(page)
        await page.wait_for_timeout(1200)

        seen_signatures = set()

        for page_index in range(1, MAYO_MAX_PAGES + 1):
            await auto_scroll(page)
            await page.wait_for_timeout(500)
            html = await page.content()

            soup = BeautifulSoup(html, "html.parser")
            names = []
            for a in soup.select("a[href*='/research/faculty/']"):
                n = _person_anchor_name(a)
                if n:
                    names.append(n.casefold())
            for h in soup.select("h2,h3,h4"):
                n = clean_person_name_candidate(h.get_text(" ", strip=True))
                if plausible_name(n):
                    names.append(n.casefold())

            signature = tuple(sorted(set(names))[:25])
            if signature and signature in seen_signatures:
                break
            if signature:
                seen_signatures.add(signature)

            html_pages.append(html)

            # Prefer explicit next control.
            next_clicked = False
            selectors = [
                "a[rel='next']",
                "button[aria-label*='Next']",
                "a[aria-label*='Next']",
                ".pagination-next a",
                ".pagination .next a",
            ]
            for selector in selectors:
                try:
                    loc = page.locator(selector).first
                    if await loc.count() and await loc.is_visible():
                        disabled = await loc.get_attribute("disabled")
                        aria_disabled = await loc.get_attribute("aria-disabled")
                        if disabled is not None or aria_disabled == "true":
                            continue
                        await loc.click(timeout=2500)
                        await page.wait_for_timeout(1000)
                        next_clicked = True
                        break
                except Exception:
                    pass

            if not next_clicked:
                # Mayo UI may expose page numbers rather than a literal Next.
                target_text = str(page_index + 1)
                try:
                    loc = page.get_by_role("link", name=target_text, exact=True).last
                    if await loc.count() and await loc.is_visible():
                        await loc.click(timeout=2500)
                        await page.wait_for_timeout(1000)
                        next_clicked = True
                except Exception:
                    pass

            if not next_clicked:
                try:
                    loc = page.get_by_role("button", name=target_text, exact=True).last
                    if await loc.count() and await loc.is_visible():
                        await loc.click(timeout=2500)
                        await page.wait_for_timeout(1000)
                        next_clicked = True
                except Exception:
                    pass

            if not next_clicked:
                break

        return html_pages

    finally:
        await page.close()


async def scrape_mayo_cancer(
    context, source_url: str, profile_sem: asyncio.Semaphore, errors: List[Dict]
) -> List[PersonRecord]:
    # HTTP first page is useful even when browser pagination later fails.
    html_pages = []
    r = await fetch_http(source_url)
    if r["ok"]:
        html_pages.append(r["html"])

    try:
        rendered = await _mayo_rendered_pages(context, source_url)
        # Deduplicate identical HTML snapshots.
        seen_html = set()
        merged_pages = []
        for html in html_pages + rendered:
            sig = hash(html)
            if sig not in seen_html:
                seen_html.add(sig)
                merged_pages.append(html)
        html_pages = merged_pages
    except Exception as exc:
        errors.append({
            "type": "mayo_pagination_warning",
            "source_url": source_url,
            "error": excel_safe(str(exc)),
        })

    if not html_pages:
        errors.append({
            "type": "fetch_failed",
            "source_url": source_url,
            "error": "No Mayo faculty HTML could be fetched",
        })
        return []

    records = []
    for html in html_pages:
        records.extend(_mayo_card_records(html, source_url))

    records = dedupe_records(records)

    # Follow profiles only for rows where the listing itself did not expose email.
    records = await _enrich_cancer_placeholders(
        context, records, profile_sem, errors, source_url
    )

    for rec in records:
        rec.country = rec.country or "United States"
        rec.country_source = rec.country_source or "trusted_source_site"

    return dedupe_records(records)


async def scrape_source(
    context,
    source_url: str,
    llm_enabled: bool,
    llm_sem: asyncio.Semaphore,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    """
    V10 dispatcher.

    These three cancer sites use profile-first extraction because broad page-wide
    email scraping is unsafe.
    """
    print(f"\n   [SOURCE] {source_url}", flush=True)

    if _is_jhmi_cancer(source_url):
        print("   [ADAPTER] Johns Hopkins Radiation Oncology profile crawler", flush=True)
        records = await scrape_jhmi_cancer(context, source_url, profile_sem, errors)

    elif _is_uh_cancer(source_url) and "clinical-faculty" in source_url.casefold():
        print("   [ADAPTER] UH Cancer Center clinical faculty profile crawler", flush=True)
        records = await scrape_uh_cancer(context, source_url, profile_sem, errors)

    elif _is_mayo_cancer(source_url):
        print("   [ADAPTER] Mayo Cancer Center faculty + pagination crawler", flush=True)
        records = await scrape_mayo_cancer(context, source_url, profile_sem, errors)

    else:
        # Safe generic fallback for other university/editorial URLs.
        all_records: List[PersonRecord] = []
        current_url = source_url
        visited = set()

        for page_num in range(1, MAX_PAGES_PER_URL + 1):
            if not current_url or current_url in visited:
                break
            visited.add(current_url)

            r = await fetch_http(current_url)
            method = "http"
            if not r["ok"]:
                r = await playwright_fetch(context, current_url)
                method = "playwright"
            if not r["ok"]:
                errors.append({
                    "type": "fetch_failed",
                    "source_url": source_url,
                    "page_url": current_url,
                    "error": excel_safe(r["error"]),
                })
                break

            actual = r.get("url") or current_url
            page_type, page_records = await parse_html(
                r["html"], actual, llm_enabled, llm_sem
            )
            page_records = await enrich_missing_profiles(
                context, page_records, profile_sem
            )
            page_records = await llm_filter_records(
                page_records, llm_enabled, llm_sem, errors, source_url
            )

            for rec in page_records:
                rec.source_url = source_url
                ok, reason = validate_record(rec)
                if ok:
                    all_records.append(rec)
                else:
                    errors.append({
                        "type": "rejected_record",
                        "reason": reason,
                        "name": excel_safe(rec.name),
                        "email": excel_safe(rec.email),
                        "profile_url": excel_safe(rec.profile_url),
                        "source_url": source_url,
                    })

            if not FOLLOW_PAGINATION or page_type == "EDITORIAL_BOARD":
                break

            nxt = next_page_url(r["html"], actual)
            if not nxt or nxt in visited:
                break
            if _host(nxt) != _host(source_url):
                break
            current_url = nxt
            await polite_delay()

        records = dedupe_records(all_records)

    # Final deterministic validation. Qwen is deliberately NOT required for
    # profile-verified cancer records.
    final = []
    for rec in records:
        rec.source_url = source_url
        ok, reason = validate_record(rec)
        if ok:
            final.append(rec)
        else:
            errors.append({
                "type": "rejected_record",
                "reason": reason,
                "name": excel_safe(rec.name),
                "email": excel_safe(rec.email),
                "country": excel_safe(rec.country),
                "profile_url": excel_safe(rec.profile_url),
                "source_url": source_url,
                "method": rec.extraction_method,
            })

    final = dedupe_records(final)
    print(
        f"   [QUALITY] accepted={len(final)} "
        f"blank_names={sum(1 for x in final if not x.name)}",
        flush=True,
    )
    for rec in final:
        print(
            f"      {(rec.name or '')[:38]:38} | "
            f"{rec.email[:36]:36} | {rec.profile_url[:45]}",
            flush=True,
        )

    return final


# =============================================================================
# V11 ROBUST PSYCHIATRY / UNIVERSITY DIRECTORY OVERRIDES
# =============================================================================
# Adds robust support for:
#   - University of Iowa Psychiatry filtered profile directory
#   - University of Iowa Psychiatry research-team cards
#   - UPenn Psychiatry legacy faculty database -> individual faculty profiles
#
# Also improves the generic profile-crawling path:
#   - listing card -> visible name + profile URL + local email
#   - profile click/fetch -> direct mailto + visible text + JSON-LD + data attrs
#   - Cloudflare data-cfemail
#   - common JS/onclick email encodings
#   - one trusted institutional secondary-profile hop
#   - browser rendering when HTTP returns a shell
#   - Qwen only for genuinely ambiguous mappings
#
# IMPORTANT:
#   Email remains the only mandatory export field.
#   We prefer missing rows to wrong person/email associations, but every
#   discovered person profile is attempted and failures are explicitly logged.

from html import unescape as _html_unescape


V11_PROFILE_CONCURRENCY = 2
V11_MAX_PROFILE_LINKS_PER_SOURCE = 1500
V11_SECONDARY_PROFILE_HOPS = 2

V11_TRUSTED_INSTITUTION_DOMAINS = (
    "uiowa.edu",
    "upenn.edu",
    "pennmedicine.org",
    "pennmedicine.upenn.edu",
    "profiles.upenn.edu",
    "med.upenn.edu",
)

V11_PROFILE_LINK_TEXT_BAD = {
    "profile", "view profile", "read more", "learn more", "more",
    "email", "contact", "website", "personal website", "homepage",
    "research", "publications", "clinical trials", "faculty", "staff",
    "directory", "people", "next", "previous",
}

V11_NONPERSON_PATH_FRAGMENTS = (
    "/news/", "/events/", "/event/", "/search", "/category/", "/tag/",
    "/about/", "/contact/", "/education/", "/research/", "/patient",
    "/giving/", "/donate/", "/privacy", "/terms", "/jobs", "/careers",
)


def _v11_host(url: str) -> str:
    return urlparse(normalize_url(url)).netloc.casefold().replace("www.", "")


def _v11_root_domain(host: str) -> str:
    host = (host or "").casefold().strip(".")
    parts = host.split(".")
    if len(parts) <= 2:
        return host
    # Sufficient for the university domains handled here.
    return ".".join(parts[-2:])


def _v11_same_org_url(a: str, b: str) -> bool:
    ha = _v11_host(a)
    hb = _v11_host(b)
    if not ha or not hb:
        return False
    return ha == hb or _v11_root_domain(ha) == _v11_root_domain(hb)


def _v11_is_uiowa_psychiatry(url: str) -> bool:
    return _v11_host(url) == "psychiatry.medicine.uiowa.edu"


def _v11_is_uiowa_directory(url: str) -> bool:
    p = urlparse(url)
    return _v11_is_uiowa_psychiatry(url) and p.path.rstrip("/").casefold() == "/profile"


def _v11_is_uiowa_research_team(url: str) -> bool:
    return (
        _v11_is_uiowa_psychiatry(url)
        and urlparse(url).path.rstrip("/").casefold() == "/research-team"
    )


def _v11_is_upenn_psych(url: str) -> bool:
    return "med.upenn.edu" in _v11_host(url) and "psychiatry" in url.casefold()


def _v11_is_upenn_faculty_database(url: str) -> bool:
    return (
        _v11_is_upenn_psych(url)
        and urlparse(url).path.casefold().endswith("/psychiatry/faculty_database.html")
    )


def _v11_clean_listing_name(value: str) -> str:
    value = clean_text(value)
    value = re.sub(r"(?i)^\s*profile\s+of\s+", "", value)
    value = re.sub(r"(?i)^\s*(?:photo|picture)\s+of\s+", "", value)
    value = re.sub(r"(?i)\s*[-–—]\s*(?:university of iowa|penn medicine|university of pennsylvania)\s*$", "", value)
    value = clean_person_name_candidate(value)
    return value if plausible_name(value) else ""


def _v11_decode_cfemail(hex_string: str) -> str:
    try:
        raw = bytes.fromhex(clean_text(hex_string))
        if len(raw) < 2:
            return ""
        key = raw[0]
        return normalize_email("".join(chr(x ^ key) for x in raw[1:]))
    except Exception:
        return ""


def _v11_extract_script_emails(text: str) -> List[str]:
    """
    Extract common JavaScript-obfuscated addresses without inventing anything.

    Handles examples such as:
      'john' + '@' + 'uiowa.edu'
      user='john'; domain='upenn.edu'
      john [at] upenn [dot] edu
      john AT som.upenn.edu
    """
    text = _html_unescape(str(text or ""))
    out = []

    # Standard / textual obfuscation.
    out.extend(extract_text_emails(text))

    # Quoted string concatenation.
    for m in re.finditer(
        r"""['"]([A-Z0-9._%+\-]{1,64})['"]\s*\+\s*['"]@['"]\s*\+\s*['"]([A-Z0-9.\-]+\.[A-Z]{2,24})['"]""",
        text,
        re.I,
    ):
        e = normalize_email(f"{m.group(1)}@{m.group(2)}")
        if e:
            out.append(e)

    # user/domain variables in nearby JS.
    users = re.findall(
        r"""(?i)\b(?:user|username|mailbox|emailname)\s*=\s*['"]([A-Z0-9._%+\-]{1,64})['"]""",
        text,
    )
    domains = re.findall(
        r"""(?i)\b(?:domain|host|emaildomain)\s*=\s*['"]([A-Z0-9.\-]+\.[A-Z]{2,24})['"]""",
        text,
    )
    if len(users) == 1 and len(domains) == 1:
        e = normalize_email(f"{users[0]}@{domains[0]}")
        if e:
            out.append(e)

    return unique(out)


def _v11_extract_email_evidence(root: Tag) -> List[Tuple[int, str, Optional[Tag], str]]:
    """
    Comprehensive email evidence collector.

    Returns (score, email, node, evidence_type).
    Higher score means more direct/local evidence.
    """
    found: Dict[str, Tuple[int, Optional[Tag], str]] = {}

    def add(raw, score, node, kind):
        e = normalize_email(raw)
        if not e:
            return
        old = found.get(e)
        if old is None or score > old[0]:
            found[e] = (score, node, kind)

    # Direct mailto (including uppercase and parameters).
    for a in root.select("a[href]"):
        href = clean_text(a.get("href", ""))
        if href.casefold().startswith("mailto:"):
            add(href.split("?", 1)[0], 190, a, "mailto_href")
            visible = normalize_email(a.get_text(" ", strip=True))
            if visible:
                add(visible, 195, a, "mailto_visible")

    # JSON-LD Person.
    for script in root.select('script[type="application/ld+json"]'):
        raw = script.string or script.get_text("", strip=True)
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            continue

        stack = [data]
        while stack:
            obj = stack.pop()
            if isinstance(obj, dict):
                typ = obj.get("@type")
                types = typ if isinstance(typ, list) else [typ]
                if any(str(x).casefold() == "person" for x in types if x):
                    values = obj.get("email", [])
                    values = values if isinstance(values, list) else [values]
                    for v in values:
                        add(v, 200, script, "jsonld_person")
                stack.extend(obj.values())
            elif isinstance(obj, list):
                stack.extend(obj)

    # Data attributes and Cloudflare.
    attrs = (
        "data-email", "data-mail", "data-address", "data-email-address",
        "data-contact-email", "data-user-email",
    )
    for el in root.select("*"):
        for attr in attrs:
            if el.has_attr(attr):
                add(el.get(attr, ""), 175, el, attr)

        if el.has_attr("data-cfemail"):
            e = _v11_decode_cfemail(el.get("data-cfemail", ""))
            if e:
                add(e, 180, el, "cloudflare_cfemail")

        # JavaScript and custom links.
        for attr in ("onclick", "onmouseover", "data-content", "href"):
            val = clean_text(el.get(attr, ""))
            if not val:
                continue
            for e in _v11_extract_script_emails(val):
                add(e, 155, el, f"{attr}_script")

    # Small local text containers.
    selectors = (
        "p, li, td, dd, dt, address, "
        "[class*='email'], [class*='contact'], [id*='email'], [id*='contact'], "
        ".field, .profile-contact, .contact-info"
    )
    for el in root.select(selectors):
        txt = clean_text(el.get_text(" ", strip=True))
        if not txt or len(txt) > 800:
            continue
        for e in extract_text_emails(txt):
            add(e, 125, el, "local_visible_text")
        for e in _v11_extract_script_emails(str(el)):
            add(e, 130, el, "local_html_obfuscation")

    return sorted(
        [(score, e, node, kind) for e, (score, node, kind) in found.items()],
        key=lambda x: x[0],
        reverse=True,
    )


def _v11_extract_profile_name(soup: BeautifulSoup, name_hint: str = "") -> str:
    candidates: List[Tuple[int, str]] = []

    # Listing identity.
    hint = _v11_clean_listing_name(name_hint)
    if hint:
        candidates.append((175, hint))

    # Structured names.
    for person in _jsonld_people(soup):
        n = _v11_clean_listing_name(person.get("name", ""))
        if n:
            candidates.append((210, n))

    for selector, score in (
        ('[itemprop="name"]', 205),
        ("main h1", 200),
        ("article h1", 200),
        (".profile-name", 195),
        (".person-name", 195),
        (".faculty-name", 195),
        (".headline__heading", 190),
        ("h1", 180),
        ("main h2", 170),
        ("article h2", 170),
    ):
        for el in soup.select(selector):
            n = _v11_clean_listing_name(el.get_text(" ", strip=True))
            if n:
                candidates.append((score, n))

    og = soup.select_one('meta[property="og:title"], meta[name="twitter:title"]')
    if og:
        n = _v11_clean_listing_name(og.get("content", ""))
        if n:
            candidates.append((150, n))

    if not candidates:
        return ""

    candidates.sort(reverse=True)

    if hint:
        # Prefer a candidate compatible with the listing name.
        for _, n in candidates:
            if _profile_identity_matches(hint, n):
                return n
        return hint

    return candidates[0][1]


def _v11_local_context(node: Optional[Tag], email: str, limit: int = 1200) -> str:
    if node is None:
        return ""
    try:
        block = _nearest_person_container(node, email, max_chars=limit)
        return clean_text(block.get_text(" ", strip=True))
    except Exception:
        return clean_text(node.get_text(" ", strip=True))[:limit]


def _v11_rank_profile_email(
    name: str,
    evidence: List[Tuple[int, str, Optional[Tag], str]],
) -> Tuple[str, str, int, str]:
    """
    Pick one email from a person's profile page.

    Strong direct profile evidence is accepted even if a university username is
    not lexically similar to the person's name. Footer/general addresses are not.
    """
    ranked = []

    for base, email, node, kind in evidence:
        if not email or is_generic_email(email) or is_strict_generic_email(email):
            continue

        score = base
        name_score = email_name_score(name, email)
        score += min(45, name_score)

        context = _v11_local_context(node, email).casefold()
        person_tokens = _name_tokens(name)

        if person_tokens and any(tok in context for tok in person_tokens):
            score += 30

        if any(x in context for x in (
            "administrative contact", "assistant to", "media contact",
            "press contact", "general inquiries", "general information",
            "webmaster", "development office", "employment opportunities",
        )):
            score -= 140

        # Direct profile mailto/JSON-LD is intrinsically strong.
        if kind in {"mailto_href", "mailto_visible", "jsonld_person"}:
            score += 15

        ranked.append((score, email, kind, context))

    if not ranked:
        return "", "", 0, ""

    ranked.sort(reverse=True)
    score, email, kind, context = ranked[0]

    # Require strong profile evidence.
    if score < 150:
        return "", "", score, kind

    return email, kind, score, context


def _v11_extract_profile_metadata(
    soup: BeautifulSoup,
    root: Tag,
    name: str,
    email: str,
    email_context: str,
) -> Dict[str, str]:
    data = {
        "academic_title": "",
        "academic_rank": "",
        "affiliation": "",
        "university": "",
        "faculty": "",
        "school": "",
        "department": "",
        "institute": "",
        "division": "",
        "city": "",
        "address": "",
        "phone": "",
    }

    # Titles directly below/near profile name.
    for sel in (
        ".profile-title", ".person-title", ".faculty-title", ".title",
        ".field--name-field-person-title", ".subtitle", "main h2", "article h2"
    ):
        el = soup.select_one(sel)
        if el:
            t = clean_text(el.get_text(" ", strip=True))
            if t and len(t) <= 180 and t.casefold() != name.casefold():
                data["academic_title"] = t
                break

    # Contact/address.
    for sel in (
        "address", ".address", ".contact-address", "[itemprop='address']",
        ".profile-contact", ".contact-info"
    ):
        el = soup.select_one(sel)
        if el:
            t = clean_text(el.get_text(" ", strip=True))
            if t and len(t) <= 600:
                data["address"] = t
                break

    data["phone"] = extract_phone(root)

    # Local contact/affiliation, not full biography.
    local = clean_text(email_context)
    if local:
        local = re.sub(re.escape(email), " ", local, flags=re.I)
        if name:
            local = re.sub(re.escape(name), " ", local, flags=re.I)
        local = clean_text(local)
        if len(local) <= 700:
            data["affiliation"] = local

    return data


def extract_profile_record_v11(
    html: str,
    profile_url: str,
    source_url: str,
    name_hint: str = "",
) -> Optional[PersonRecord]:
    soup = BeautifulSoup(html, "html.parser")
    root = (
        soup.select_one("main")
        or soup.select_one("article")
        or soup.select_one("#main-content")
        or soup.select_one("#content")
        or soup.select_one(".main-content")
        or soup.body
        or soup
    )

    name = _v11_extract_profile_name(soup, name_hint)
    if name_hint and name and not _profile_identity_matches(name_hint, name):
        # Listing identity and profile identity disagree: do not cross-map.
        return None

    evidence = _v11_extract_email_evidence(root)
    email, email_kind, email_score, email_context = _v11_rank_profile_email(
        name, evidence
    )
    if not email:
        return None

    metadata = _v11_extract_profile_metadata(
        soup, root, name, email, email_context
    )
    links = extract_links(root, profile_url)

    country = ""
    country_source = ""

    # Country from email ccTLD or explicit short local/address evidence only.
    c = country_from_tld(email)
    if c:
        country = c
        country_source = "email_tld"
    elif metadata["address"]:
        c = extract_country_from_text(metadata["address"])
        if c:
            country = c
            country_source = "profile_address"

    # Known US institutions.
    h = _v11_host(profile_url)
    if not country and any(
        h == d or h.endswith("." + d)
        for d in ("uiowa.edu", "upenn.edu", "pennmedicine.org")
    ):
        country = "United States"
        country_source = "trusted_institution_domain"

    rec = PersonRecord(
        name=name if plausible_name(name) else "",
        email=email,
        country=country,
        country_source=country_source,
        alternate_emails="",
        email_type="personal",
        email_conflict="no",
        page_type="UNIVERSITY_DIRECTORY",
        academic_title=metadata["academic_title"],
        academic_rank=metadata["academic_rank"],
        affiliation=metadata["affiliation"],
        university=metadata["university"],
        faculty=metadata["faculty"],
        school=metadata["school"],
        department=metadata["department"],
        institute=metadata["institute"],
        division=metadata["division"],
        city=metadata["city"],
        address=metadata["address"],
        phone=metadata["phone"],
        orcid=links.get("orcid", ""),
        google_scholar=links.get("google_scholar", ""),
        pubmed=links.get("pubmed", ""),
        profile_url=profile_url,
        personal_homepage=links.get("personal_homepage", ""),
        source_url=source_url,
        confidence=min(100, max(92, email_score)),
        extraction_method=f"profile_v11:{email_kind}",
        scrape_status="accepted",
    )
    return rec


def _v11_anchor_is_person_profile(
    anchor: Tag,
    source_url: str,
) -> Tuple[int, str, str]:
    href = clean_text(anchor.get("href", ""))
    if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
        return 0, "", ""

    url = normalize_url(urljoin(source_url, href))
    if not url.startswith(("http://", "https://")):
        return 0, "", ""

    path = urlparse(url).path.casefold()
    host = _v11_host(url)

    name = _v11_clean_listing_name(anchor.get_text(" ", strip=True))
    if not name:
        # Iowa images often carry exact person name.
        img = anchor.select_one("img[alt]")
        if img:
            name = _v11_clean_listing_name(img.get("alt", ""))
        if not name:
            name = _v11_clean_listing_name(
                anchor.get("aria-label", "") or anchor.get("title", "")
            )
    if not name:
        return 0, "", ""

    score = 0

    # Strong profile patterns.
    if re.search(r"/profile/[^/?#]+/?$", path):
        score += 100
    if re.search(r"/profiles?/[^/?#]+/?$", path):
        score += 90
    if "/apps/faculty/index.php/" in path:
        score += 120
    if "/faculty/" in path and len([x for x in path.split("/") if x]) >= 2:
        score += 55
    if any(x in path for x in ("/people/", "/person/", "/staff/")):
        score += 55

    # Same institution is preferred.
    if _v11_same_org_url(url, source_url):
        score += 30
    elif any(host == d or host.endswith("." + d) for d in V11_TRUSTED_INSTITUTION_DOMAINS):
        score += 20

    # Exclude obvious non-person content.
    if any(x in path for x in V11_NONPERSON_PATH_FRAGMENTS):
        score -= 120

    # Human-name anchor itself is evidence.
    if len(_name_tokens(name)) >= 2:
        score += 25

    return score, url, name


def discover_uiowa_directory_records(
    html: str,
    source_url: str,
) -> List[PersonRecord]:
    """
    University of Iowa filtered directory.

    The rendered listing can already contain:
      h2 person name
      /profile/<slug>
      direct mailto

    We capture email locally if present, AND keep profile placeholders for
    anyone without a listing email.
    """
    soup = BeautifulSoup(html, "html.parser")
    records = []

    cards = soup.select(
        ".directory-listing li, "
        ".directory-listing .card, "
        "ol.grid > li"
    )

    for card in cards:
        name_el = card.select_one(
            "h2 a, h2 .headline__heading, .card-body h2 a, .headline__heading"
        )
        name = _v11_clean_listing_name(
            name_el.get_text(" ", strip=True) if name_el else ""
        )

        profile_url = ""
        for a in card.select("a[href]"):
            score, url, candidate_name = _v11_anchor_is_person_profile(
                a, source_url
            )
            if score >= 100:
                profile_url = url
                if not name:
                    name = candidate_name
                break

        evidence = _v11_extract_email_evidence(card)
        email, kind, score, context = _v11_rank_profile_email(name, evidence)

        if email:
            rec = PersonRecord(
                name=name,
                email=email,
                country="United States",
                country_source="trusted_institution_domain",
                page_type="UNIVERSITY_DIRECTORY",
                profile_url=profile_url,
                source_url=source_url,
                confidence=min(100, max(96, score)),
                extraction_method=f"uiowa_directory_card:{kind}",
            )
            title = card.select_one("h3")
            if title:
                rec.academic_title = clean_text(
                    title.get_text(" ", strip=True)
                )[:180]
            records.append(rec)
        elif profile_url:
            records.append(PersonRecord(
                name=name,
                profile_url=profile_url,
                source_url=source_url,
                page_type="UNIVERSITY_DIRECTORY",
                country="United States",
                country_source="trusted_institution_domain",
                confidence=90,
                extraction_method="uiowa_directory_profile_placeholder",
            ))

    return dedupe_records(records)


def discover_uiowa_research_team_records(
    html: str,
    source_url: str,
) -> List[PersonRecord]:
    soup = BeautifulSoup(html, "html.parser")
    records = []

    cards = soup.select(
        ".block-inline-blockuiowa-card .card, "
        ".uiowa-card .card, "
        ".card--layout-left, "
        ".click-container.card"
    )

    seen_card_text = set()
    for card in cards:
        signature = clean_text(card.get_text(" ", strip=True))[:600]
        if not signature or signature in seen_card_text:
            continue
        seen_card_text.add(signature)

        heading = card.select_one(
            "h2 .headline__heading, h2 a, .headline__heading, h3 a"
        )
        name = _v11_clean_listing_name(
            heading.get_text(" ", strip=True) if heading else ""
        )
        if not name:
            img = card.select_one("img[alt]")
            if img:
                name = _v11_clean_listing_name(img.get("alt", ""))
        if not name:
            continue

        # Local email if available.
        evidence = _v11_extract_email_evidence(card)
        email, kind, score, _ = _v11_rank_profile_email(name, evidence)

        profile_candidates = []
        for a in card.select("a[href]"):
            href = normalize_url(urljoin(source_url, a.get("href", "")))
            if not href.startswith(("http://", "https://")):
                continue

            anchor_name = _v11_clean_listing_name(
                a.get_text(" ", strip=True)
            )
            path = urlparse(href).path.casefold()
            host = _v11_host(href)

            pscore = 0
            if re.search(r"/profile/[^/?#]+/?$", path):
                pscore += 120
            if _v11_same_org_url(href, source_url):
                pscore += 30
            if host.endswith("uiowa.edu"):
                pscore += 25
            # External Iowa lab sites are still authoritative.
            if host.endswith("uiowa.edu") and host != _v11_host(source_url):
                pscore += 35
            if anchor_name and _profile_identity_matches(name, anchor_name):
                pscore += 25

            if pscore >= 50:
                profile_candidates.append((pscore, href))

        profile_candidates.sort(reverse=True)
        profile_url = profile_candidates[0][1] if profile_candidates else ""

        if email:
            records.append(PersonRecord(
                name=name,
                email=email,
                country="United States",
                country_source="trusted_institution_domain",
                page_type="UNIVERSITY_DIRECTORY",
                profile_url=profile_url,
                source_url=source_url,
                confidence=min(100, max(94, score)),
                extraction_method=f"uiowa_research_team_card:{kind}",
            ))
        elif profile_url:
            records.append(PersonRecord(
                name=name,
                country="United States",
                country_source="trusted_institution_domain",
                page_type="UNIVERSITY_DIRECTORY",
                profile_url=profile_url,
                source_url=source_url,
                confidence=88,
                extraction_method="uiowa_research_team_profile_placeholder",
            ))

    return dedupe_records(records)


def discover_upenn_faculty_records(
    html: str,
    source_url: str,
) -> List[PersonRecord]:
    """
    UPenn legacy table:
      <td class="member_name">
        <a href="/apps/faculty/index.php/g332/p8527978">Richa Aggarwal Dutta</a>
      </td>

    Every faculty link is retained as a placeholder and individually fetched.
    """
    soup = BeautifulSoup(html, "html.parser")
    records = []
    seen = set()

    selectors = (
        "td.member_name a[href], "
        "table a[href*='/apps/faculty/index.php/'], "
        "a[href*='/apps/faculty/index.php/']"
    )

    for a in soup.select(selectors):
        name = _v11_clean_listing_name(a.get_text(" ", strip=True))
        href = normalize_url(urljoin(source_url, a.get("href", "")))
        if not name or not href:
            continue
        if "/apps/faculty/index.php/" not in urlparse(href).path.casefold():
            continue

        key = href.casefold()
        if key in seen:
            continue
        seen.add(key)

        records.append(PersonRecord(
            name=name,
            profile_url=href,
            source_url=source_url,
            page_type="UNIVERSITY_DIRECTORY",
            country="United States",
            country_source="trusted_institution_domain",
            confidence=94,
            extraction_method="upenn_faculty_database_placeholder",
        ))

    return records[:V11_MAX_PROFILE_LINKS_PER_SOURCE]


def discover_generic_person_profiles_v11(
    html: str,
    source_url: str,
) -> List[PersonRecord]:
    """
    Conservative generic person-profile discovery.

    This deliberately operates on repeated cards/list rows and person-looking
    anchors instead of following every site link.
    """
    soup = BeautifulSoup(html, "html.parser")
    records = []
    seen = set()

    # First structured containers.
    containers = soup.select(
        "article, li, tr, .card, .person, .staff-member, .faculty-member, "
        ".team-member, .profile-card, .directory-item, [class*='person-card']"
    )

    for block in containers:
        text = clean_text(block.get_text(" ", strip=True))
        if not text or len(text) > 2500:
            continue

        # Direct local email can create a record without profile click.
        mailto = block.select_one("a[href^='mailto:'],a[href^='MAILTO:']")
        local_email = normalize_email(mailto.get("href", "")) if mailto else ""

        best = (0, "", "")
        for a in block.select("a[href]"):
            candidate = _v11_anchor_is_person_profile(a, source_url)
            if candidate[0] > best[0]:
                best = candidate

        profile_score, profile_url, name = best

        if not name:
            candidates = collect_name_candidates(block, local_email)
            if candidates:
                name = _v11_clean_listing_name(candidates[0][1])

        if local_email and name and not is_generic_email(local_email):
            records.append(PersonRecord(
                name=name,
                email=local_email,
                profile_url=profile_url,
                source_url=source_url,
                page_type="UNIVERSITY_DIRECTORY",
                confidence=95 if profile_url else 88,
                extraction_method="generic_v11_local_mailto",
            ))

        if profile_score >= 100 and profile_url and name:
            key = profile_url.casefold()
            if key not in seen:
                seen.add(key)
                records.append(PersonRecord(
                    name=name,
                    profile_url=profile_url,
                    source_url=source_url,
                    page_type="UNIVERSITY_DIRECTORY",
                    confidence=min(95, 55 + profile_score // 2),
                    extraction_method="generic_v11_profile_placeholder",
                ))

    # Then strong standalone profile links.
    for a in soup.select("a[href]"):
        score, profile_url, name = _v11_anchor_is_person_profile(a, source_url)
        if score < 125 or not profile_url or not name:
            continue
        key = profile_url.casefold()
        if key in seen:
            continue
        seen.add(key)
        records.append(PersonRecord(
            name=name,
            profile_url=profile_url,
            source_url=source_url,
            page_type="UNIVERSITY_DIRECTORY",
            confidence=min(95, 55 + score // 2),
            extraction_method="generic_v11_profile_placeholder",
        ))
        if len(records) >= V11_MAX_PROFILE_LINKS_PER_SOURCE:
            break

    return dedupe_records(records)


async def _v11_fetch_html(context, url: str, force_browser: bool = False) -> Dict:
    if not force_browser:
        r = await fetch_http(url)
        if r["ok"]:
            return r
    return await playwright_fetch(context, url)


def _v11_secondary_profile_urls(
    html: str,
    current_url: str,
    person_name: str,
) -> List[str]:
    """
    Trusted one-hop profile/lab/directory continuation.

    Useful when a research-team card links to an institutional lab page rather
    than the department's /profile page.
    """
    soup = BeautifulSoup(html, "html.parser")
    scored = []

    for a in soup.select("a[href]"):
        href = normalize_url(urljoin(current_url, a.get("href", "")))
        if not href.startswith(("http://", "https://")):
            continue
        if href.rstrip("/") == current_url.rstrip("/"):
            continue

        host = _v11_host(href)
        if not any(host == d or host.endswith("." + d) for d in V11_TRUSTED_INSTITUTION_DOMAINS):
            continue

        text = _v11_clean_listing_name(a.get_text(" ", strip=True))
        path = urlparse(href).path.casefold()
        score = 0

        if re.search(r"/profile/[^/?#]+/?$", path):
            score += 100
        if "/apps/faculty/index.php/" in path:
            score += 100
        if any(x in path for x in ("/people/", "/person/", "/faculty/", "/staff/")):
            score += 50
        if text and _profile_identity_matches(person_name, text):
            score += 50
        if _v11_same_org_url(href, current_url):
            score += 15

        if score >= 60:
            scored.append((score, href))

    out = []
    seen = set()
    for _, url in sorted(scored, reverse=True):
        if url not in seen:
            seen.add(url)
            out.append(url)
    return out[:V11_SECONDARY_PROFILE_HOPS]


async def enrich_profile_placeholders_v11(
    context,
    records: List[PersonRecord],
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
    source_url: str,
) -> List[PersonRecord]:
    """
    Keep direct email records and fetch every distinct email-less profile.

    Crucially, a profile failure is logged with the person's listing name and
    URL so missing coverage is visible instead of silently disappearing.
    """
    direct = [r for r in records if normalize_email(r.email)]
    placeholders = [
        r for r in records
        if not normalize_email(r.email) and normalize_url(r.profile_url)
    ]

    unique = []
    seen = set()
    for r in placeholders:
        key = normalize_url(r.profile_url).casefold()
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(r)

    async def one(placeholder: PersonRecord):
        async with profile_sem:
            # UIowa filtered listing is JS rendered. Person profile itself is
            # normally server-rendered, but browser fallback remains available.
            r = await _v11_fetch_html(context, placeholder.profile_url)
            if not r["ok"]:
                return None, "profile_fetch_failed"

            final_url = r.get("url") or placeholder.profile_url
            rec = extract_profile_record_v11(
                r["html"], final_url, source_url, placeholder.name
            )
            if rec and rec.email:
                return rec, ""

            # Try trusted secondary institutional profile/lab links.
            for second in _v11_secondary_profile_urls(
                r["html"], final_url, placeholder.name
            ):
                r2 = await _v11_fetch_html(context, second)
                if not r2["ok"]:
                    continue
                final2 = r2.get("url") or second
                rec = extract_profile_record_v11(
                    r2["html"], final2, source_url, placeholder.name
                )
                if rec and rec.email:
                    # Keep the authoritative contact page.
                    rec.profile_url = final2
                    return rec, ""

            return None, "profile_no_verified_email"

    results = await asyncio.gather(*(one(r) for r in unique))

    for placeholder, (rec, reason) in zip(unique, results):
        if rec and rec.email:
            direct.append(rec)
        else:
            errors.append({
                "type": reason,
                "source_url": source_url,
                "name": excel_safe(placeholder.name),
                "profile_url": excel_safe(placeholder.profile_url),
            })

    return dedupe_records(direct)


async def scrape_uiowa_directory_v11(
    context,
    source_url: str,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    # Force browser because the user-supplied HTML demonstrates the populated
    # directory is client-rendered.
    r = await _v11_fetch_html(context, source_url, force_browser=True)
    if not r["ok"]:
        errors.append({
            "type": "fetch_failed",
            "source_url": source_url,
            "error": excel_safe(r["error"]),
        })
        return []

    records = discover_uiowa_directory_records(r["html"], source_url)

    # The listing itself has mailto for many/all faculty. Profile enrichment is
    # still attempted for records lacking email.
    records = await enrich_profile_placeholders_v11(
        context, records, profile_sem, errors, source_url
    )

    return dedupe_records(records)


async def scrape_uiowa_research_team_v11(
    context,
    source_url: str,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    r = await _v11_fetch_html(context, source_url, force_browser=True)
    if not r["ok"]:
        errors.append({
            "type": "fetch_failed",
            "source_url": source_url,
            "error": excel_safe(r["error"]),
        })
        return []

    records = discover_uiowa_research_team_records(r["html"], source_url)
    records = await enrich_profile_placeholders_v11(
        context, records, profile_sem, errors, source_url
    )
    return dedupe_records(records)


async def scrape_upenn_faculty_v11(
    context,
    source_url: str,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    r = await _v11_fetch_html(context, source_url)
    if not r["ok"]:
        errors.append({
            "type": "fetch_failed",
            "source_url": source_url,
            "error": excel_safe(r["error"]),
        })
        return []

    placeholders = discover_upenn_faculty_records(r["html"], source_url)
    if not placeholders:
        # Some legacy pages can behave differently in a browser.
        br = await _v11_fetch_html(context, source_url, force_browser=True)
        if br["ok"]:
            placeholders = discover_upenn_faculty_records(
                br["html"], source_url
            )

    errors.append({
        "type": "source_discovery_summary",
        "source_url": source_url,
        "profiles_discovered": len(placeholders),
    })

    return await enrich_profile_placeholders_v11(
        context, placeholders, profile_sem, errors, source_url
    )


async def scrape_generic_v11(
    context,
    source_url: str,
    llm_enabled: bool,
    llm_sem: asyncio.Semaphore,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    """
    More robust generic pipeline:
      HTTP -> parse direct structured records/profile links
           -> if weak/empty, browser render
           -> follow discovered profiles
           -> deterministic validation
           -> Qwen only for ambiguous non-profile records
    """
    r = await _v11_fetch_html(context, source_url)
    if not r["ok"]:
        errors.append({
            "type": "fetch_failed",
            "source_url": source_url,
            "error": excel_safe(r["error"]),
        })
        return []

    actual = r.get("url") or source_url
    page_type, existing = await parse_html(
        r["html"], actual, llm_enabled, llm_sem
    )

    profiles = []
    if page_type != "EDITORIAL_BOARD":
        profiles = discover_generic_person_profiles_v11(
            r["html"], actual
        )

    combined = dedupe_records(existing + profiles)

    # If no profile/direct records, browser-render before giving up.
    if page_type != "EDITORIAL_BOARD" and not combined:
        br = await _v11_fetch_html(context, source_url, force_browser=True)
        if br["ok"]:
            actual = br.get("url") or source_url
            _, existing2 = await parse_html(
                br["html"], actual, llm_enabled, llm_sem
            )
            profiles2 = discover_generic_person_profiles_v11(
                br["html"], actual
            )
            combined = dedupe_records(existing2 + profiles2)

    # Follow email-less profiles.
    combined = await enrich_profile_placeholders_v11(
        context, combined, profile_sem, errors, source_url
    )

    # Only remaining direct/ambiguous records need the Qwen gate.
    profile_verified = [
        r for r in combined
        if "profile_v11:" in (r.extraction_method or "")
    ]
    other = [
        r for r in combined
        if "profile_v11:" not in (r.extraction_method or "")
    ]

    if other:
        other = await llm_filter_records(
            other, llm_enabled, llm_sem, errors, source_url
        )

    return dedupe_records(profile_verified + other)


async def scrape_source(
    context,
    source_url: str,
    llm_enabled: bool,
    llm_sem: asyncio.Semaphore,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    """
    V11 final dispatcher.

    Existing cancer-specific v10 adapters remain available, plus the new Iowa
    and UPenn psychiatry adapters. All other university sources use the stronger
    generic profile-first pipeline.
    """
    print(f"\n   [SOURCE] {source_url}", flush=True)

    # ---- Existing cancer-specific adapters ---------------------------------
    if _is_jhmi_cancer(source_url):
        print("   [ADAPTER] Johns Hopkins Radiation Oncology", flush=True)
        records = await scrape_jhmi_cancer(
            context, source_url, profile_sem, errors
        )

    elif _is_uh_cancer(source_url) and "clinical-faculty" in source_url.casefold():
        print("   [ADAPTER] UH Cancer Center clinical faculty", flush=True)
        records = await scrape_uh_cancer(
            context, source_url, profile_sem, errors
        )

    elif _is_mayo_cancer(source_url):
        print("   [ADAPTER] Mayo Cancer Center faculty", flush=True)
        records = await scrape_mayo_cancer(
            context, source_url, profile_sem, errors
        )

    # ---- New psychiatry adapters -------------------------------------------
    elif _v11_is_uiowa_directory(source_url):
        print(
            "   [ADAPTER] University of Iowa Psychiatry directory "
            "(rendered cards + mailto + profiles)",
            flush=True,
        )
        records = await scrape_uiowa_directory_v11(
            context, source_url, profile_sem, errors
        )

    elif _v11_is_uiowa_research_team(source_url):
        print(
            "   [ADAPTER] University of Iowa Psychiatry research team "
            "(cards + institutional profile/lab follow)",
            flush=True,
        )
        records = await scrape_uiowa_research_team_v11(
            context, source_url, profile_sem, errors
        )

    elif _v11_is_upenn_faculty_database(source_url):
        print(
            "   [ADAPTER] UPenn Psychiatry faculty database "
            "(all table profile links + individual profile fetch)",
            flush=True,
        )
        records = await scrape_upenn_faculty_v11(
            context, source_url, profile_sem, errors
        )

    else:
        print(
            "   [ADAPTER] Robust generic profile-first university parser",
            flush=True,
        )
        records = await scrape_generic_v11(
            context,
            source_url,
            llm_enabled,
            llm_sem,
            profile_sem,
            errors,
        )

    # ---- Final deterministic validation ------------------------------------
    final = []
    for rec in dedupe_records(records):
        rec.source_url = source_url
        ok, reason = validate_record(rec)
        if ok:
            final.append(rec)
        else:
            errors.append({
                "type": "rejected_record",
                "reason": reason,
                "name": excel_safe(rec.name),
                "email": excel_safe(rec.email),
                "country": excel_safe(rec.country),
                "profile_url": excel_safe(rec.profile_url),
                "source_url": source_url,
                "method": excel_safe(rec.extraction_method),
            })

    final = dedupe_records(final)

    print(
        f"   [QUALITY] accepted={len(final)} "
        f"blank_names={sum(1 for x in final if not x.name)}",
        flush=True,
    )
    for rec in final:
        print(
            f"      {(rec.name or '')[:40]:40} | "
            f"{rec.email[:38]:38} | "
            f"{(rec.profile_url or '')[:55]}",
            flush=True,
        )

    if not final:
        errors.append({
            "type": "no_verified_people",
            "source_url": source_url,
            "message": (
                "No verified email records were produced. Review profile_* "
                "diagnostics in errors.json for the exact missed profiles."
            ),
        })

    return final


# =============================================================================
# V12 UNIVERSAL EMAIL-FIRST SAFETY NET
# =============================================================================
# Why this exists:
#   Some pages do NOT put the person's name and email in the same DOM card.
#   Example: Columbia COAP has a "Location and Contact Information" block
#   containing kidpower@nys... and a separate "Principal Investigator" block.
#
# Previous versions tried too hard to create a person mapping, so perfectly
# valid page emails could disappear. V12 changes the invariant:
#
#   1. NEVER miss a valid email found in HTML/rendered DOM/profile.
#   2. Name is optional.
#   3. Only attach a name when the name/email association is strong.
#   4. If association is ambiguous, export the email with name="".
#   5. Strong profile/person rows still take precedence over email-only rows.
#
# This makes the generic scraper work on simple contact pages as well as
# structured faculty directories without manufacturing identities.

V12_EMAIL_FIRST = True
V12_RENDER_WHEN_EMAIL_COVERAGE_LOW = True
V12_MAX_EMAILS_PER_PAGE = 5000

V12_CONTACT_SECTION_LABELS = {
    "location and contact information",
    "contact information",
    "contact us",
    "general information",
    "general inquiries",
    "program contact",
    "project contact",
    "lab contact",
    "research contact",
    "administrative contact",
    "media contact",
}

V12_PERSON_SECTION_LABELS = {
    "principal investigator",
    "co-principal investigator",
    "investigator",
    "faculty",
    "researcher",
    "research team",
    "team member",
    "staff",
}


def _v12_all_valid_page_emails(soup: BeautifulSoup) -> List[str]:
    """
    Exhaustive page-level email inventory.

    Includes all mechanisms already supported by V11 plus raw document text.
    This is used for coverage accounting and the final email-only safety net.
    """
    emails = []

    root = soup.body or soup
    for _, email, _, _ in _v11_extract_email_evidence(root):
        e = normalize_email(email)
        if e:
            emails.append(e)

    # Raw HTML catches addresses hidden in scripts/data that DOM selectors miss.
    raw = str(soup)
    emails.extend(_v11_extract_script_emails(raw))
    emails.extend(extract_text_emails(raw))

    return unique(emails)[:V12_MAX_EMAILS_PER_PAGE]


def _v12_heading_context(node: Optional[Tag]) -> List[str]:
    out = []
    if node is None:
        return out
    try:
        cur = node
        for _ in range(5):
            if not isinstance(cur, Tag):
                break

            # Same-container headings.
            for h in cur.select("h1,h2,h3,h4,h5,h6,legend,th"):
                t = clean_text(h.get_text(" ", strip=True))
                if t and len(t) <= 160:
                    out.append(t)

            # Nearby previous headings.
            for h in cur.find_all_previous(
                ["h1", "h2", "h3", "h4", "h5", "h6", "legend", "th"],
                limit=4,
            ):
                t = clean_text(h.get_text(" ", strip=True))
                if t and len(t) <= 160:
                    out.append(t)

            cur = cur.parent
    except Exception:
        pass
    return unique(out)


def _v12_is_contact_context(node: Optional[Tag]) -> bool:
    for heading in _v12_heading_context(node):
        low = heading.casefold()
        if any(label in low for label in V12_CONTACT_SECTION_LABELS):
            return True
    return False


def _v12_is_person_context(node: Optional[Tag]) -> bool:
    for heading in _v12_heading_context(node):
        low = heading.casefold()
        if any(label in low for label in V12_PERSON_SECTION_LABELS):
            return True
    return False


def _v12_safe_name_for_email(
    email: str,
    node: Optional[Tag],
    source_url: str,
) -> Tuple[str, int, str]:
    """
    Return (name, confidence, reason).

    A name is attached only when evidence is local and strong.
    Otherwise return blank name; the email is still exported.
    """
    if node is None:
        return "", 0, "no_local_node"

    try:
        block = _nearest_person_container(node, email, max_chars=1400)
    except Exception:
        block = node

    candidates = collect_name_candidates(block, email)
    if not candidates:
        return "", 0, "no_name_candidate"

    # collect_name_candidates returns tuples (score, text)
    top_score, top_name = candidates[0]
    top_name = _v11_clean_listing_name(top_name)
    if not top_name:
        return "", 0, "invalid_name_candidate"

    direct_match = email_name_score(top_name, email)

    # Explicit profile/person context is strong.
    profile_link = ""
    profile_score = 0
    try:
        for a in block.select("a[href]"):
            ps, pu, pn = _v11_anchor_is_person_profile(a, source_url)
            if ps > profile_score and (
                not pn or _profile_identity_matches(top_name, pn)
            ):
                profile_score = ps
                profile_link = pu
    except Exception:
        pass

    # Critical anti-cross-column rule:
    # a contact-information mailbox next to a PI column must NOT be assigned
    # to the PI merely because both share a parent card.
    if _v12_is_contact_context(node) and not _v12_is_person_context(node):
        if direct_match < 40 and profile_score < 100:
            return "", 0, "contact_section_not_person_specific"

    # Strong lexical email/name match.
    if direct_match >= 40:
        return top_name, 98, "strong_name_email_match"

    # Profile-local block with one email and one credible human.
    local_emails = []
    try:
        local_emails = [
            e for _, e, _, _ in _v11_extract_email_evidence(block)
            if normalize_email(e)
        ]
    except Exception:
        pass
    local_emails = unique(local_emails)

    credible_names = []
    for score, name in candidates[:8]:
        n = _v11_clean_listing_name(name)
        if n and n not in credible_names:
            credible_names.append(n)

    if (
        profile_score >= 100
        and len(local_emails) == 1
        and len(credible_names) == 1
    ):
        return top_name, 94, "single_person_profile_block"

    # Same tight card/list item with exactly one human and one email.
    if (
        len(clean_text(block.get_text(" ", strip=True))) <= 900
        and len(local_emails) == 1
        and len(credible_names) == 1
        and not _v12_is_contact_context(node)
    ):
        return top_name, 90, "single_person_single_email_card"

    return "", 0, "ambiguous_name_email_association"


def harvest_email_first_records_v12(
    html: str,
    source_url: str,
) -> List[PersonRecord]:
    """
    Build one safe record for EVERY valid email visible/embedded on the page.

    If a name cannot be safely mapped, it remains blank.
    """
    soup = BeautifulSoup(html, "html.parser")
    root = soup.body or soup

    evidence = _v11_extract_email_evidence(root)

    # Ensure raw-only emails are also represented.
    evidence_by_email: Dict[str, Tuple[int, str, Optional[Tag], str]] = {}
    for score, email, node, kind in evidence:
        e = normalize_email(email)
        if not e:
            continue
        old = evidence_by_email.get(e)
        if old is None or score > old[0]:
            evidence_by_email[e] = (score, e, node, kind)

    for e in _v12_all_valid_page_emails(soup):
        if e not in evidence_by_email:
            evidence_by_email[e] = (90, e, None, "raw_page_email")

    records = []

    for email, (base, _, node, kind) in evidence_by_email.items():
        email = normalize_email(email)
        if not email:
            continue

        name, name_conf, name_reason = _v12_safe_name_for_email(
            email, node, source_url
        )

        # Generic/shared emails are still real emails. Since email is the only
        # mandatory field, preserve them as email-only records rather than
        # falsely assigning them to a person.
        generic = is_generic_email(email) or is_strict_generic_email(email)
        if generic:
            name = ""
            name_conf = 0
            name_reason = "shared_or_generic_email_preserved_without_person"

        local_text = _v11_local_context(node, email, 900)
        affiliation = ""
        if local_text:
            temp = BeautifulSoup(f"<div>{local_text}</div>", "html.parser").div
            affiliation = split_affiliation_from_block(
                temp, name, [email], ""
            )
            affiliation = clean_text(affiliation)[:700]

        country = country_from_tld(email)
        country_source = "email_tld" if country else ""

        rec = PersonRecord(
            name=name,
            email=email,
            country=country,
            country_source=country_source,
            alternate_emails="",
            email_type="shared/role" if generic else "personal",
            email_conflict="no",
            page_type="UNIVERSITY_DIRECTORY",
            affiliation=affiliation,
            source_url=source_url,
            confidence=max(75, min(100, base // 2 + name_conf // 2)),
            extraction_method=f"email_first_v12:{kind}:{name_reason}",
            scrape_status="accepted",
        )
        records.append(rec)

    return dedupe_records(records)


def merge_email_coverage_v12(
    records: List[PersonRecord],
    email_first: List[PersonRecord],
) -> List[PersonRecord]:
    """
    Merge by email while preferring strong person/profile records.

    This guarantees all discovered emails survive without downgrading a good
    named/profile row to an email-only row.
    """
    by_email: Dict[str, PersonRecord] = {}

    def quality(r: PersonRecord) -> int:
        q = int(r.confidence or 0)
        if r.name and plausible_name(r.name):
            q += 50
        if r.profile_url:
            q += 40
        if "profile_v11:" in (r.extraction_method or ""):
            q += 50
        if "email_first_v12:" in (r.extraction_method or ""):
            q -= 10
        return q

    for rec in records + email_first:
        e = normalize_email(rec.email)
        if not e:
            continue
        rec.email = e

        if e not in by_email:
            by_email[e] = rec
            continue

        cur = by_email[e]
        if quality(rec) > quality(cur):
            best, other = rec, cur
        else:
            best, other = cur, rec

        # Fill only missing metadata from the weaker observation.
        for field_name in COLUMNS:
            if not hasattr(best, field_name) or not hasattr(other, field_name):
                continue
            if field_name in {"email", "confidence", "extraction_method"}:
                continue
            if not getattr(best, field_name) and getattr(other, field_name):
                setattr(best, field_name, getattr(other, field_name))

        best.confidence = max(
            int(best.confidence or 0),
            int(other.confidence or 0),
        )

        methods = unique([
            x for x in (
                best.extraction_method,
                other.extraction_method,
            ) if x
        ])
        best.extraction_method = "+".join(methods)

        by_email[e] = best

    return list(by_email.values())


def validate_record_v12(r: PersonRecord) -> Tuple[bool, str]:
    """
    V12 validation honors the actual business rule:
      email mandatory; name/country optional.

    A valid shared/generic email is NOT discarded. It is exported with blank
    name unless there is strong person-specific evidence.
    """
    r.email = normalize_email(r.email)
    if not r.email:
        return False, "missing_or_invalid_email"

    # Never manufacture a person identity for a shared mailbox.
    if is_generic_email(r.email) or is_strict_generic_email(r.email):
        if "profile_v11:" not in (r.extraction_method or ""):
            r.name = ""
        r.email_type = "shared/role"

    if r.name:
        cleaned = _v11_clean_listing_name(r.name)
        r.name = cleaned if cleaned and plausible_name(cleaned) else ""

    r.country = normalize_country(r.country)

    # Sanitize alternates.
    alts = []
    for raw in str(r.alternate_emails or "").split("|"):
        e = normalize_email(raw.strip())
        if e and e != r.email:
            alts.append(e)
    r.alternate_emails = " | ".join(unique(alts))

    return True, ""


# Override final validator globally.
validate_record = validate_record_v12


async def scrape_generic_v12(
    context,
    source_url: str,
    llm_enabled: bool,
    llm_sem: asyncio.Semaphore,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    """
    Universal generic source strategy.

    1. HTTP fetch.
    2. Structured parser + profile discovery.
    3. Exhaustive email-first harvest.
    4. Browser render if:
         - HTTP yielded no emails/records, OR
         - browser is likely to expose additional JS-rendered content.
    5. Follow every unique discovered person profile.
    6. Merge all emails by primary address.
    7. Use Qwen only for ambiguous named records, never as a gate that can
       delete email-only records.
    """
    r = await _v11_fetch_html(context, source_url)
    if not r["ok"]:
        errors.append({
            "type": "fetch_failed",
            "source_url": source_url,
            "error": excel_safe(r["error"]),
        })
        return []

    actual = r.get("url") or source_url
    page_type, structured = await parse_html(
        r["html"], actual, llm_enabled, llm_sem
    )

    profiles = []
    if page_type != "EDITORIAL_BOARD":
        profiles = discover_generic_person_profiles_v11(
            r["html"], actual
        )

    email_first_http = harvest_email_first_records_v12(
        r["html"], source_url
    )

    combined = dedupe_records(structured + profiles)

    # Browser-render when HTTP coverage is weak OR source is clearly dynamic.
    should_render = (
        V12_RENDER_WHEN_EMAIL_COVERAGE_LOW
        and (
            len(email_first_http) == 0
            or len(combined) == 0
            or any(x in r["html"].casefold() for x in (
                "__next_data__", "data-reactroot", "nuxt",
                "drupal-settings-json", "vue", "ng-app"
            ))
        )
    )

    email_first_browser = []
    if should_render and page_type != "EDITORIAL_BOARD":
        br = await _v11_fetch_html(
            context, source_url, force_browser=True
        )
        if br["ok"]:
            bactual = br.get("url") or source_url
            _, structured2 = await parse_html(
                br["html"], bactual, llm_enabled, llm_sem
            )
            profiles2 = discover_generic_person_profiles_v11(
                br["html"], bactual
            )
            email_first_browser = harvest_email_first_records_v12(
                br["html"], source_url
            )
            combined = dedupe_records(
                combined + structured2 + profiles2
            )

    # Enrich every email-less profile.
    combined = await enrich_profile_placeholders_v11(
        context, combined, profile_sem, errors, source_url
    )

    # Preserve profile-verified rows directly.
    profile_verified = [
        x for x in combined
        if "profile_v11:" in (x.extraction_method or "")
    ]

    # Direct named rows can use Qwen when ambiguous, but email-only safety rows
    # are never passed to Qwen as a rejection gate.
    other_named = [
        x for x in combined
        if "profile_v11:" not in (x.extraction_method or "")
        and x.name
        and normalize_email(x.email)
    ]
    other_email_only = [
        x for x in combined
        if "profile_v11:" not in (x.extraction_method or "")
        and not x.name
        and normalize_email(x.email)
    ]

    if other_named:
        other_named = await llm_filter_records(
            other_named,
            llm_enabled,
            llm_sem,
            errors,
            source_url,
        )

    base_records = (
        profile_verified
        + other_named
        + other_email_only
    )

    # The exhaustive safety net is merged last, ensuring every discovered email
    # is represented at least once.
    all_email_first = dedupe_records(
        email_first_http + email_first_browser
    )

    merged = merge_email_coverage_v12(
        base_records,
        all_email_first,
    )

    errors.append({
        "type": "email_coverage_summary",
        "source_url": source_url,
        "http_emails": len({
            normalize_email(x.email)
            for x in email_first_http
            if normalize_email(x.email)
        }),
        "browser_emails": len({
            normalize_email(x.email)
            for x in email_first_browser
            if normalize_email(x.email)
        }),
        "profiles_or_structured": len(combined),
        "final_unique_emails": len({
            normalize_email(x.email)
            for x in merged
            if normalize_email(x.email)
        }),
    })

    return merged


async def scrape_source(
    context,
    source_url: str,
    llm_enabled: bool,
    llm_sem: asyncio.Semaphore,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    """
    V12 final dispatcher.
    """
    print(f"\n   [SOURCE] {source_url}", flush=True)

    if _is_jhmi_cancer(source_url):
        print("   [ADAPTER] Johns Hopkins Radiation Oncology", flush=True)
        records = await scrape_jhmi_cancer(
            context, source_url, profile_sem, errors
        )

    elif _is_uh_cancer(source_url) and "clinical-faculty" in source_url.casefold():
        print("   [ADAPTER] UH Cancer Center clinical faculty", flush=True)
        records = await scrape_uh_cancer(
            context, source_url, profile_sem, errors
        )

    elif _is_mayo_cancer(source_url):
        print("   [ADAPTER] Mayo Cancer Center faculty", flush=True)
        records = await scrape_mayo_cancer(
            context, source_url, profile_sem, errors
        )

    elif _v11_is_uiowa_directory(source_url):
        print(
            "   [ADAPTER] University of Iowa Psychiatry directory",
            flush=True,
        )
        records = await scrape_uiowa_directory_v11(
            context, source_url, profile_sem, errors
        )

    elif _v11_is_uiowa_research_team(source_url):
        print(
            "   [ADAPTER] University of Iowa Psychiatry research team",
            flush=True,
        )
        records = await scrape_uiowa_research_team_v11(
            context, source_url, profile_sem, errors
        )

    elif _v11_is_upenn_faculty_database(source_url):
        print(
            "   [ADAPTER] UPenn Psychiatry faculty database",
            flush=True,
        )
        records = await scrape_upenn_faculty_v11(
            context, source_url, profile_sem, errors
        )

    else:
        print(
            "   [ADAPTER] Universal email-first + profile-first parser",
            flush=True,
        )
        records = await scrape_generic_v12(
            context,
            source_url,
            llm_enabled,
            llm_sem,
            profile_sem,
            errors,
        )

    final = []
    for rec in dedupe_records(records):
        rec.source_url = source_url
        ok, reason = validate_record_v12(rec)
        if ok:
            final.append(rec)
        else:
            errors.append({
                "type": "rejected_record",
                "reason": reason,
                "name": excel_safe(rec.name),
                "email": excel_safe(rec.email),
                "country": excel_safe(rec.country),
                "profile_url": excel_safe(rec.profile_url),
                "source_url": source_url,
                "method": excel_safe(rec.extraction_method),
            })

    final = merge_email_coverage_v12([], final)

    print(
        f"   [QUALITY] accepted={len(final)} "
        f"unique_emails={len({x.email for x in final if x.email})} "
        f"blank_names={sum(1 for x in final if not x.name)}",
        flush=True,
    )

    for rec in final:
        print(
            f"      {(rec.name or '[email-only]')[:40]:40} | "
            f"{rec.email[:38]:38} | "
            f"{(rec.profile_url or '')[:55]}",
            flush=True,
        )

    if not final:
        errors.append({
            "type": "no_verified_emails",
            "source_url": source_url,
            "message": "No syntactically valid emails found after HTTP + browser + profile checks.",
        })

    return final




# =============================================================================
# V14 ROBUST OPHTHALMOLOGY + FINAL DISPATCHER
# =============================================================================
# This section intentionally replaces V13 rather than stacking another set of
# conflicting overrides on top of it. Earlier generic/cancer/psychiatry helpers
# remain available, but THIS section owns the final dedupe, validation and
# scrape_source dispatcher used by main().
#
# Core rules:
#   * email is mandatory
#   * name/country are optional
#   * if a personal email has no trustworthy name, derive a display name from
#     the local-part; never manufacture a person name from a shared mailbox
#   * exact person-card/profile associations may legitimately reuse one shared
#     service mailbox across several people
#   * global/footer emails must never be attached to profile people
#   * site-specific adapters run before the universal parser

from urllib.parse import parse_qs, urlencode, urlunparse

V14_FETCH_RETRIES = 3
V14_FETCH_BACKOFF = (0.8, 1.6, 2.8)
V14_PROFILE_CONCURRENCY = max(1, PROFILE_CONCURRENCY)
V14_MAX_PROFILE_RETRIES = 2

V14_TRUSTED_METHODS = (
    'usz_card_v14',
    'usz_profile_v14',
    'eyedoctors_profile_v14',
)

V14_SHARED_LOCALPARTS = {
    'info', 'contact', 'office', 'admin', 'administrator', 'webmaster',
    'support', 'help', 'reception', 'secretary', 'faculty', 'staff', 'team',
    'department', 'dept', 'editor', 'editorial', 'enquiries', 'inquiries',
    'communications', 'media', 'augenklinik', 's.retina', 'hornhaut', 'lid',
    'augalg', 'glaukom', 'uveitis', 'orthoptik', 'notfall', 'sekretariat',
}


def _v14_host(url: str) -> str:
    return urlparse(normalize_url(url)).netloc.casefold().replace('www.', '')


def _v14_is_trusted(rec: PersonRecord) -> bool:
    method = (rec.extraction_method or '').casefold()
    return any(x in method for x in V14_TRUSTED_METHODS)


def _v14_is_shared_email(email: str) -> bool:
    e = normalize_email(email)
    if not e:
        return False
    local = e.split('@', 1)[0].casefold()
    if local in V14_SHARED_LOCALPARTS:
        return True
    if is_generic_email(e) or is_strict_generic_email(e):
        return True
    return any(
        local.startswith(prefix)
        for prefix in (
            'info.', 'info_', 'contact.', 'contact_', 'office.', 'office_',
            'admin.', 'admin_', 'secretariat', 'sekretariat', 'clinic',
            'department', 'faculty', 'team', 'service',
        )
    )


def _v14_clean_name(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ''

    value = re.sub(
        r'^(?:Mr|Mrs|Ms|Miss|Dr|Prof|Professor|PD)\.?\s+',
        '', value, flags=re.I,
    )

    # Remove credentials after comma or at end while keeping the visible person.
    parts = [clean_text(x) for x in value.split(',')]
    if len(parts) > 1:
        suffix = ', '.join(parts[1:]).casefold()
        if re.search(
            r'\b(?:pd|prof|dr|med|rer|nat|phd|msc|mph|fmh|univ|md|frcs|frcophth|facog|facs)\b',
            suffix, re.I,
        ):
            value = parts[0]

    value = re.sub(
        r'(?i)(?:,|\s)+(?:MD|M\.D\.|PhD|Ph\.D\.|MSc|MPH|MBA|MBBS|FRCS(?:\([^)]+\))?|FRCOphth|FACOG|FACS)\.?\s*$',
        '', value,
    )
    value = clean_person_name_candidate(value)
    return value if plausible_name(value) else ''


def _v14_name_from_email(email: str) -> str:
    """User-requested fallback: derive a readable name only for personal mailboxes."""
    e = normalize_email(email)
    if not e or _v14_is_shared_email(e):
        return ''
    local = e.split('@', 1)[0].split('+', 1)[0].strip().lower()
    if not local:
        return ''

    # Split explicit separators first.
    parts = [p for p in re.split(r'[._\-]+', local) if p]
    if len(parts) == 1:
        # Compact institutional usernames are weak evidence. Still provide a
        # last-resort display name as requested, but keep confidence lower.
        token = re.sub(r'\d+$', '', parts[0])
        token = re.sub(r'[^a-zà-öø-ÿā-ž]', '', token, flags=re.I)
        if len(token) < 3:
            return ''
        candidate = token[:1].upper() + token[1:]
        return candidate if plausible_name(candidate) else ''

    cleaned = []
    for token in parts:
        token = re.sub(r'\d+$', '', token)
        token = re.sub(r'[^a-zà-öø-ÿā-ž]', '', token, flags=re.I)
        if not token:
            continue
        if len(token) == 1:
            cleaned.append(token.upper() + '.')
        else:
            cleaned.append(token[:1].upper() + token[1:])
    candidate = ' '.join(cleaned)
    return candidate if candidate and plausible_name(candidate) else ''


def _v14_finalize_record(rec: PersonRecord) -> Tuple[bool, str]:
    rec.email = normalize_email(rec.email)
    if not rec.email:
        return False, 'missing_or_invalid_email'

    cleaned = _v14_clean_name(rec.name)
    if cleaned:
        rec.name = cleaned
    elif not _v14_is_shared_email(rec.email):
        rec.name = _v14_name_from_email(rec.email)
    else:
        rec.name = ''

    rec.email_type = 'shared/role' if _v14_is_shared_email(rec.email) else 'personal'

    c = normalize_country(rec.country)
    rec.country = c if c else ''

    alts = []
    for raw in re.split(r'[|;]+', str(rec.alternate_emails or '')):
        e = normalize_email(raw.strip())
        if e and e != rec.email:
            alts.append(e)
    rec.alternate_emails = ' | '.join(unique(alts))
    rec.email_conflict = 'yes' if rec.email_conflict == 'yes' else 'no'
    rec.affiliation = clean_text(rec.affiliation)[:700]
    rec.scrape_status = 'accepted'
    return True, ''


def dedupe_records_v14(records: List[PersonRecord]) -> List[PersonRecord]:
    """
    Final dedupe rules.

    EyeDoctors.ie:
      * one row per doctor profile
      * prefer a personal email as primary
      * retain every other valid email in alternate_emails

    USZ:
      * one row per explicit person/profile + mailbox association
      * shared clinic/service mailboxes may legitimately repeat for many people

    Generic sources:
      * dedupe by primary email where no stronger person/profile identity exists
    """
    merged: Dict[Tuple, PersonRecord] = {}

    def _alts(rec: PersonRecord) -> List[str]:
        vals = []
        for raw in re.split(r'[|;]+', str(rec.alternate_emails or '')):
            e = normalize_email(raw.strip())
            if e:
                vals.append(e)
        return unique(vals)

    def _prefer_email(a: str, b: str, name: str = '') -> Tuple[str, str]:
        """Return (preferred_primary, other_email)."""
        a = normalize_email(a)
        b = normalize_email(b)
        if not a:
            return b, ''
        if not b:
            return a, ''

        def score(e: str) -> int:
            value = 0
            if not _v14_is_shared_email(e):
                value += 100
            value += min(60, email_name_score(name, e))
            return value

        if score(b) > score(a):
            return b, a
        return a, b

    for rec in records:
        rec.email = normalize_email(rec.email)
        purl = normalize_url(rec.profile_url).casefold()
        nkey = normalize_name_key(rec.name) or clean_text(rec.name).casefold()
        method = (rec.extraction_method or '').casefold()
        trusted = _v14_is_trusted(rec)
        shared = _v14_is_shared_email(rec.email) if rec.email else False

        # EyeDoctors profiles represent one doctor. Do not emit one row per
        # practice email; merge all verified profile emails into one record.
        if purl and 'eyedoctors_profile_v14' in method:
            key = ('eyedoctors_profile', purl)
        # USZ explicitly binds the person card to a displayed clinic mailbox.
        elif rec.email and purl and ('usz_card_v14' in method or 'usz_profile_v14' in method):
            key = ('usz_profile_email', purl, rec.email)
        elif rec.email and nkey and trusted:
            key = ('trusted_name_email', nkey, rec.email)
        elif rec.email and nkey and shared:
            key = ('shared_name_email', nkey, rec.email)
        elif rec.email:
            key = ('email', rec.email)
        elif purl:
            key = ('profile', purl)
        else:
            key = ('name_source', nkey, normalize_url(rec.source_url).casefold())

        if key not in merged:
            rec.alternate_emails = ' | '.join(
                e for e in _alts(rec) if e != rec.email
            )
            merged[key] = rec
            continue

        cur = merged[key]

        # If two EyeDoctors observations have different primary emails, retain
        # both, preferring the personal/name-matching address as primary.
        if key[0] == 'eyedoctors_profile':
            preferred, other = _prefer_email(cur.email, rec.email, cur.name or rec.name)
            accumulated = _alts(cur) + _alts(rec)
            if other:
                accumulated.append(other)
            cur.email = preferred
            cur.email_type = 'shared/role' if _v14_is_shared_email(preferred) else 'personal'
            cur.alternate_emails = ' | '.join(unique([
                e for e in accumulated
                if normalize_email(e) and normalize_email(e) != preferred
            ]))

        if (not _v14_clean_name(cur.name)) and _v14_clean_name(rec.name):
            cur.name = rec.name
        if not cur.profile_url and rec.profile_url:
            cur.profile_url = rec.profile_url
        if not cur.country and rec.country:
            cur.country = rec.country
            cur.country_source = rec.country_source

        for field_name in COLUMNS:
            if not hasattr(cur, field_name) or not hasattr(rec, field_name):
                continue
            if field_name in {'email', 'email_type', 'alternate_emails'} and key[0] == 'eyedoctors_profile':
                continue
            old = getattr(cur, field_name)
            new = getattr(rec, field_name)
            if field_name == 'confidence':
                cur.confidence = max(int(cur.confidence or 0), int(rec.confidence or 0))
            elif field_name == 'alternate_emails':
                vals = _alts(cur) + _alts(rec)
                cur.alternate_emails = ' | '.join(unique([
                    e for e in vals
                    if normalize_email(e) and normalize_email(e) != normalize_email(cur.email)
                ]))
            elif not old and new:
                setattr(cur, field_name, new)

        methods = unique([x for x in (cur.extraction_method, rec.extraction_method) if x])
        cur.extraction_method = '+'.join(methods)

    return list(merged.values())


# Make main()/generic downstream use the final deduper.
dedupe_records = dedupe_records_v14
validate_record = _v14_finalize_record


# -----------------------------------------------------------------------------
# Robust fetch cascade
# -----------------------------------------------------------------------------

async def _v14_context_request_fetch(context, url: str) -> Dict:
    """Playwright APIRequestContext fallback; avoids page.goto networking quirks."""
    try:
        response = await context.request.get(
            url,
            timeout=max(PLAYWRIGHT_TIMEOUT, 30000),
            fail_on_status_code=False,
            headers={
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Referer': url,
            },
        )
        status = response.status
        text = await response.text()
        if status >= 400:
            return {'ok': False, 'error': f'context.request HTTP {status}', 'html': text, 'url': url}
        if len(text or '') < 250:
            return {'ok': False, 'error': 'context.request HTML too small', 'html': text, 'url': url}
        low = text.casefold()
        if any(x in low for x in ('verify you are human', 'captcha', 'access denied', 'checking your browser')):
            return {'ok': False, 'error': 'context.request anti-bot page', 'html': text, 'url': url}
        return {'ok': True, 'error': '', 'html': text, 'url': url}
    except Exception as exc:
        return {'ok': False, 'error': f'context.request: {exc}', 'html': '', 'url': url}


async def _v14_browser_fetch(context, url: str) -> Dict:
    last_error = ''
    waits = ('domcontentloaded', 'commit')
    for attempt in range(V14_FETCH_RETRIES):
        page = await context.new_page()
        try:
            wait_until = waits[min(attempt, len(waits)-1)]
            response = await page.goto(
                url,
                wait_until=wait_until,
                timeout=max(PLAYWRIGHT_TIMEOUT, 30000),
            )
            try:
                await page.wait_for_selector('body', timeout=8000)
            except Exception:
                pass
            await dismiss_popups(page)
            await auto_scroll(page)
            await click_load_more(page)
            await page.wait_for_timeout(500)
            html = await page.content()
            status = response.status if response else 200
            if status >= 400:
                last_error = f'HTTP {status}'
            elif len(html or '') >= 250:
                return {'ok': True, 'error': '', 'html': html, 'url': page.url or url}
            else:
                last_error = 'rendered HTML too small'
        except Exception as exc:
            last_error = str(exc)
        finally:
            await page.close()

        await asyncio.sleep(V14_FETCH_BACKOFF[min(attempt, len(V14_FETCH_BACKOFF)-1)])

    return {'ok': False, 'error': last_error or 'browser fetch failed', 'html': '', 'url': url}


async def fetch_robust_v14(context, urls) -> Dict:
    if isinstance(urls, str):
        urls = [urls]
    candidates = unique([normalize_url(x) for x in urls if normalize_url(x)])
    last = {'ok': False, 'error': 'no candidate URL', 'html': '', 'url': candidates[0] if candidates else ''}

    for url in candidates:
        for attempt in range(V14_FETCH_RETRIES):
            r = await fetch_http(url)
            if r.get('ok'):
                return r
            last = r
            if attempt < V14_FETCH_RETRIES - 1:
                await asyncio.sleep(V14_FETCH_BACKOFF[min(attempt, len(V14_FETCH_BACKOFF)-1)])

    for url in candidates:
        r = await _v14_context_request_fetch(context, url)
        if r.get('ok'):
            return r
        last = r

    for url in candidates:
        r = await _v14_browser_fetch(context, url)
        if r.get('ok'):
            return r
        last = r

    return last


# -----------------------------------------------------------------------------
# USZ
# -----------------------------------------------------------------------------

def _v14_is_usz(url: str) -> bool:
    p = urlparse(url)
    return _v14_host(url) == 'usz.ch' and p.path.rstrip('/').casefold() == '/team'


def _v14_usz_page_urls(source_url: str, page_no: int) -> List[str]:
    p = urlparse(source_url)
    qs = parse_qs(p.query, keep_blank_values=True)
    clinic_id = clean_text((qs.get('clinic_id') or [''])[0])
    if not clinic_id:
        return [source_url]

    canonical_qs = {
        'search': '',
        'letter': '',
        'vested_interests': '',
        'clinic_id': clinic_id,
        'cpage': str(max(1, int(page_no))),
    }
    canonical = urlunparse((
        p.scheme or 'https', p.netloc, p.path or '/team/', '',
        urlencode(canonical_qs), ''
    ))
    amp = (
        f"{p.scheme or 'https'}://{p.netloc}{p.path or '/team/'}"
        f"?search&letter&vested_interests&clinic_id={clinic_id}&cpage={max(1, int(page_no))}"
    )
    original = source_url if page_no == 1 else ''
    return unique([original, canonical, amp])


def _v14_usz_parse(html: str, source_url: str) -> Tuple[List[PersonRecord], List[int]]:
    soup = BeautifulSoup(html, 'html.parser')
    records: List[PersonRecord] = []
    pages = set()

    for a in soup.select('a[href*="cpage="]'):
        try:
            q = parse_qs(urlparse(urljoin(source_url, a.get('href', ''))).query, keep_blank_values=True)
            n = int((q.get('cpage') or [''])[0])
            if n > 0:
                pages.add(n)
        except Exception:
            pass

    for card in soup.select('div.person.list-item, .listing-elements .person, .person.list-item'):
        name_el = card.select_one('.person-name')
        raw_name = clean_text(name_el.get_text(' ', strip=True) if name_el else '')
        name = _v14_clean_name(raw_name)
        if not name:
            continue

        profile_url = ''
        if name_el:
            a = name_el.select_one('a[href]')
            if a:
                profile_url = normalize_url(urljoin(source_url, a.get('href', '')))

        position = clean_text((card.select_one('.person-positions') or card).get_text(' ', strip=True))
        if card.select_one('.person-positions'):
            position = clean_text(card.select_one('.person-positions').get_text(' ', strip=True))
        specialty_el = card.select_one('.person-specialities .value')
        specialty = clean_text(specialty_el.get_text(' ', strip=True)) if specialty_el else ''

        phone = ''
        for row in card.select('.contact-box > div'):
            label_el = row.select_one('.label')
            label = clean_text(label_el.get_text(' ', strip=True) if label_el else '').casefold()
            if label.startswith('tel'):
                value_el = row.select_one('.value')
                phone = clean_text(value_el.get_text(' ', strip=True) if value_el else row.get_text(' ', strip=True))
                break

        emails = []
        for a in card.select('a[href^="mailto:"], a[href^="MAILTO:"]'):
            e = normalize_email(a.get('href', '')) or normalize_email(a.get_text(' ', strip=True))
            if e:
                emails.append(e)
        emails = unique(emails)

        for email in emails:
            records.append(PersonRecord(
                name=name,
                email=email,
                country='Switzerland',
                country_source='site_adapter',
                email_type='shared/role' if _v14_is_shared_email(email) else 'personal',
                email_conflict='no',
                page_type='UNIVERSITY_DIRECTORY',
                academic_title=position,
                specialty=specialty,
                affiliation='University Hospital Zurich (USZ), Department of Ophthalmology',
                university='University Hospital Zurich (USZ)',
                department='Ophthalmology',
                phone=phone,
                profile_url=profile_url,
                source_url=source_url,
                confidence=99,
                extraction_method='usz_card_v14',
                scrape_status='accepted',
            ))

    return dedupe_records_v14(records), sorted(pages)


def _v14_usz_profile_emails(html: str, profile_url: str, source_url: str, name_hint: str) -> List[PersonRecord]:
    """Harvest profile-local emails only; footer/nav is removed first."""
    soup = BeautifulSoup(html, 'html.parser')
    normalize_dom(soup)
    root = soup.select_one('main') or soup.select_one('article') or soup.select_one('#content') or soup.body or soup
    name = _v14_clean_name((root.select_one('h1') or root.select_one('.person-name') or root).get_text(' ', strip=True))
    if not name:
        name = _v14_clean_name(name_hint)

    records = []
    seen = set()
    for a in root.select('a[href^="mailto:"], a[href^="MAILTO:"]'):
        email = normalize_email(a.get('href', '')) or normalize_email(a.get_text(' ', strip=True))
        if not email or email in seen:
            continue
        seen.add(email)
        block = _nearest_person_container(a, email, max_chars=1200)
        text = clean_text(block.get_text(' ', strip=True))
        # Reject obvious administrative/global context unless the listing itself
        # already used the same address for this person.
        low = text.casefold()
        if any(x in low for x in ('media contact', 'webmaster', 'general information')):
            continue
        records.append(PersonRecord(
            name=name,
            email=email,
            country='Switzerland',
            country_source='site_adapter',
            email_type='shared/role' if _v14_is_shared_email(email) else 'personal',
            page_type='UNIVERSITY_DIRECTORY',
            affiliation='University Hospital Zurich (USZ), Department of Ophthalmology',
            university='University Hospital Zurich (USZ)',
            department='Ophthalmology',
            phone=extract_phone(block),
            profile_url=profile_url,
            source_url=source_url,
            confidence=97,
            extraction_method='usz_profile_v14',
        ))
    return dedupe_records_v14(records)


async def scrape_usz_v14(context, source_url: str, profile_sem: asyncio.Semaphore, errors: List[Dict]) -> List[PersonRecord]:
    print('   [V14] USZ cards + explicit cpage crawl + optional profile enrichment', flush=True)
    queue = [1]
    queued = {1}
    visited = set()
    records: List[PersonRecord] = []
    profile_hints: Dict[str, str] = {}
    successful_pages = 0

    while queue and len(visited) < MAX_PAGES_PER_URL:
        page_no = queue.pop(0)
        if page_no in visited:
            continue
        visited.add(page_no)

        r = await fetch_robust_v14(context, _v14_usz_page_urls(source_url, page_no))
        if not r.get('ok'):
            errors.append({
                'type': 'usz_page_fetch_failed', 'source_url': source_url,
                'page': page_no, 'error': excel_safe(r.get('error', '')),
            })
            continue

        successful_pages += 1
        page_records, page_numbers = _v14_usz_parse(r['html'], source_url)
        records.extend(page_records)
        for rec in page_records:
            if rec.profile_url and rec.name:
                profile_hints[rec.profile_url] = rec.name

        for n in page_numbers:
            if n not in visited and n not in queued and n <= MAX_PAGES_PER_URL:
                queued.add(n)
                queue.append(n)

        print(f'   [USZ PAGE] {page_no}: records={len(page_records)} discovered_pages={page_numbers}', flush=True)
        await polite_delay()

    # Profile enrichment is best-effort and never removes listing-card records.
    async def profile_one(url: str, name: str):
        async with profile_sem:
            r = await fetch_robust_v14(context, [url])
            if not r.get('ok'):
                return []
            return _v14_usz_profile_emails(r['html'], url, source_url, name)

    if profile_hints:
        results = await asyncio.gather(*(profile_one(u, n) for u, n in profile_hints.items()))
        for recs in results:
            records.extend(recs)

    records = dedupe_records_v14(records)
    errors.append({
        'type': 'email_coverage_summary',
        'source_url': source_url,
        'adapter': 'usz_v14',
        'pages_fetched': successful_pages,
        'profiles_discovered': len(profile_hints),
        'final_people_rows': len(records),
        'final_unique_emails': len({r.email for r in records if r.email}),
    })
    return records


# -----------------------------------------------------------------------------
# EyeDoctors.ie
# -----------------------------------------------------------------------------

def _v14_is_eyedoctors(url: str) -> bool:
    p = urlparse(url)
    return _v14_host(url) == 'eyedoctors.ie' and '/opthalmologists/' in p.path.casefold()


def _v14_is_eyedoctors_profile(url: str) -> bool:
    return _v14_is_eyedoctors(url) and bool(
        re.search(r'/opthalmologists/doctor/\d+/[^/]+/?$', urlparse(url).path, re.I)
    )


def discover_eyedoctors_profiles_v14(html: str, source_url: str) -> List[Tuple[str, str]]:
    soup = BeautifulSoup(html, 'html.parser')
    found: Dict[str, str] = {}
    for a in soup.select('a[href*="/opthalmologists/doctor/"]'):
        url = normalize_url(urljoin(source_url, a.get('href', '')))
        if not _v14_is_eyedoctors_profile(url):
            continue
        name = _v14_clean_name(a.get_text(' ', strip=True))
        if not name:
            title = clean_text(a.get('title', ''))
            m = re.search(r'See\s+(.+?)(?:\'s|’s)\s+record', title, re.I)
            if m:
                name = _v14_clean_name(m.group(1))
        if url not in found or (name and not found[url]):
            found[url] = name
    return list(found.items())


def _v14_dt_dd_pairs(root: Tag) -> List[Tuple[str, Tag]]:
    out = []
    for dt in root.select('dt'):
        dd = dt.find_next_sibling('dd')
        if dd:
            out.append((clean_text(dt.get_text(' ', strip=True)), dd))
    return out


def parse_eyedoctors_profile_v14(
    html: str,
    profile_url: str,
    source_url: str,
    name_hint: str = '',
) -> List[PersonRecord]:
    """
    Parse exactly ONE EyeDoctors.ie doctor profile into at most ONE row.

    Critical rules:
      * never scrape footer/global site emails
      * only inspect dl.member-details and its practice/private-room blocks
      * collect every valid profile email
      * prefer personal/name-matching email as primary
      * move all remaining verified emails to alternate_emails
      * if only a shared practice mailbox exists, keep it (email is mandatory)
        because it is explicitly present on that doctor's profile
    """
    soup = BeautifulSoup(html, 'html.parser')

    details = (
        soup.select_one('#details dl.member-details')
        or soup.select_one('dl.member-details')
    )
    if details is None:
        return []

    root = soup.select_one('#details') or details.parent or details
    h1 = root.select_one('h1')
    name = _v14_clean_name(
        h1.get_text(' ', strip=True) if h1 else name_hint
    )
    if not name:
        name = _v14_clean_name(name_hint)

    specialty: List[str] = []
    contact_nodes: List[Tag] = []

    for label, dd in _v14_dt_dd_pairs(details):
        low = label.casefold()
        text = clean_text(dd.get_text(' ', strip=True))

        if (
            'sub-specialty' in low
            or 'sub specialty' in low
            or 'specialty' in low
        ):
            if text:
                specialty.append(text)

        # Only these local blocks are allowed to provide a doctor contact.
        if any(x in low for x in (
            'private rooms', 'private room', 'practice', 'clinic',
            'contact details', 'contact detail',
        )):
            contact_nodes.append(dd)

    # Some profiles place Email/Tel as their own dt/dd pair rather than inside
    # Private Rooms. Include those DDs, still strictly inside member-details.
    for label, dd in _v14_dt_dd_pairs(details):
        low = label.casefold().strip(' :')
        if low in {'email', 'e-mail', 'contact', 'telephone', 'tel'}:
            if dd not in contact_nodes:
                contact_nodes.append(dd)

    nodes = contact_nodes or [details]

    # email -> strongest local node
    email_map: Dict[str, Tag] = {}
    for node in nodes:
        for a in node.select('a[href^="mailto:"], a[href^="MAILTO:"]'):
            e = (
                normalize_email(a.get('href', ''))
                or normalize_email(a.get_text(' ', strip=True))
            )
            if e:
                email_map.setdefault(e, node)

    # Plain-text fallback is allowed only inside the same member-details blocks.
    for node in nodes:
        text = clean_text(node.get_text(' ', strip=True))
        if len(text) > 1200:
            continue
        for e in extract_text_emails(text):
            email_map.setdefault(e, node)

    if not email_map:
        return []

    emails = list(email_map.keys())

    def email_rank(email: str) -> Tuple[int, int, str]:
        # Personal/name-correlated addresses outrank shared clinic mailboxes.
        personal_bonus = 100 if not _v14_is_shared_email(email) else 0
        match = email_name_score(name, email) if name else 0
        # deterministic tie-breaker keeps output stable
        return (personal_bonus + match, match, email)

    emails.sort(key=email_rank, reverse=True)
    primary = emails[0]
    alternates = emails[1:]
    primary_node = email_map[primary]

    # Build bounded practice/contact metadata from all local contact blocks,
    # not from the full page.
    affiliation_parts: List[str] = []
    phone = ''
    for node in nodes:
        text = clean_text(node.get_text(' ', strip=True))
        if not text:
            continue
        # Remove email strings/labels but retain practice/address text.
        for e in emails:
            text = re.sub(re.escape(e), ' ', text, flags=re.I)
        text = re.sub(r'(?i)\bE-?mail\s*:\s*', ' ', text)
        text = clean_text(text)
        if text:
            affiliation_parts.append(text)
        if not phone:
            phone = extract_phone(node)

    affiliation = ' | '.join(unique(affiliation_parts))[:700]

    rec = PersonRecord(
        name=name,
        email=primary,
        alternate_emails=' | '.join(alternates),
        country='Ireland',
        country_source='site_adapter',
        email_type='shared/role' if _v14_is_shared_email(primary) else 'personal',
        email_conflict='no',
        page_type='UNIVERSITY_DIRECTORY',
        specialty='; '.join(unique(specialty)),
        affiliation=affiliation,
        phone=phone or extract_phone(primary_node),
        profile_url=profile_url,
        source_url=source_url,
        confidence=100 if name and not _v14_is_shared_email(primary) else (96 if name else 90),
        extraction_method='eyedoctors_profile_v14',
        scrape_status='accepted',
    )

    return [rec]


async def scrape_eyedoctors_v14(context, source_url: str, profile_sem: asyncio.Semaphore, errors: List[Dict]) -> List[PersonRecord]:
    print('   [V14] EyeDoctors listing -> every doctor profile -> member-details email only', flush=True)

    if _v14_is_eyedoctors_profile(source_url):
        items = [(source_url, '')]
    else:
        listing = await fetch_robust_v14(context, [source_url])
        if not listing.get('ok'):
            errors.append({'type': 'fetch_failed', 'source_url': source_url, 'error': excel_safe(listing.get('error', ''))})
            return []
        items = discover_eyedoctors_profiles_v14(listing['html'], listing.get('url') or source_url)

    print(f'   [EYEDOCTORS] profiles_discovered={len(items)}', flush=True)

    async def one(profile_url: str, name_hint: str):
        async with profile_sem:
            last_error = ''
            for attempt in range(V14_MAX_PROFILE_RETRIES):
                r = await fetch_robust_v14(context, [profile_url])
                if r.get('ok'):
                    recs = parse_eyedoctors_profile_v14(r['html'], profile_url, source_url, name_hint)
                    if recs:
                        return recs, None
                    return [], {
                        'type': 'profile_no_verified_email', 'source_url': source_url,
                        'profile_url': profile_url, 'name': name_hint,
                    }
                last_error = r.get('error', '')
                await asyncio.sleep(0.8 * (attempt + 1))
            return [], {
                'type': 'profile_fetch_failed', 'source_url': source_url,
                'profile_url': profile_url, 'name': name_hint, 'error': excel_safe(last_error),
            }

    results = await asyncio.gather(*(one(u, n) for u, n in items))
    records: List[PersonRecord] = []
    no_email = 0
    fetch_failed = 0
    for recs, diag in results:
        records.extend(recs)
        if diag:
            errors.append(diag)
            if diag['type'] == 'profile_no_verified_email':
                no_email += 1
            elif diag['type'] == 'profile_fetch_failed':
                fetch_failed += 1

    records = dedupe_records_v14(records)
    errors.append({
        'type': 'email_coverage_summary',
        'source_url': source_url,
        'adapter': 'eyedoctors_v14',
        'profiles_discovered': len(items),
        'profiles_with_email': len({r.profile_url for r in records if r.profile_url}),
        'profiles_without_email': no_email,
        'profile_fetch_failed': fetch_failed,
        'final_people_rows': len(records),
        'final_unique_emails': len({r.email for r in records if r.email}),
    })
    return records



# =============================================================================
# V16 GENERIC SEMANTIC UNIVERSITY CRAWLER
# =============================================================================
# Goals:
#   * profile links can have ANY URL format; URL path patterns are only hints
#   * identify profile links from human-name anchors + local person structure
#   * verify identity on fetched profile pages
#   * extract emails only from profile-local evidence
#   * support URL pagination, numeric pagination, next buttons, load-more,
#     AJAX pagination and infinite scroll
#   * stop on content stability, not on one specific pagination pattern
# =============================================================================

V16_MAX_LISTING_STATES = 120
V16_MAX_PROFILES = 2500
V16_STABLE_ROUNDS = 2
V16_BROWSER_INTERACTIONS = 80
V16_PROFILE_FETCH_RETRIES = 2

V16_BAD_HOST_PARTS = (
    'linkedin.com', 'facebook.com', 'instagram.com', 'twitter.com', 'x.com',
    'youtube.com', 'researchgate.net', 'orcid.org', 'scholar.google',
    'scopus.com', 'doi.org', 'pubmed.ncbi.nlm.nih.gov',
)
V16_BAD_EXT = (
    '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.zip', '.rar',
    '.jpg', '.jpeg', '.png', '.gif', '.svg', '.webp', '.mp4', '.mp3', '.ics',
)
V16_BAD_LINK_TEXT = {
    'home', 'about', 'contact', 'contacts', 'news', 'events', 'event', 'research',
    'publications', 'publication', 'projects', 'project', 'read more', 'learn more',
    'more', 'details', 'website', 'homepage', 'email', 'e-mail', 'phone',
    'privacy', 'cookies', 'cookie policy', 'next', 'previous', 'back', 'login',
    'download', 'downloads', 'apply', 'search', 'menu',
}
V16_PERSON_CUES = (
    'professor', 'researcher', 'scientist', 'lecturer', 'faculty', 'staff',
    'postdoc', 'postdoctoral', 'doctoral', 'phd', 'research fellow',
    'senior scientist', 'assistant professor', 'associate professor',
    'technician', 'engineer', 'group leader', 'principal investigator',
)
V16_PAGINATION_TEXT = {
    'next', 'next page', 'older', 'more', 'load more', 'show more', 'view more',
    '>', '›', '»', '→',
}


def _v16_host(url: str) -> str:
    try:
        return urlparse(normalize_url(url)).netloc.casefold().replace('www.', '').split(':')[0]
    except Exception:
        return ''


def _v16_root_domain(url: str) -> str:
    host = _v16_host(url)
    if not host:
        return ''
    # Preserve common academic/public suffixes.
    parts = host.split('.')
    if len(parts) <= 2:
        return host
    multi = {'ac.uk', 'co.uk', 'edu.au', 'ac.jp', 'ac.kr', 'edu.cn'}
    suffix2 = '.'.join(parts[-2:])
    if suffix2 in multi and len(parts) >= 3:
        return '.'.join(parts[-3:])
    return '.'.join(parts[-2:])


def _v16_same_org(a: str, b: str) -> bool:
    ra, rb = _v16_root_domain(a), _v16_root_domain(b)
    return bool(ra and rb and ra == rb)


def _v16_bad_target(url: str) -> bool:
    if not url:
        return True
    p = urlparse(url)
    host = p.netloc.casefold().replace('www.', '')
    path = (p.path or '').casefold()
    if any(x in host for x in V16_BAD_HOST_PARTS):
        return True
    if path.endswith(V16_BAD_EXT):
        return True
    if p.scheme not in {'http', 'https'}:
        return True
    return False


def _v16_name_from_anchor(a: Tag) -> str:
    vals = [
        clean_text(a.get_text(' ', strip=True)),
        clean_text(a.get('aria-label', '')),
        clean_text(a.get('title', '')),
    ]
    img = a.select_one('img[alt]')
    if img:
        vals.append(clean_text(img.get('alt', '')))
    for value in vals:
        if not value:
            continue
        # Remove common prefixes while keeping the original spelling.
        value = re.sub(r'(?i)^\s*(?:prof(?:essor)?|dr|mr|mrs|ms|miss)\.?\s+', '', value)
        n = clean_person_name_candidate(value)
        if n and plausible_name(n) and n.casefold() not in V16_BAD_LINK_TEXT:
            return n
    return ''


def _v16_person_container(anchor: Tag) -> Optional[Tag]:
    """Return a compact local block likely representing one person."""
    best = None
    best_score = -999
    cur = anchor
    for depth in range(7):
        if not isinstance(cur, Tag):
            break
        parent = cur if depth == 0 else cur.parent
        if not isinstance(parent, Tag) or parent.name in {'body', 'html'}:
            break
        txt = clean_text(parent.get_text(' ', strip=True))
        if len(txt) > 2600:
            cur = parent
            continue
        score = 0
        cls = ' '.join(parent.get('class', [])).casefold()
        if parent.name in {'article', 'li', 'tr'}:
            score += 25
        if any(k in cls for k in ('person', 'people', 'staff', 'faculty', 'member', 'team', 'profile', 'card', 'researcher')):
            score += 35
        low = txt.casefold()
        if any(cue in low for cue in V16_PERSON_CUES):
            score += 20
        # prefer blocks with only a few links, typical for person cards
        links = parent.select('a[href]')
        if 1 <= len(links) <= 8:
            score += 15
        score -= depth * 3
        if score > best_score:
            best_score, best = score, parent
        cur = parent
    return best or (anchor.parent if isinstance(anchor.parent, Tag) else anchor)


def _v16_score_person_link(anchor: Tag, source_url: str) -> Tuple[int, str, str]:
    href = clean_text(anchor.get('href', ''))
    if not href or href.casefold().startswith(('#', 'mailto:', 'tel:', 'javascript:')):
        return 0, '', ''
    url = normalize_url(urljoin(source_url, href))
    if not url.startswith(('http://', 'https://')) or _v16_bad_target(url):
        return 0, '', ''
    if normalize_url(url).rstrip('/') == normalize_url(source_url).rstrip('/'):
        return 0, '', ''

    name = _v16_name_from_anchor(anchor)
    if not name:
        return 0, '', ''

    score = 60  # human name on link is the main signal
    tokens = _name_tokens(name)
    if len(tokens) >= 2:
        score += 25

    block = _v16_person_container(anchor)
    if block is not None:
        btxt = clean_text(block.get_text(' ', strip=True)).casefold()
        if any(cue in btxt for cue in V16_PERSON_CUES):
            score += 25
        cls = ' '.join(block.get('class', [])).casefold()
        if any(k in cls for k in ('person', 'people', 'staff', 'faculty', 'member', 'team', 'profile', 'card', 'researcher')):
            score += 25

    # Same institution is a strong hint, but NOT a requirement.
    if _v16_same_org(source_url, url):
        score += 35
    else:
        # external links are allowed only when very strongly person-associated
        score -= 15

    # Conventional profile-like URL is only a hint, never a requirement.
    path = (urlparse(url).path or '').casefold()
    if any(x in path for x in ('profile', 'people', 'person', 'staff', 'faculty', 'team', 'member', 'researcher', 'sitoweb')):
        score += 10

    return score, url, name


def discover_person_profiles_v16(html: str, source_url: str) -> List[PersonRecord]:
    soup = BeautifulSoup(html or '', 'html.parser')
    out: List[PersonRecord] = []
    seen = set()

    # Evaluate every link. We are structural/semantic, not path-based.
    for a in soup.select('a[href]'):
        score, url, name = _v16_score_person_link(a, source_url)
        if score < 95 or not url or not name:
            continue
        key = url.casefold()
        if key in seen:
            continue
        seen.add(key)

        block = _v16_person_container(a)
        title = ''
        affiliation = ''
        if block is not None:
            text = clean_text(block.get_text(' ', strip=True))
            for sel in ('.title', '.role', '.position', '.job-title', '.academic-title', 'h3', 'h4', 'p'):
                el = block.select_one(sel)
                if el:
                    t = clean_text(el.get_text(' ', strip=True))
                    if t and t != name and len(t) <= 220 and not plausible_name(t):
                        title = t
                        break
            affiliation = text.replace(name, '', 1).strip(' ,;:-')[:700]

        out.append(PersonRecord(
            name=name,
            page_type='UNIVERSITY_DIRECTORY',
            academic_title=title,
            affiliation=affiliation,
            profile_url=url,
            source_url=source_url,
            confidence=min(95, score),
            extraction_method='semantic_profile_discovery_v16',
        ))
        if len(out) >= V16_MAX_PROFILES:
            break
    return out


def _v16_page_signature(html: str, url: str) -> Tuple[str, Tuple[str, ...], Tuple[str, ...]]:
    """Stable signature based on people/profile links and emails, not raw HTML."""
    profiles = discover_person_profiles_v16(html, url)
    pset = tuple(sorted({normalize_url(r.profile_url).casefold() for r in profiles if r.profile_url}))
    emails = tuple(sorted(set(extract_text_emails(clean_text(BeautifulSoup(html or '', 'html.parser').get_text(' ', strip=True))))))
    return normalize_url(url).casefold(), pset, emails


def _v16_url_pagination_candidates(html: str, current_url: str, source_url: str) -> List[str]:
    soup = BeautifulSoup(html or '', 'html.parser')
    scored = []
    for a in soup.select('a[href]'):
        href = clean_text(a.get('href', ''))
        if not href or href.casefold().startswith(('#', 'mailto:', 'tel:', 'javascript:')):
            continue
        url = normalize_url(urljoin(current_url, href))
        if not url.startswith(('http://', 'https://')):
            continue
        if not _v16_same_org(url, source_url):
            continue
        if normalize_url(url).casefold() == normalize_url(current_url).casefold():
            continue

        text = clean_text(a.get_text(' ', strip=True)).casefold()
        rel = [str(x).casefold() for x in (a.get('rel') or [])]
        p = urlparse(url)
        qs = parse_qs(p.query)
        score = 0
        if 'next' in rel:
            score += 120
        if text in V16_PAGINATION_TEXT:
            score += 100
        if re.fullmatch(r'\d{1,4}', text):
            score += 45
        if any(k.casefold() in {'page', 'p', 'pg', 'offset', 'start', 'from', 'cursor'} for k in qs):
            score += 35
        if re.search(r'/(?:page|p)/?\d+/?$', p.path, re.I):
            score += 35
        aria = clean_text(a.get('aria-label', '')).casefold()
        if 'next' in aria or 'page' in aria:
            score += 45
        cls = ' '.join(a.get('class', [])).casefold()
        if 'pagination' in cls or 'pager' in cls or 'next' in cls:
            score += 30
        if score >= 40:
            scored.append((score, url))

    seen = set()
    out = []
    for _, u in sorted(scored, key=lambda x: x[0], reverse=True):
        k = u.casefold()
        if k not in seen:
            seen.add(k)
            out.append(u)
    return out


async def _v16_browser_collect_listing_states(context, start_url: str, errors: List[Dict]) -> List[Tuple[str, str]]:
    """Collect rendered listing states via next/numeric/load-more/infinite-scroll."""
    states: List[Tuple[str, str]] = []
    page = await context.new_page()
    seen_signatures = set()
    stable = 0
    interactions = 0
    current_numeric = 1
    try:
        try:
            await page.goto(start_url, wait_until='domcontentloaded', timeout=PLAYWRIGHT_TIMEOUT)
        except Exception:
            try:
                await page.goto(start_url, wait_until='commit', timeout=PLAYWRIGHT_TIMEOUT)
            except Exception as exc:
                errors.append({'type': 'browser_listing_fetch_failed', 'source_url': start_url, 'error': excel_safe(str(exc))})
                return states

        await dismiss_popups(page)
        await page.wait_for_timeout(700)

        while interactions < V16_BROWSER_INTERACTIONS and len(states) < V16_MAX_LISTING_STATES:
            try:
                await auto_scroll(page)
            except Exception:
                pass
            await page.wait_for_timeout(350)
            html = await page.content()
            sig = _v16_page_signature(html, page.url)
            if sig not in seen_signatures:
                seen_signatures.add(sig)
                states.append((page.url, html))
                stable = 0
            else:
                stable += 1

            # 1) load more/show more/view more
            clicked = False
            for label in ('Load more', 'Load More', 'Show more', 'Show More', 'View more', 'View More', 'More results'):
                try:
                    loc = page.get_by_text(label, exact=True).first
                    if await loc.count() and await loc.is_visible():
                        disabled = await loc.get_attribute('disabled')
                        aria_disabled = await loc.get_attribute('aria-disabled')
                        if disabled is None and aria_disabled != 'true':
                            await loc.click(timeout=2200)
                            await page.wait_for_timeout(700)
                            clicked = True
                            interactions += 1
                            break
                except Exception:
                    pass
            if clicked:
                continue

            # 2) explicit next controls
            selectors = [
                "a[rel='next']", "button[rel='next']", ".pagination-next a", ".pager-next a",
                ".pagination .next a", "a[aria-label*='Next']", "button[aria-label*='Next']",
                "a[title*='Next']", "button[title*='Next']",
            ]
            for sel in selectors:
                try:
                    loc = page.locator(sel).first
                    if await loc.count() and await loc.is_visible():
                        disabled = await loc.get_attribute('disabled')
                        aria_disabled = await loc.get_attribute('aria-disabled')
                        if disabled is None and aria_disabled != 'true':
                            await loc.click(timeout=2200)
                            await page.wait_for_timeout(700)
                            clicked = True
                            interactions += 1
                            break
                except Exception:
                    pass
            if clicked:
                current_numeric += 1
                continue

            # 3) text-based next
            for label in ('Next', 'NEXT', 'next', 'Next page', '›', '»', '>'):
                try:
                    loc = page.get_by_text(label, exact=True).last
                    if await loc.count() and await loc.is_visible():
                        await loc.click(timeout=2200)
                        await page.wait_for_timeout(700)
                        clicked = True
                        interactions += 1
                        break
                except Exception:
                    pass
            if clicked:
                current_numeric += 1
                continue

            # 4) numeric next page
            target = str(current_numeric + 1)
            for role in ('link', 'button'):
                try:
                    loc = page.get_by_role(role, name=target, exact=True).last
                    if await loc.count() and await loc.is_visible():
                        disabled = await loc.get_attribute('disabled')
                        aria_disabled = await loc.get_attribute('aria-disabled')
                        if disabled is None and aria_disabled != 'true':
                            await loc.click(timeout=2200)
                            await page.wait_for_timeout(700)
                            clicked = True
                            interactions += 1
                            current_numeric += 1
                            break
                except Exception:
                    pass
            if clicked:
                continue

            # 5) infinite scroll: one extra content-growth attempt
            try:
                before = await page.evaluate('document.body.scrollHeight')
                await page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                await page.wait_for_timeout(900)
                after = await page.evaluate('document.body.scrollHeight')
                interactions += 1
                if after > before:
                    continue
            except Exception:
                pass

            if stable >= V16_STABLE_ROUNDS:
                break
            # no control and no growth => stable termination
            stable += 1
            if stable >= V16_STABLE_ROUNDS:
                break

    finally:
        await page.close()
    return states


def _v16_profile_name_match(expected: str, html: str) -> bool:
    if not expected:
        return True
    soup = BeautifulSoup(html or '', 'html.parser')
    expected_tokens = set(_name_tokens(expected))
    if not expected_tokens:
        return True
    candidates = []
    for sel in ("[itemprop='name']", 'main h1', 'article h1', '.person-name', '.profile-name', '.staff-name', '.faculty-name', 'h1'):
        for el in soup.select(sel):
            n = clean_person_name_candidate(el.get_text(' ', strip=True))
            if n and plausible_name(n):
                candidates.append(n)
    if not candidates:
        return True  # don't reject profiles solely because markup lacks heading
    for n in candidates:
        actual = set(_name_tokens(n))
        if actual and expected_tokens & actual:
            return True
    return False


def _v16_extract_profile(html: str, profile_url: str, source_url: str, name_hint: str) -> Optional[PersonRecord]:
    if not html:
        return None
    if not _v16_profile_name_match(name_hint, html):
        return None

    soup = BeautifulSoup(html, 'html.parser')
    root = soup.select_one('main') or soup.select_one('article') or soup.select_one('#main-content') or soup.select_one('#content') or soup.body or soup

    # Name: listing hint first, otherwise profile heading/JSON-LD.
    name = clean_person_name_candidate(name_hint) if name_hint else ''
    if not (name and plausible_name(name)):
        try:
            name = _v11_extract_profile_name(soup, '')
        except Exception:
            name = ''
    if not name:
        for sel in ("[itemprop='name']", 'main h1', 'article h1', '.person-name', '.profile-name', '.staff-name', '.faculty-name', 'h1'):
            el = soup.select_one(sel)
            if el:
                n = clean_person_name_candidate(el.get_text(' ', strip=True))
                if n and plausible_name(n):
                    name = n
                    break

    # Comprehensive profile-local email evidence.
    try:
        evidence = _v11_extract_email_evidence(root)
    except Exception:
        evidence = []

    # Fallback direct mailto/plain contact blocks.
    if not evidence:
        for a in root.select('a[href]'):
            href = clean_text(a.get('href', ''))
            if href.casefold().startswith('mailto:'):
                e = normalize_email(href.split('?', 1)[0]) or normalize_email(a.get_text(' ', strip=True))
                if e:
                    evidence.append((190, e, a, 'mailto_v16'))
        for el in root.select("p,li,dd,dt,address,[class*='email'],[class*='contact'],[id*='email'],[id*='contact']"):
            txt = clean_text(el.get_text(' ', strip=True))
            if txt and len(txt) <= 700:
                for e in extract_text_emails(txt):
                    evidence.append((110, e, el, 'local_text_v16'))

    ranked = []
    seen_emails = set()
    for base, raw_email, node, kind in evidence:
        email = normalize_email(raw_email)
        if not email or email in seen_emails:
            continue
        seen_emails.add(email)
        score = int(base)
        if name:
            score += min(50, email_name_score(name, email))
        local = ''
        if node is not None:
            try:
                local = clean_text(_nearest_person_container(node, email, max_chars=1200).get_text(' ', strip=True)).casefold()
            except Exception:
                try:
                    local = clean_text(node.get_text(' ', strip=True)).casefold()
                except Exception:
                    local = ''
        if name and local:
            if any(t in local for t in _name_tokens(name)):
                score += 25
        if any(x in local for x in ('webmaster', 'press office', 'media contact', 'general enquiries', 'general inquiries', 'privacy')):
            score -= 140
        # Shared/generic allowed only when it is genuinely profile-local; personal preferred.
        if is_generic_email(email) or is_strict_generic_email(email):
            if base < 150:
                score -= 100
            else:
                score -= 25
        ranked.append((score, email, node, kind))

    ranked.sort(key=lambda x: x[0], reverse=True)
    if not ranked or ranked[0][0] < 90:
        return None

    best_score, email, email_node, email_kind = ranked[0]
    alternates = []
    for score, e, _, _ in ranked[1:]:
        if e != email and score >= 110 and e not in alternates:
            alternates.append(e)

    # Metadata from a compact profile-local block where possible.
    local_block = root
    if email_node is not None:
        try:
            local_block = _nearest_person_container(email_node, email, max_chars=1500)
        except Exception:
            pass
    local_text = clean_text(local_block.get_text(' ', strip=True))
    links = extract_links(root, profile_url)
    affiliation = split_affiliation_from_block(local_block, name, [email] + alternates, '')[:700]

    title = ''
    for sel in ('.job-title', '.person-title', '.profile-title', '.role', '.position', '.academic-title', '[class*="title"]'):
        el = root.select_one(sel)
        if el:
            t = clean_text(el.get_text(' ', strip=True))
            if t and t != name and len(t) <= 220:
                title = t
                break

    country = country_from_tld(email)
    country_source = 'email_tld' if country else ''
    if not country:
        c = extract_country_from_text(affiliation)
        if c:
            country, country_source = c, 'affiliation'

    return PersonRecord(
        name=name if name and plausible_name(name) else name_from_email(email),
        email=email,
        country=country,
        alternate_emails=' | '.join(unique(alternates)),
        email_type='shared/role' if (is_generic_email(email) or is_strict_generic_email(email)) else 'personal',
        email_conflict='no',
        page_type='UNIVERSITY_DIRECTORY',
        academic_title=title,
        affiliation=affiliation,
        phone=extract_phone(local_block),
        orcid=links.get('orcid', '') or extract_orcid_text(local_text),
        google_scholar=links.get('google_scholar', ''),
        scopus_author_id=extract_scopus(local_text),
        pubmed=links.get('pubmed', ''),
        profile_url=profile_url,
        personal_homepage=links.get('personal_homepage', ''),
        source_url=source_url,
        country_source=country_source,
        confidence=min(100, max(90, best_score)),
        extraction_method=f'semantic_profile_v16:{email_kind}',
        scrape_status='accepted',
    )


async def _v16_fetch_profile(context, rec: PersonRecord, sem: asyncio.Semaphore) -> Tuple[Optional[PersonRecord], Optional[Dict]]:
    async with sem:
        last_error = ''
        for attempt in range(V16_PROFILE_FETCH_RETRIES + 1):
            try:
                r = await fetch_http(rec.profile_url)
                if not r.get('ok'):
                    r = await playwright_fetch(context, rec.profile_url)
                if r.get('ok'):
                    parsed = _v16_extract_profile(r.get('html', ''), r.get('url') or rec.profile_url, rec.source_url, rec.name)
                    if parsed and normalize_email(parsed.email):
                        return parsed, None
                    return None, {
                        'type': 'profile_no_verified_email',
                        'source_url': rec.source_url,
                        'profile_url': rec.profile_url,
                        'name': rec.name,
                    }
                last_error = r.get('error', '')
            except Exception as exc:
                last_error = str(exc)
            if attempt < V16_PROFILE_FETCH_RETRIES:
                await asyncio.sleep(0.6 * (attempt + 1))
        return None, {
            'type': 'profile_fetch_failed',
            'source_url': rec.source_url,
            'profile_url': rec.profile_url,
            'name': rec.name,
            'error': excel_safe(last_error),
        }


async def scrape_generic_v16(
    context,
    source_url: str,
    llm_enabled: bool,
    llm_sem: asyncio.Semaphore,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    print('   [V16] Semantic profile discovery + universal pagination', flush=True)

    # ---------------------------------------------------------
    # Phase 1: gather listing states using BOTH direct URL pagination and browser interaction
    # ---------------------------------------------------------
    listing_states: List[Tuple[str, str]] = []
    queue = [source_url]
    queued = {normalize_url(source_url).casefold()}
    visited = set()

    while queue and len(listing_states) < V16_MAX_LISTING_STATES:
        url = queue.pop(0)
        key = normalize_url(url).casefold()
        if key in visited:
            continue
        visited.add(key)
        r = await fetch_http(url)
        if not r.get('ok'):
            r = await playwright_fetch(context, url)
        if not r.get('ok'):
            errors.append({'type': 'listing_fetch_failed', 'source_url': source_url, 'page_url': url, 'error': excel_safe(r.get('error', ''))})
            continue
        actual = r.get('url') or url
        html = r.get('html', '')
        listing_states.append((actual, html))
        for nxt in _v16_url_pagination_candidates(html, actual, source_url):
            nk = normalize_url(nxt).casefold()
            if nk not in queued and nk not in visited:
                queued.add(nk)
                queue.append(nxt)

    # Rendered pagination/Load-more/infinite-scroll, always attempted once because
    # many institutional directories expose pagination only client-side.
    try:
        rendered_states = await _v16_browser_collect_listing_states(context, source_url, errors)
        # merge by semantic signature
        sig_seen = set()
        merged_states = []
        for u, h in listing_states + rendered_states:
            sig = _v16_page_signature(h, u)
            if sig not in sig_seen:
                sig_seen.add(sig)
                merged_states.append((u, h))
        listing_states = merged_states[:V16_MAX_LISTING_STATES]
    except Exception as exc:
        errors.append({'type': 'pagination_warning', 'source_url': source_url, 'error': excel_safe(str(exc))})

    # ---------------------------------------------------------
    # Phase 2: collect direct email rows + semantic profile placeholders
    # ---------------------------------------------------------
    direct: List[PersonRecord] = []
    placeholders: Dict[str, PersonRecord] = {}

    for page_url, html in listing_states:
        # Existing local email parser remains useful for pages that expose emails directly.
        try:
            _, rows = await parse_html(html, page_url, llm_enabled, llm_sem)
            for r in rows:
                if normalize_email(r.email):
                    r.source_url = source_url
                    direct.append(r)
        except Exception as exc:
            errors.append({'type': 'listing_parse_warning', 'source_url': source_url, 'page_url': page_url, 'error': excel_safe(str(exc))})

        for p in discover_person_profiles_v16(html, page_url):
            p.source_url = source_url
            key = normalize_url(p.profile_url).casefold()
            if not key:
                continue
            old = placeholders.get(key)
            if old is None or (not old.name and p.name):
                placeholders[key] = p
            if len(placeholders) >= V16_MAX_PROFILES:
                break

    print(
        f'   [DISCOVERY] listing_states={len(listing_states)} '
        f'profiles={len(placeholders)} direct_email_rows={len(direct)}',
        flush=True,
    )

    # ---------------------------------------------------------
    # Phase 3: fetch EVERY unique discovered profile
    # ---------------------------------------------------------
    profile_items = list(placeholders.values())[:V16_MAX_PROFILES]
    results = await asyncio.gather(*(_v16_fetch_profile(context, p, profile_sem) for p in profile_items))
    profile_rows: List[PersonRecord] = []
    for rec, diag in results:
        if rec:
            profile_rows.append(rec)
        if diag:
            errors.append(diag)

    # ---------------------------------------------------------
    # Phase 4: final merge and validation
    # ---------------------------------------------------------
    merged = dedupe_records_v14(direct + profile_rows)
    final: List[PersonRecord] = []
    for rec in merged:
        rec.source_url = source_url
        ok, reason = _v14_finalize_record(rec)
        if ok:
            final.append(rec)
        else:
            errors.append({
                'type': 'rejected_record', 'reason': reason,
                'name': excel_safe(rec.name), 'email': excel_safe(rec.email),
                'country': excel_safe(rec.country), 'profile_url': excel_safe(rec.profile_url),
                'source_url': source_url, 'method': excel_safe(rec.extraction_method),
            })

    final = dedupe_records_v14(final)
    errors.append({
        'type': 'email_coverage_summary',
        'source_url': source_url,
        'adapter': 'semantic_generic_v16',
        'listing_states': len(listing_states),
        'profiles_discovered': len(placeholders),
        'profiles_with_email': len({r.profile_url for r in profile_rows if r.profile_url and r.email}),
        'direct_email_rows': len(direct),
        'final_rows': len(final),
        'final_unique_emails': len({r.email for r in final if r.email}),
    })
    print(
        f'   [V16 QUALITY] listing_states={len(listing_states)} '
        f'profiles={len(placeholders)} profile_rows={len(profile_rows)} '
        f'final={len(final)} unique_emails={len({r.email for r in final if r.email})}',
        flush=True,
    )
    return final

# -----------------------------------------------------------------------------
# Final dispatcher
# -----------------------------------------------------------------------------

async def scrape_source(
    context,
    source_url: str,
    llm_enabled: bool,
    llm_sem: asyncio.Semaphore,
    profile_sem: asyncio.Semaphore,
    errors: List[Dict],
) -> List[PersonRecord]:
    print(f'\n   [SOURCE] {source_url}', flush=True)

    if _v14_is_usz(source_url):
        records = await scrape_usz_v14(context, source_url, profile_sem, errors)

    elif _v14_is_eyedoctors(source_url):
        records = await scrape_eyedoctors_v14(context, source_url, profile_sem, errors)

    elif _is_jhmi_cancer(source_url):
        print('   [ADAPTER] Johns Hopkins Radiation Oncology', flush=True)
        records = await scrape_jhmi_cancer(context, source_url, profile_sem, errors)

    elif _is_uh_cancer(source_url) and 'clinical-faculty' in source_url.casefold():
        print('   [ADAPTER] UH Cancer Center clinical faculty', flush=True)
        records = await scrape_uh_cancer(context, source_url, profile_sem, errors)

    elif _is_mayo_cancer(source_url):
        print('   [ADAPTER] Mayo Cancer Center faculty', flush=True)
        records = await scrape_mayo_cancer(context, source_url, profile_sem, errors)

    elif _v11_is_uiowa_directory(source_url):
        print('   [ADAPTER] University of Iowa Psychiatry directory', flush=True)
        records = await scrape_uiowa_directory_v11(context, source_url, profile_sem, errors)

    elif _v11_is_uiowa_research_team(source_url):
        print('   [ADAPTER] University of Iowa Psychiatry research team', flush=True)
        records = await scrape_uiowa_research_team_v11(context, source_url, profile_sem, errors)

    elif _v11_is_upenn_faculty_database(source_url):
        print('   [ADAPTER] UPenn Psychiatry faculty database', flush=True)
        records = await scrape_upenn_faculty_v11(context, source_url, profile_sem, errors)

    else:
        print('   [ADAPTER] Universal email-first + profile-first parser', flush=True)
        records = await scrape_generic_v16(
            context, source_url, llm_enabled, llm_sem, profile_sem, errors
        )

    final: List[PersonRecord] = []
    for rec in dedupe_records_v14(records):
        rec.source_url = source_url
        ok, reason = _v14_finalize_record(rec)
        if ok:
            final.append(rec)
        else:
            errors.append({
                'type': 'rejected_record', 'reason': reason,
                'name': excel_safe(rec.name), 'email': excel_safe(rec.email),
                'country': excel_safe(rec.country), 'profile_url': excel_safe(rec.profile_url),
                'source_url': source_url, 'method': excel_safe(rec.extraction_method),
            })

    final = dedupe_records_v14(final)
    print(
        f'   [QUALITY] accepted={len(final)} '
        f'unique_emails={len({r.email for r in final if r.email})} '
        f'blank_names={sum(1 for r in final if not r.name)}',
        flush=True,
    )

    if not final:
        errors.append({
            'type': 'no_verified_emails',
            'source_url': source_url,
            'message': 'No syntactically valid emails found after robust HTTP/request/browser/profile checks.',
        })

    return final


if __name__ == '__main__':
    asyncio.run(main())