from __future__ import annotations

import argparse
import json
import logging
import random
import re
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from urllib.parse import parse_qs, quote, urljoin, urlparse

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import sync_playwright


BASE_URL = "https://www.english.ids-cologne.de"
LIST_URL = (
    BASE_URL
    + "/ids-cologne-exhibitors/list-of-exhibitors/"
)

OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "ids_cologne_dental_exhibitors.xlsx"
CHECKPOINT_FILE = OUTPUT_DIR / "ids_cologne_checkpoint.csv"
PROFILE_DIR = Path("browser_profiles") / "ids_cologne"

LOGIN_WAIT_SECONDS = 900
PAGE_WAIT_MS = 2500
PROFILE_WAIT_MS = 1800
MIN_PROFILE_DELAY = 2.0
MAX_PROFILE_DELAY = 4.5
MIN_PAGE_DELAY = 6.0
MAX_PAGE_DELAY = 12.0
LONG_BREAK_EVERY_PAGES = 5
LONG_BREAK_MIN = 25.0
LONG_BREAK_MAX = 45.0
PAGINATION_RETRIES = 4
PAGE_SIZE = 20

PAGINATE_VALUES = (
    "%7B%22stichwort%22%3A%22%22%2C"
    "%22suchart%22%3A%22alle%22%7D"
)


@dataclass
class Exhibitor:
    company_name: str
    country: str
    hall_stand: str
    website: str
    email: str
    address: str
    telephone: str
    profile_link: str
    source_url: str
    list_offset: int


def clean(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def is_login_page(html_text: str) -> bool:
    text = html_text.lower()

    return any(
        marker in text
        for marker in [
            "accounts.google.com",
            "sign in with google",
            "performing security verification",
            "just a moment",
            "challenge-platform",
            "cf-chl",
        ]
    )


def has_exhibitor_results(html_text: str) -> bool:
    soup = BeautifulSoup(html_text, "html.parser")

    return bool(
        soup.select_one(
            ".esr.search-results .item a.db-aslink[href]"
        )
    )


def wait_for_results(page, timeout_seconds: int) -> None:
    deadline = time.time() + timeout_seconds

    while time.time() < deadline:
        html_text = page.content()

        if is_login_page(html_text):
            logging.info(
                "Waiting for Google login or security verification..."
            )
            page.wait_for_timeout(3000)
            continue

        if has_exhibitor_results(html_text):
            logging.info("IDS exhibitor results loaded.")
            return

        logging.info("Waiting for exhibitor results...")
        page.wait_for_timeout(2500)

    raise RuntimeError(
        "Exhibitor results did not load before timeout."
    )


def build_list_url(offset: int) -> str:
    if offset == 0:
        return LIST_URL

    return (
        LIST_URL
        + "?route=aussteller/blaettern"
        + f"&&start={offset}"
        + f"&paginatevalues={PAGINATE_VALUES}"
    )


def extract_offsets(html_text: str) -> list[int]:
    soup = BeautifulSoup(html_text, "html.parser")
    offsets = {0}

    for anchor in soup.select(
        'a[href*="route=aussteller/blaettern"]'
        '[href*="start="]'
    ):
        href = clean(anchor.get("href", ""))

        match = re.search(
            r"(?:[?&]|&&)start=(\d+)",
            href,
            re.I,
        )

        if match:
            offsets.add(int(match.group(1)))

    return sorted(offsets)


def parse_list_page(
    html_text: str,
    source_url: str,
    offset: int,
) -> list[Exhibitor]:
    soup = BeautifulSoup(html_text, "html.parser")
    records: list[Exhibitor] = []

    for item in soup.select(
        ".esr.search-results > .item, "
        ".esr.search-results .item"
    ):
        link = item.select_one(
            ".col1ergebnis a.db-aslink[href]"
        )

        if not link:
            continue

        company_name = clean(
            link.get("title")
            or link.get_text(" ", strip=True)
        )
        profile_link = urljoin(
            BASE_URL,
            clean(link.get("href", "")),
        )

        if not company_name or not profile_link:
            continue

        country = ""

        first_column = item.select_one(
            ".col1ergebnis"
        )

        if first_column:
            paragraphs = first_column.select("p")

            for paragraph in paragraphs:
                value = clean(
                    paragraph.get_text(
                        " ",
                        strip=True,
                    )
                )

                if value:
                    country = value
                    break

        hall_stand = ""

        hall_link = item.select_one(
            'a[href*="route=hallenplan2/index"]'
        )

        if hall_link:
            hall_stand = clean(
                hall_link.get_text(
                    " ",
                    strip=True,
                )
            )

        records.append(
            Exhibitor(
                company_name=company_name,
                country=country,
                hall_stand=hall_stand,
                website="",
                email="",
                address="",
                telephone="",
                profile_link=profile_link,
                source_url=source_url,
                list_offset=offset,
            )
        )

    unique: dict[str, Exhibitor] = {}

    for record in records:
        unique[
            record.profile_link.lower().rstrip("/")
        ] = record

    return list(unique.values())


def find_exhibitor_jsonld(
    soup: BeautifulSoup,
) -> dict:
    candidates: list[dict] = []

    for script in soup.select(
        'script[type="application/ld+json"]'
    ):
        raw = script.string or script.get_text()

        if not raw:
            continue

        raw = raw.strip()

        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            continue

        if isinstance(payload, list):
            candidates.extend(
                item
                for item in payload
                if isinstance(item, dict)
            )
        elif isinstance(payload, dict):
            graph = payload.get("@graph")

            if isinstance(graph, list):
                candidates.extend(
                    item
                    for item in graph
                    if isinstance(item, dict)
                )

            candidates.append(payload)

    for candidate in candidates:
        candidate_type = str(
            candidate.get("@type", "")
        ).lower()

        if (
            candidate.get("email")
            or candidate.get("telephone")
            or candidate.get("address")
        ) and candidate_type not in {
            "event",
            "organization",
            "webpage",
            "breadcrumblist",
        }:
            return candidate

    for candidate in candidates:
        if (
            candidate.get("email")
            or candidate.get("telephone")
            or candidate.get("address")
        ):
            name = clean(candidate.get("name"))

            if "koelnmesse" not in name.lower():
                return candidate

    return {}


def format_address(address_value) -> str:
    if isinstance(address_value, str):
        return clean(address_value)

    if not isinstance(address_value, dict):
        return ""

    parts = [
        clean(address_value.get("streetAddress")),
        clean(address_value.get("postalCode")),
        clean(address_value.get("addressLocality")),
        clean(address_value.get("addressRegion")),
        clean(address_value.get("addressCountry")),
    ]

    return ", ".join(
        part
        for part in parts
        if part
    )


def extract_email(
    soup: BeautifulSoup,
    jsonld: dict,
) -> str:
    value = clean(jsonld.get("email"))

    if value:
        value = value.replace("mailto:", "")

        match = re.search(
            r"[A-Z0-9._%+\-]+"
            r"@[A-Z0-9.\-]+\.[A-Z]{2,}",
            value,
            re.I,
        )

        if match:
            return match.group(0).lower()

    mail_link = soup.select_one(
        'a[href^="mailto:"]'
    )

    if mail_link:
        value = clean(
            mail_link.get("href", "")
        ).replace("mailto:", "")

        value = value.split("?", 1)[0]

        match = re.search(
            r"[A-Z0-9._%+\-]+"
            r"@[A-Z0-9.\-]+\.[A-Z]{2,}",
            value,
            re.I,
        )

        if match:
            return match.group(0).lower()

    return ""


def extract_website(
    soup: BeautifulSoup,
    jsonld: dict,
) -> str:
    value = clean(jsonld.get("url"))

    if value and value.startswith(
        ("http://", "https://")
    ):
        if "ids-cologne.de" not in value:
            return value

    for anchor in soup.select(
        ".sico.ico_link a[href], "
        'a[href^="http"]'
    ):
        href = clean(anchor.get("href", ""))

        if not href:
            continue

        host = urlparse(href).netloc.lower()

        if any(
            blocked in host
            for blocked in [
                "ids-cologne.de",
                "koelnmesse.",
                "facebook.com",
                "instagram.com",
                "linkedin.com",
                "youtube.com",
                "twitter.com",
                "x.com",
            ]
        ):
            continue

        return href

    return ""


def scrape_profile(
    context,
    record: Exhibitor,
) -> Exhibitor:
    page = context.new_page()

    try:
        page.goto(
            record.profile_link,
            wait_until="domcontentloaded",
            timeout=90000,
        )

        page.wait_for_timeout(PROFILE_WAIT_MS)

        deadline = time.time() + 180

        while time.time() < deadline:
            html_text = page.content()

            if not is_login_page(html_text):
                break

            logging.info(
                "Login verification active on profile: %s",
                record.company_name,
            )

            page.wait_for_timeout(3000)

        soup = BeautifulSoup(
            page.content(),
            "html.parser",
        )

        jsonld = find_exhibitor_jsonld(soup)

        name = clean(jsonld.get("name"))

        if name and "koelnmesse" not in name.lower():
            record.company_name = name

        record.email = extract_email(
            soup,
            jsonld,
        )
        record.website = extract_website(
            soup,
            jsonld,
        )
        record.telephone = clean(
            jsonld.get("telephone")
        )
        record.address = format_address(
            jsonld.get("address")
        )

        if not record.telephone:
            tel_link = soup.select_one(
                'a[href^="tel:"]'
            )

            if tel_link:
                record.telephone = clean(
                    tel_link.get("href", "")
                ).replace("tel:", "")

        return record

    finally:
        page.close()


def load_checkpoint() -> dict[str, Exhibitor]:
    if not CHECKPOINT_FILE.exists():
        return {}

    dataframe = pd.read_csv(
        CHECKPOINT_FILE,
        dtype=str,
    ).fillna("")

    records: dict[str, Exhibitor] = {}

    for row in dataframe.to_dict(
        orient="records"
    ):
        record = Exhibitor(
            company_name=row.get(
                "company_name",
                "",
            ),
            country=row.get(
                "country",
                "",
            ),
            hall_stand=row.get(
                "hall_stand",
                "",
            ),
            website=row.get(
                "website",
                "",
            ),
            email=row.get(
                "email",
                "",
            ),
            address=row.get(
                "address",
                "",
            ),
            telephone=row.get(
                "telephone",
                "",
            ),
            profile_link=row.get(
                "profile_link",
                "",
            ),
            source_url=row.get(
                "source_url",
                "",
            ),
            list_offset=int(
                row.get("list_offset", "0")
                or 0
            ),
        )

        if record.profile_link:
            records[
                record.profile_link.lower()
                .rstrip("/")
            ] = record

    return records


def save_checkpoint(
    records: dict[str, Exhibitor],
) -> None:
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    dataframe = pd.DataFrame(
        [
            asdict(record)
            for record in records.values()
        ]
    )

    dataframe.to_csv(
        CHECKPOINT_FILE,
        index=False,
        encoding="utf-8-sig",
    )


def export_excel(
    records: list[Exhibitor],
) -> None:
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    columns = [
        "company_name",
        "country",
        "hall_stand",
        "website",
        "email",
        "address",
        "telephone",
        "profile_link",
        "source_url",
        "list_offset",
    ]

    dataframe = pd.DataFrame(
        [asdict(record) for record in records],
        columns=columns,
    )

    if not dataframe.empty:
        dataframe = dataframe.sort_values(
            by=["company_name", "country"],
            na_position="last",
        )

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:
        dataframe.to_excel(
            writer,
            sheet_name="IDS Exhibitors",
            index=False,
        )

        worksheet = writer.sheets[
            "IDS Exhibitors"
        ]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        widths = {
            "A": 42,
            "B": 24,
            "C": 26,
            "D": 52,
            "E": 42,
            "F": 70,
            "G": 24,
            "H": 70,
            "I": 70,
            "J": 14,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        for row in worksheet.iter_rows(
            min_row=2
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    logging.info(
        "Excel saved: %s",
        OUTPUT_FILE.resolve(),
    )




def human_delay(
    minimum: float,
    maximum: float,
    label: str,
) -> None:
    delay = random.uniform(
        minimum,
        maximum,
    )

    logging.info(
        "%s delay: %.1f seconds",
        label,
        delay,
    )

    time.sleep(delay)


def wait_for_result_list_ready(
    page,
    timeout_seconds: int = 60,
) -> None:
    deadline = time.time() + timeout_seconds

    while time.time() < deadline:
        try:
            count = page.locator(
                ".esr.search-results .item "
                ".col1ergebnis a.db-aslink"
            ).count()

            if count > 0:
                return

        except Exception:
            pass

        page.wait_for_timeout(1000)

    raise RuntimeError(
        "IDS result list did not recover before timeout"
    )


def load_ajax_page(
    page,
    offset: int,
    timeout_seconds: int = 75,
) -> str:
    """
    Load an IDS page through blaettern(offset), with retries and recovery.
    """
    if offset == 0:
        html_text = page.content()

        if not has_exhibitor_results(
            html_text
        ):
            wait_for_results(
                page,
                LOGIN_WAIT_SECONDS,
            )
            html_text = page.content()

        return html_text

    last_error = None

    for attempt in range(
        1,
        PAGINATION_RETRIES + 1,
    ):
        try:
            wait_for_result_list_ready(
                page,
                timeout_seconds=60,
            )

            locator = page.locator(
                ".esr.search-results .item "
                ".col1ergebnis a.db-aslink"
            ).first

            first_link_before = (
                locator.get_attribute(
                    "href",
                    timeout=60000,
                )
            )

            function_exists = page.evaluate(
                "() => typeof window.blaettern === 'function'"
            )

            if not function_exists:
                raise RuntimeError(
                    "IDS blaettern() is unavailable"
                )

            if attempt > 1:
                human_delay(
                    8.0,
                    16.0,
                    f"Pagination retry {attempt}",
                )

            logging.info(
                "Calling IDS AJAX paginator: "
                "blaettern('%d') | attempt %d/%d",
                offset,
                attempt,
                PAGINATION_RETRIES,
            )

            page.evaluate(
                "(offset) => "
                "window.blaettern(String(offset))",
                offset,
            )

            deadline = (
                time.time()
                + timeout_seconds
            )

            while time.time() < deadline:
                page.wait_for_timeout(1000)

                try:
                    current_items = page.locator(
                        ".esr.search-results .item "
                        ".col1ergebnis a.db-aslink"
                    )

                    if current_items.count() == 0:
                        continue

                    first_link_after = (
                        current_items.first
                        .get_attribute(
                            "href",
                            timeout=10000,
                        )
                    )

                except Exception:
                    continue

                if (
                    first_link_after
                    and first_link_after
                    != first_link_before
                ):
                    logging.info(
                        "AJAX page loaded: "
                        "offset=%d | first=%s",
                        offset,
                        first_link_after,
                    )

                    page.wait_for_timeout(
                        PAGE_WAIT_MS
                    )

                    return page.content()

            raise RuntimeError(
                "AJAX result list did not change"
            )

        except Exception as exc:
            last_error = exc

            logging.warning(
                "Pagination attempt %d/%d failed "
                "for offset %d: %s",
                attempt,
                PAGINATION_RETRIES,
                offset,
                exc,
            )

            if attempt < PAGINATION_RETRIES:
                logging.info(
                    "Reloading IDS list page before retry..."
                )

                page.goto(
                    LIST_URL,
                    wait_until="domcontentloaded",
                    timeout=90000,
                )

                wait_for_results(
                    page,
                    LOGIN_WAIT_SECONDS,
                )

                human_delay(
                    10.0,
                    18.0,
                    "Recovery",
                )

    raise RuntimeError(
        f"IDS pagination failed for offset "
        f"{offset} after "
        f"{PAGINATION_RETRIES} attempts: "
        f"{last_error}"
    )



def scrape(
    *,
    max_pages: int | None,
    fresh: bool,
) -> list[Exhibitor]:
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )
    PROFILE_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    if fresh and CHECKPOINT_FILE.exists():
        CHECKPOINT_FILE.unlink()

    existing = load_checkpoint()

    logging.info(
        "Checkpoint records: %d",
        len(existing),
    )

    completed_offsets = {
        record.list_offset
        for record in existing.values()
    }

    with sync_playwright() as playwright:
        context = (
            playwright.chromium
            .launch_persistent_context(
                user_data_dir=str(
                    PROFILE_DIR.resolve()
                ),
                headless=False,
                channel="chrome",
                viewport=None,
                args=[
                    "--start-maximized",
                    (
                        "--disable-blink-features="
                        "AutomationControlled"
                    ),
                ],
            )
        )

        page = (
            context.pages[0]
            if context.pages
            else context.new_page()
        )

        page.goto(
            LIST_URL,
            wait_until="domcontentloaded",
            timeout=90000,
        )

        print(
            "\nChrome is open.\n"
            "Log in manually if IDS asks for Google login.\n"
            "Do not put the password in this script or terminal.\n"
        )

        wait_for_results(
            page,
            LOGIN_WAIT_SECONDS,
        )

        first_html = page.content()
        visible_offsets = extract_offsets(first_html)

        if not visible_offsets:
            visible_offsets = [0]

        max_offset = max(visible_offsets)

        # IDS currently exposes the final offset in the pager.
        # Example: 2000 means offsets 0..2000 = 101 pages.
        total_pages = (max_offset // PAGE_SIZE) + 1

        if max_pages is not None:
            total_pages = min(
                total_pages,
                max_pages,
            )

        logging.info(
            "Detected final offset: %d",
            max_offset,
        )
        logging.info(
            "Dynamic total pages: %d",
            total_pages,
        )

        offsets = [
            page_index * PAGE_SIZE
            for page_index in range(total_pages)
        ]

        for page_index, offset in enumerate(
            offsets,
            start=1,
        ):
            current_url = build_list_url(offset)

            logging.info(
                "List page %d/%d | offset=%d",
                page_index,
                total_pages,
                offset,
            )

            # Page 1 must be parsed from the initial loaded document.
            # Every later page uses IDS's own AJAX function.
            try:
                html_text = load_ajax_page(
                    page,
                    offset,
                )
            except Exception as exc:
                logging.warning(
                    "Pagination failed at offset %d: %s",
                    offset,
                    exc,
                )
                continue

            if not has_exhibitor_results(
                html_text
            ):
                logging.warning(
                    "No exhibitor records at offset %d",
                    offset,
                )
                continue

            records = parse_list_page(
                html_text,
                current_url,
                offset,
            )

            logging.info(
                "Exhibitors found: %d",
                len(records),
            )

            if not records:
                continue

            new_records = 0

            for index, record in enumerate(
                records,
                start=1,
            ):
                key = (
                    record.profile_link.lower()
                    .rstrip("/")
                )

                if key in existing:
                    logging.info(
                        "[%d/%d] Already saved: %s",
                        index,
                        len(records),
                        record.company_name,
                    )
                    continue

                logging.info(
                    "[%d/%d] %s",
                    index,
                    len(records),
                    record.company_name,
                )

                try:
                    record = scrape_profile(
                        context,
                        record,
                    )

                    logging.info(
                        "Saved | %s | %s",
                        record.company_name,
                        record.email or "No email",
                    )

                except Exception as exc:
                    logging.warning(
                        "Profile failed: %s | %s",
                        record.profile_link,
                        exc,
                    )

                existing[key] = record
                new_records += 1

                save_checkpoint(
                    existing
                )

                human_delay(
                    MIN_PROFILE_DELAY,
                    MAX_PROFILE_DELAY,
                    "Profile",
                )

            export_excel(
                list(existing.values())
            )

            completed_offsets.add(offset)

            logging.info(
                "Completed page %d/%d | new=%d | total=%d",
                page_index,
                total_pages,
                new_records,
                len(existing),
            )

            if (
                page_index
                % LONG_BREAK_EVERY_PAGES
                == 0
            ):
                human_delay(
                    LONG_BREAK_MIN,
                    LONG_BREAK_MAX,
                    "Long rest",
                )
            else:
                human_delay(
                    MIN_PAGE_DELAY,
                    MAX_PAGE_DELAY,
                    "Page",
                )

        context.close()

    return list(existing.values())



def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Scrape IDS Cologne dental exhibitors using "
            "the exact IDS list and profile HTML structure."
        )
    )

    parser.add_argument(
        "--max-pages",
        type=int,
        default=None,
        help="Limit pages for testing.",
    )

    parser.add_argument(
        "--email-only",
        action="store_true",
    )

    parser.add_argument(
        "--fresh",
        action="store_true",
    )

    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format=(
            "%(asctime)s | %(levelname)s | "
            "%(message)s"
        ),
    )

    records = scrape(
        max_pages=args.max_pages,
        fresh=args.fresh,
    )

    if args.email_only:
        records = [
            record
            for record in records
            if record.email
        ]

    export_excel(records)

    logging.info(
        "Final exhibitors exported: %d",
        len(records),
    )
    logging.info(
        "Rows with email: %d",
        sum(
            bool(record.email)
            for record in records
        ),
    )
    logging.info(
        "Rows with telephone: %d",
        sum(
            bool(record.telephone)
            for record in records
        ),
    )
    logging.info(
        "Rows with website: %d",
        sum(
            bool(record.website)
            for record in records
        ),
    )


if __name__ == "__main__":
    main()