from __future__ import annotations

import argparse
import io
import json
import re
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import unquote, urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright


DEFAULT_CDP_URL = "http://127.0.0.1:9222"
DEFAULT_OLLAMA_URL = "http://127.0.0.1:11434"
DEFAULT_OLLAMA_MODEL = "qwen3:4b"

REQUEST_TIMEOUT = 60
PAGE_TIMEOUT_MS = 60_000
OLLAMA_TIMEOUT = 180

EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)
DOI_RE = re.compile(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+\b", re.I)
YEAR_RE = re.compile(r"\b(?:19|20)\d{2}\b", re.I)

OUTPUT_COLUMNS = [
    "employee_name",
    "employee_email",
    "file_name",
    "source_url",
    "journal_name",
    "volume",
    "issue",
    "year",
    "issue_url",
    "article_title",
    "article_url",
    "pdf_url",
    "doi",
    "author_name",
    "email",
    "affiliation",
    "country",
    "is_corresponding_author",
    "author_source",
    "email_source",
    "mapping_method",
    "mapping_confidence",
    "status",
    "error",
]


@dataclass
class Issue:
    volume: str = ""
    issue: str = ""
    year: int | None = None
    url: str = ""


@dataclass
class Article:
    title: str = ""
    url: str = ""
    pdf_url: str = ""
    doi: str | None = None
    volume: str = ""
    issue: str = ""
    year: int | None = None
    issue_url: str = ""


@dataclass
class AuthorCandidate:
    name: str
    affiliation: str = ""
    country: str = ""
    source: str = "html"


@dataclass
class EmailCandidate:
    email: str
    context: str = ""
    source: str = "html"
    native_author_name: str = ""
    affiliation: str = ""
    country: str = ""
    is_corresponding_author: bool = False


@dataclass
class OutputRow:
    employee_name: str = ""
    employee_email: str = ""
    file_name: str = ""
    source_url: str = ""
    journal_name: str = ""
    volume: str = ""
    issue: str = ""
    year: int | None = None
    issue_url: str = ""
    article_title: str = ""
    article_url: str = ""
    pdf_url: str = ""
    doi: str | None = None
    author_name: str = ""
    email: str = ""
    affiliation: str = ""
    country: str = ""
    is_corresponding_author: bool = False
    author_source: str = ""
    email_source: str = ""
    mapping_method: str = ""
    mapping_confidence: float = 0.0
    status: str = ""
    error: str = ""


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def make_soup(html: str) -> BeautifulSoup:
    try:
        return BeautifulSoup(html or "", "lxml")
    except Exception:
        return BeautifulSoup(html or "", "html.parser")


def normalize_url(url: str, base: str = "", drop_query: bool = False) -> str:
    absolute = urljoin(base or url, url)
    p = urlparse(absolute)
    return urlunparse(
        (
            p.scheme or "https",
            p.netloc.lower(),
            p.path.rstrip("/") or "/",
            "",
            "" if drop_query else p.query,
            "",
        )
    )


def same_host(a: str, b: str) -> bool:
    ha = (urlparse(a).hostname or "").lower()
    hb = (urlparse(b).hostname or "").lower()
    return ha == hb or ha.endswith("." + hb) or hb.endswith("." + ha)


def uniq(values: Iterable[str]) -> list[str]:
    out = []
    seen = set()
    for value in values:
        value = clean_text(value)
        if not value:
            continue
        key = value.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(value)
    return out


def normalize_email(value: str) -> str:
    value = re.sub(r"^mailto:", "", clean_text(value), flags=re.I).split("?", 1)[0]
    m = EMAIL_RE.search(value)
    return m.group(0).lower() if m else ""


def normalize_doi(value: str | None) -> str | None:
    text = unquote(clean_text(value))
    text = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", text, flags=re.I)
    text = re.sub(r"^doi:\s*", "", text, flags=re.I)
    m = DOI_RE.search(text)
    return m.group(0).rstrip(").,;").lower() if m else None


def extract_year(text: str) -> int | None:
    m = YEAR_RE.search(text or "")
    return int(m.group(0)) if m else None


def extract_country(affiliation: str) -> str:
    parts = [x.strip(" .;") for x in clean_text(affiliation).split(",") if x.strip(" .;")]
    return parts[-1] if parts else ""


def first_meta(soup: BeautifulSoup, *names: str) -> str:
    for name in names:
        node = soup.select_one(f"meta[name='{name}'],meta[property='{name}']")
        if node:
            value = clean_text(node.get("content"))
            if value:
                return value
    return ""


def all_meta(soup: BeautifulSoup, name: str) -> list[str]:
    return uniq(
        clean_text(node.get("content"))
        for node in soup.select(f"meta[name='{name}']")
        if clean_text(node.get("content"))
    )


def name_tokens(name: str) -> set[str]:
    text = re.sub(r"[^a-z0-9]+", " ", clean_text(name).casefold())
    stop = {"dr", "prof", "professor", "mr", "mrs", "ms", "md", "phd", "mbbs"}
    return {t for t in text.split() if len(t) >= 2 and t not in stop}


def names_match(a: str, b: str) -> bool:
    aa = name_tokens(a)
    bb = name_tokens(b)
    if not aa or not bb:
        return False
    if aa == bb:
        return True
    overlap = aa & bb
    return len(overlap) >= 2 and (aa <= bb or bb <= aa)


def _normalized_name_text(value: str) -> str:
    return clean_text(
        re.sub(
            r"[^a-z0-9]+",
            " ",
            clean_text(value).casefold(),
        )
    )


def author_context_distance(
    author_name: str,
    context: str,
    email: str,
) -> int | None:
    if not author_name or not context:
        return None

    raw = clean_text(context).casefold()
    email_pos = raw.find(clean_text(email).casefold())
    if email_pos < 0:
        email_pos = len(raw) // 2

    name_raw = clean_text(author_name).casefold()
    positions = []
    start = 0
    while True:
        pos = raw.find(name_raw, start)
        if pos < 0:
            break
        positions.append(pos)
        start = pos + max(1, len(name_raw))

    if positions:
        return min(abs(pos - email_pos) for pos in positions)

    tokens = [
        t for t in _normalized_name_text(author_name).split()
        if len(t) >= 3
    ]
    if len(tokens) < 2:
        return None

    first = tokens[0]
    last = tokens[-1]

    first_positions = [m.start() for m in re.finditer(rf"\\b{re.escape(first)}\\b", raw)]
    last_positions = [m.start() for m in re.finditer(rf"\\b{re.escape(last)}\\b", raw)]

    if not first_positions or not last_positions:
        return None

    best = None
    for fp in first_positions:
        for lp in last_positions:
            if abs(fp - lp) > 140:
                continue
            midpoint = (fp + lp) // 2
            distance = abs(midpoint - email_pos)
            best = distance if best is None else min(best, distance)

    return best


def supported_authors_for_email(
    authors: list["AuthorCandidate"],
    email: "EmailCandidate",
) -> list[tuple[int, "AuthorCandidate"]]:
    supported: list[tuple[int, AuthorCandidate]] = []

    for author in authors:
        distance = author_context_distance(
            author.name,
            email.context,
            email.email,
        )
        if distance is not None:
            supported.append((distance, author))

    supported.sort(key=lambda item: item[0])
    return supported


NAVIGATION_TITLES = {
    "apcs payment",
    "article processing charges",
    "case report journal articles",
    "contact",
    "current issue",
    "home",
    "peer reviewed articles",
    "research journal articles",
    "review journal articles",
    "search",
    "search articles",
    "submit article",
    "trending articles",
}

NAVIGATION_PATH_MARKERS = (
    "/about",
    "/contact",
    "/processingcharges",
    "/search",
    "/submit",
    "/editorial",
)

GENERIC_MAILBOX_PREFIXES = {
    "admin",
    "contact",
    "editor",
    "editorial",
    "editorinchief",
    "info",
    "journal",
    "office",
    "publisher",
    "secretary",
    "submissions",
    "support",
    "webmaster",
}


def is_navigation_candidate(title: str, url: str) -> bool:
    label = clean_text(title).casefold().strip(" .:-")
    path = urlparse(url).path.casefold()

    if label in NAVIGATION_TITLES:
        return True

    if any(marker in path for marker in NAVIGATION_PATH_MARKERS):
        return True

    return False


def is_generic_mailbox(email: str) -> bool:
    local = clean_text(email).casefold().split("@", 1)[0]
    local = re.sub(r"[^a-z]", "", local)
    return local in GENERIC_MAILBOX_PREFIXES


def has_scholarly_jsonld(soup: BeautifulSoup) -> bool:
    for node in soup.select("script[type='application/ld+json']"):
        raw = node.string or node.get_text()
        if not raw:
            continue
        try:
            payload = json.loads(raw)
        except Exception:
            continue

        objects = payload if isinstance(payload, list) else [payload]
        for obj in objects:
            if not isinstance(obj, dict):
                continue
            kind = clean_text(obj.get("@type")).casefold()
            if any(
                token in kind
                for token in (
                    "scholarlyarticle",
                    "article",
                    "medicalscholarlyarticle",
                )
            ):
                return True
    return False


def sanitize_filename(value: str, fallback: str) -> str:
    value = clean_text(value) or fallback
    value = re.sub(r'[<>:"/\\\\|?*]+', "_", value)
    value = value.strip(" ._")
    return value[:180] or fallback


class BrowserSession:
    def __init__(self, cdp_url: str):
        self.cdp_url = cdp_url
        self._pw = None
        self.browser = None
        self.context = None

    def __enter__(self):
        self._pw = sync_playwright().start()
        self.browser = self._pw.chromium.connect_over_cdp(self.cdp_url)
        self.context = self.browser.contexts[0] if self.browser.contexts else self.browser.new_context()
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            if self._pw:
                self._pw.stop()
        except Exception:
            pass

    def render(self, url: str) -> tuple[str, str]:
        page = self.context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
            return page.url, page.content()
        finally:
            page.close()


class Fetcher:
    def __init__(self, browser: BrowserSession | None):
        self.browser = browser
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/151 Safari/537.36"
                )
            }
        )

    def get(self, url: str, referer: str = "") -> tuple[str, str]:
        headers = {"Referer": referer} if referer else {}
        try:
            r = self.session.get(url, headers=headers, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            text = r.text or ""
            blocked = (
                r.status_code >= 400
                or len(text) < 800
                or "captcha" in text.lower()
                or "access denied" in text.lower()
                or "cf-chl-" in text.lower()
                or "checking your browser" in text.lower()
            )
            if not blocked:
                return r.url, text
        except requests.RequestException:
            pass

        if not self.browser:
            raise RuntimeError(f"Unable to fetch: {url}")

        return self.browser.render(url)

    def get_binary(self, url: str, referer: str = "") -> bytes:
        headers = {"Referer": referer} if referer else {}
        r = self.session.get(url, headers=headers, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        r.raise_for_status()
        return r.content


class HierarchyDetector:
    """
    Publisher-independent hierarchy detector.

    It classifies links by structure and surrounding text instead of fixing
    behavior to one domain/publisher.
    """

    ISSUE_URL_PATTERNS = [
        re.compile(r"/issue/view/\d+", re.I),
        re.compile(r"/issue/\d+", re.I),
        re.compile(r"/issues/\d+", re.I),
        re.compile(r"/volume/[^/]+/issue/[^/]+", re.I),
        re.compile(r"/vol/[^/]+/(?:issue|suppl)/[^/]+", re.I),
        re.compile(r"/toc/[^/]+/[^/]+/[^/]+", re.I),
        re.compile(r"[?&](?:issue|issueid|issue_id)=", re.I),
    ]

    ARTICLE_URL_PATTERNS = [
        re.compile(r"/article/view/\d+", re.I),
        re.compile(r"/article/", re.I),
        re.compile(r"/articles/", re.I),
        re.compile(r"/doi/(?:full|abs|pdf|epdf)?/?10\.", re.I),
        re.compile(r"/science/article/", re.I),
        re.compile(r"/fulltext/", re.I),
        re.compile(r"/abstract/", re.I),
        re.compile(r"/content/", re.I),
        re.compile(r"[?&](?:article|articleid|article_id)=", re.I),
    ]

    PDF_PATTERNS = [
        re.compile(r"\.pdf(?:$|\?)", re.I),
        re.compile(r"/pdf/", re.I),
        re.compile(r"/epdf/", re.I),
        re.compile(r"/article/download/", re.I),
        re.compile(r"/galley/", re.I),
    ]

    def looks_like_issue(self, href: str, text: str) -> bool:
        if any(p.search(href) for p in self.ISSUE_URL_PATTERNS):
            return True
        return bool(
            re.search(
                r"\b(volume|vol\.?|issue|no\.?|том|№)\s*[0-9ivxlc]+",
                text,
                re.I,
            )
        )

    def looks_like_article(self, href: str, text: str) -> bool:
        if any(p.search(href) for p in self.PDF_PATTERNS):
            return False
        return any(p.search(href) for p in self.ARTICLE_URL_PATTERNS)

    def looks_like_pdf(self, href: str, text: str) -> bool:
        if any(p.search(href) for p in self.PDF_PATTERNS):
            return True
        return "pdf" in text.casefold()

    def parse_volume_issue_year(self, text: str, href: str) -> tuple[str, str, int | None]:
        joined = f"{clean_text(text)} {href}"

        volume_patterns = [
            r"(?:volume|vol\.?|том)\s*([A-Za-zА-Яа-я0-9.-]+)",
            r"/vol(?:ume)?/([^/?#]+)",
            r"[?&]volume=([^&#]+)",
        ]

        issue_patterns = [
            r"(?:issue|no\.?|№)\s*([A-Za-zА-Яа-я0-9.-]+)",
            r"/issue/([^/?#]+)",
            r"[?&]issue=([^&#]+)",
        ]

        volume = ""
        issue = ""

        for pattern in volume_patterns:
            m = re.search(pattern, joined, re.I)
            if m:
                volume = clean_text(m.group(1))
                break

        for pattern in issue_patterns:
            m = re.search(pattern, joined, re.I)
            if m:
                issue = clean_text(m.group(1))
                break

        # /toc/<code>/<volume>/<issue>
        if not volume or not issue:
            m = re.search(r"/toc/[^/]+/([^/?#]+)/([^/?#]+)", href, re.I)
            if m:
                volume = volume or clean_text(m.group(1))
                issue = issue or clean_text(m.group(2))

        # /vol/<volume>/suppl/C
        if not issue:
            m = re.search(r"/vol/([^/?#]+)/(?:issue|suppl)/([^/?#]+)", href, re.I)
            if m:
                volume = volume or clean_text(m.group(1))
                issue = clean_text(m.group(2))

        return volume, issue, extract_year(joined)


class GenericJournalCrawler:
    def __init__(self, fetcher: Fetcher):
        self.fetcher = fetcher
        self.detector = HierarchyDetector()

    def journal_name(self, soup: BeautifulSoup) -> str:
        return (
            first_meta(
                soup,
                "citation_journal_title",
                "dc.source",
                "og:site_name",
                "og:title",
            )
            or clean_text(soup.title.get_text(" ", strip=True) if soup.title else "")
        )

    def discover_issues(self, source_url: str, max_issues: int) -> tuple[str, list[Issue]]:
        final_url, html = self.fetcher.get(source_url)
        soup = make_soup(html)
        journal_name = self.journal_name(soup)

        found: dict[str, Issue] = {}

        for a in soup.select("a[href]"):
            href = normalize_url(a.get("href") or "", final_url)
            text = clean_text(a.get_text(" ", strip=True))

            if not same_host(href, final_url):
                continue

            if not self.detector.looks_like_issue(href, text):
                continue

            if normalize_url(href) == normalize_url(final_url):
                continue

            volume, issue, year = self.detector.parse_volume_issue_year(text, href)

            found[normalize_url(href, drop_query=False)] = Issue(
                volume=volume,
                issue=issue,
                year=year,
                url=href,
            )

        issues = list(found.values())

        # If source is already an issue page or archive page with direct article links,
        # use it as one pseudo issue.
        if not issues:
            direct_articles = self._discover_article_links_from_html(final_url, soup, limit=1)
            if direct_articles:
                issues = [Issue(url=final_url)]

        def n(value: str) -> int:
            m = re.search(r"\d+", value or "")
            return int(m.group(0)) if m else -1

        issues.sort(
            key=lambda x: (
                x.year or 0,
                n(x.volume),
                n(x.issue),
            ),
            reverse=True,
        )

        return journal_name, issues[:max_issues]

    def _discover_article_links_from_html(
        self,
        base_url: str,
        soup: BeautifulSoup,
        limit: int,
    ) -> list[Article]:
        found: dict[str, Article] = {}

        for a in soup.select("a[href]"):
            href = normalize_url(a.get("href") or "", base_url)
            text = clean_text(
                a.get("title")
                or a.get("aria-label")
                or a.get_text(" ", strip=True)
            )

            if not same_host(href, base_url):
                continue

            if not self.detector.looks_like_article(href, text):
                continue

            if is_navigation_candidate(text, href):
                continue

            if len(text) < 5:
                parent = a.find_parent(["article", "li", "div", "section"])
                heading = parent.select_one("h2,h3,h4,[class*='title']") if parent else None
                text = clean_text(
                    heading.get_text(" ", strip=True)
                    if heading else ""
                )

            if len(text) < 5:
                continue

            found[href] = Article(
                title=text,
                url=href,
            )

            if len(found) >= limit:
                break

        return list(found.values())

    def _enrich_issue_from_page(
        self,
        issue: Issue,
        soup: BeautifulSoup,
        final_url: str,
    ) -> Issue:
        heading = ""
        for selector in (
            "h1",
            ".page_title",
            ".current_issue_title",
            ".issue_title",
            ".title",
        ):
            node = soup.select_one(selector)
            if node:
                heading = clean_text(node.get_text(" ", strip=True))
                if heading:
                    break

        title_meta = first_meta(
            soup,
            "citation_journal_title",
            "dc.title",
            "og:title",
        )
        context = clean_text(f"{heading} {title_meta}")

        volume, issue_value, year = self.detector.parse_volume_issue_year(
            context,
            final_url,
        )

        # Do not accept routing words such as /issue/view/34 as issue='view'.
        if issue_value.casefold() in {"view", "archive", "current"}:
            issue_value = ""

        return Issue(
            volume=issue.volume or volume,
            issue=issue.issue if issue.issue.casefold() not in {"view", "archive", "current"} else "",
            year=issue.year or year,
            url=issue.url,
        )

    def _is_real_article_page(
        self,
        final_url: str,
        soup: BeautifulSoup,
    ) -> bool:
        citation_title = first_meta(soup, "citation_title")
        citation_authors = all_meta(soup, "citation_author")
        doi = normalize_doi(
            first_meta(
                soup,
                "citation_doi",
                "dc.identifier",
                "dc.identifier.doi",
            )
        )

        # Strong publisher-neutral evidence.
        if citation_title and citation_authors:
            return True

        if citation_title and doi:
            return True

        if has_scholarly_jsonld(soup):
            return True

        # Visible fallback: DOI + an author-like area.
        if doi and soup.select_one(
            ".authors, .author, .article-authors, "
            "[class*='author-name'], [class*='authorName']"
        ):
            return True

        return False

    def discover_articles(self, issue: Issue, max_articles: int) -> list[Article]:
        final_url, html = self.fetcher.get(issue.url)
        soup = make_soup(html)

        enriched_issue = self._enrich_issue_from_page(
            issue,
            soup,
            final_url,
        )
        issue.volume = enriched_issue.volume
        issue.issue = enriched_issue.issue
        issue.year = enriched_issue.year

        # If issue URL is actually an article page.
        if first_meta(soup, "citation_title"):
            return [
                Article(
                    title=first_meta(soup, "citation_title"),
                    url=final_url,
                    doi=normalize_doi(first_meta(soup, "citation_doi")),
                    volume=issue.volume,
                    issue=issue.issue,
                    year=issue.year,
                    issue_url=issue.url,
                )
            ]

        articles = self._discover_article_links_from_html(
            final_url,
            soup,
            max_articles,
        )

        validated: list[Article] = []

        for article in articles:
            article.volume = issue.volume
            article.issue = issue.issue
            article.year = issue.year
            article.issue_url = issue.url

            if is_navigation_candidate(article.title, article.url):
                continue

            try:
                candidate_final, candidate_html = self.fetcher.get(
                    article.url,
                    referer=issue.url,
                )
                candidate_soup = make_soup(candidate_html)

                if not self._is_real_article_page(
                    candidate_final,
                    candidate_soup,
                ):
                    continue

                # Prefer the actual article title from metadata.
                real_title = first_meta(
                    candidate_soup,
                    "citation_title",
                    "dc.title",
                    "og:title",
                )
                if real_title:
                    article.title = real_title

                validated.append(article)

                if len(validated) >= max_articles:
                    break

            except Exception:
                continue

        return validated


def extract_jsonld_authors(soup: BeautifulSoup) -> list[AuthorCandidate]:
    found: list[AuthorCandidate] = []

    for node in soup.select("script[type='application/ld+json']"):
        raw = node.string or node.get_text()
        if not raw:
            continue

        try:
            payload = json.loads(raw)
        except Exception:
            continue

        objects = payload if isinstance(payload, list) else [payload]

        for obj in objects:
            if not isinstance(obj, dict):
                continue

            authors = obj.get("author") or []
            if isinstance(authors, dict):
                authors = [authors]

            for a in authors:
                if isinstance(a, str):
                    name = clean_text(a)
                    if name:
                        found.append(AuthorCandidate(name=name, source="jsonld"))
                    continue

                if not isinstance(a, dict):
                    continue

                name = clean_text(a.get("name"))
                if not name:
                    continue

                affiliation = ""
                raw_aff = a.get("affiliation")

                if isinstance(raw_aff, dict):
                    affiliation = clean_text(raw_aff.get("name"))
                elif isinstance(raw_aff, list):
                    affiliation = "; ".join(
                        clean_text(x.get("name") if isinstance(x, dict) else x)
                        for x in raw_aff
                        if clean_text(x.get("name") if isinstance(x, dict) else x)
                    )
                elif raw_aff:
                    affiliation = clean_text(raw_aff)

                found.append(
                    AuthorCandidate(
                        name=name,
                        affiliation=affiliation,
                        country=extract_country(affiliation),
                        source="jsonld",
                    )
                )

    return found


def extract_pdf_text(data: bytes, max_pages: int = 5) -> str:
    parts: list[str] = []

    try:
        import pymupdf  # type: ignore

        doc = pymupdf.open(stream=data, filetype="pdf")
        for i in range(min(max_pages, len(doc))):
            text = clean_text(doc[i].get_text("text"))
            if text:
                parts.append(text)
        doc.close()

        if parts:
            return "\n".join(parts)
    except Exception:
        pass

    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(io.BytesIO(data))
        for page in reader.pages[:max_pages]:
            text = clean_text(page.extract_text() or "")
            if text:
                parts.append(text)
        if parts:
            return "\n".join(parts)
    except Exception:
        pass

    return ""


class ArticleExtractor:
    def __init__(self, fetcher: Fetcher):
        self.fetcher = fetcher
        self.detector = HierarchyDetector()

    def extract(
        self,
        article: Article,
        use_pdf: bool,
        pdf_max_pages: int,
    ) -> tuple[str, str | None, list[AuthorCandidate], list[EmailCandidate], str]:
        final_url, html = self.fetcher.get(article.url, referer=article.issue_url)
        soup = make_soup(html)

        title = (
            first_meta(soup, "citation_title", "dc.title", "og:title")
            or clean_text(
                soup.select_one("h1").get_text(" ", strip=True)
                if soup.select_one("h1")
                else article.title
            )
        )

        doi = normalize_doi(
            first_meta(
                soup,
                "citation_doi",
                "dc.identifier",
                "dc.identifier.doi",
            )
        )
        if not doi:
            doi = normalize_doi(soup.get_text(" ", strip=True))

        author_names = all_meta(soup, "citation_author")
        affiliations = all_meta(soup, "citation_author_institution")

        authors: list[AuthorCandidate] = []

        for i, name in enumerate(author_names):
            aff = affiliations[i] if i < len(affiliations) else ""
            authors.append(
                AuthorCandidate(
                    name=name,
                    affiliation=aff,
                    country=extract_country(aff),
                    source="html_metadata",
                )
            )

        for extra in extract_jsonld_authors(soup):
            if not any(names_match(x.name, extra.name) for x in authors):
                authors.append(extra)

        visible_author_selectors = [
            ".authors .name",
            ".authors .author",
            ".author .name",
            ".article-authors .author",
            ".entryAuthor",
            "[class*='author-name']",
            "[class*='authorName']",
        ]

        for selector in visible_author_selectors:
            for node in soup.select(selector):
                name = clean_text(node.get_text(" ", strip=True))
                if (
                    3 <= len(name) <= 180
                    and "@" not in name
                    and not any(names_match(x.name, name) for x in authors)
                ):
                    authors.append(
                        AuthorCandidate(
                            name=name,
                            source="html_visible",
                        )
                    )

        emails: dict[str, EmailCandidate] = {}

        for node in soup.select("a[href^='mailto:']"):
            email = normalize_email(node.get("href") or "")
            if not email:
                continue

            parent = node.find_parent(["p", "li", "div", "section", "article", "td"])
            context = clean_text(
                parent.get_text(" ", strip=True)
                if parent
                else node.parent.get_text(" ", strip=True)
                if node.parent
                else ""
            )

            corresponding = bool(
                re.search(
                    r"correspond(?:ing|ence)|contact author",
                    context,
                    re.I,
                )
            )

            # Header/footer/site mailboxes are not author emails.
            if is_generic_mailbox(email) and not corresponding:
                continue

            emails[email] = EmailCandidate(
                email=email,
                context=context,
                source="mailto",
                is_corresponding_author=corresponding,
            )

        body_text = clean_text(soup.get_text(" ", strip=True))

        for m in EMAIL_RE.finditer(body_text):
            email = m.group(0).lower()
            if email in emails:
                continue

            start = max(0, m.start() - 700)
            end = min(len(body_text), m.end() + 700)
            context = clean_text(body_text[start:end])

            corresponding = bool(
                re.search(
                    r"correspond(?:ing|ence)|contact author",
                    context,
                    re.I,
                )
            )

            if is_generic_mailbox(email) and not corresponding:
                continue

            emails[email] = EmailCandidate(
                email=email,
                context=context,
                source="html_visible",
                is_corresponding_author=corresponding,
            )

        pdf_url = first_meta(soup, "citation_pdf_url")

        if pdf_url:
            pdf_url = normalize_url(pdf_url, final_url)

        if not pdf_url:
            for a in soup.select("a[href]"):
                href = a.get("href") or ""
                text = clean_text(
                    a.get("title")
                    or a.get_text(" ", strip=True)
                )

                if self.detector.looks_like_pdf(href, text):
                    pdf_url = normalize_url(href, final_url)
                    break

        if use_pdf and pdf_url:
            try:
                data = self.fetcher.get_binary(pdf_url, referer=final_url)
                pdf_text = extract_pdf_text(data, max_pages=pdf_max_pages)

                for m in EMAIL_RE.finditer(pdf_text):
                    email = m.group(0).lower()
                    start = max(0, m.start() - 500)
                    end = min(len(pdf_text), m.end() + 500)
                    context = clean_text(pdf_text[start:end])

                    corresponding = bool(
                        re.search(
                            r"correspond(?:ing|ence)|e-?mail|contact author",
                            context,
                            re.I,
                        )
                    )

                    if is_generic_mailbox(email) and not corresponding:
                        continue

                    if email not in emails:
                        emails[email] = EmailCandidate(
                            email=email,
                            context=context,
                            source="pdf",
                            is_corresponding_author=corresponding,
                        )
            except Exception:
                pass

        return (
            title,
            doi or article.doi,
            authors,
            list(emails.values()),
            pdf_url,
        )


class OllamaMapper:
    def __init__(self, url: str, model: str, enabled: bool):
        self.url = url.rstrip("/")
        self.model = model
        self.enabled = enabled

    def map(
        self,
        article_title: str,
        authors: list[AuthorCandidate],
        emails: list[EmailCandidate],
    ) -> list[dict[str, Any]]:
        if not self.enabled or not authors or not emails:
            return []

        evidence: dict[str, list[str]] = {}

        for email in emails:
            supported = supported_authors_for_email(
                authors,
                email,
            )
            candidate_names = [
                author.name
                for distance, author in supported
                if distance <= 500
            ]
            if candidate_names:
                evidence[email.email.casefold()] = candidate_names

        if not evidence:
            return []

        prompt_emails = []
        for email in emails:
            allowed = evidence.get(
                email.email.casefold(),
                [],
            )
            if not allowed:
                continue

            prompt_emails.append(
                {
                    "email": email.email,
                    "context": email.context[:2200],
                    "source": email.source,
                    "allowed_author_names": allowed,
                }
            )

        payload = {
            "task": (
                "Map each email only to one of its allowed_author_names. "
                "Use explicit correspondence evidence in the supplied context. "
                "Never invent an author or email. "
                "If the context does not clearly identify one author, "
                "return author_name as null."
            ),
            "article_title": article_title,
            "emails": prompt_emails,
            "required_json": {
                "matches": [
                    {
                        "author_name": "one allowed_author_names value or null",
                        "email": "one supplied email",
                        "confidence": 0.0,
                        "reason": "brief evidence",
                    }
                ]
            },
        }

        r = requests.post(
            f"{self.url}/api/chat",
            json={
                "model": self.model,
                "stream": False,
                "format": "json",
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "You are a strict bibliographic entity-resolution engine. "
                            "You may only select an author explicitly listed in "
                            "allowed_author_names for that email. Return JSON only."
                        ),
                    },
                    {
                        "role": "user",
                        "content": json.dumps(
                            payload,
                            ensure_ascii=False,
                        ),
                    },
                ],
                "options": {"temperature": 0},
            },
            timeout=OLLAMA_TIMEOUT,
        )
        r.raise_for_status()

        response_json = r.json()
        content = (
            response_json.get("message", {}).get("content")
            or response_json.get("response")
            or "{}"
        )

        data = json.loads(content)
        valid: list[dict[str, Any]] = []

        for item in data.get("matches") or []:
            email_value = clean_text(
                item.get("email")
            ).casefold()

            allowed_names = evidence.get(
                email_value,
                [],
            )
            if not allowed_names:
                continue

            author_value = clean_text(
                item.get("author_name")
            )
            if not author_value:
                continue

            canonical_author = next(
                (
                    name
                    for name in allowed_names
                    if name.casefold() == author_value.casefold()
                ),
                None,
            )
            if not canonical_author:
                continue

            try:
                confidence = float(
                    item.get("confidence", 0)
                )
            except Exception:
                confidence = 0.0

            if confidence < 0.95:
                continue

            valid.append(
                {
                    "author_name": canonical_author,
                    "email": email_value,
                    "confidence": min(confidence, 1.0),
                }
            )

        return valid


def deterministic_mapping(
    authors: list[AuthorCandidate],
    emails: list[EmailCandidate],
) -> tuple[list[dict[str, Any]], list[EmailCandidate]]:
    mapped: list[dict[str, Any]] = []
    unresolved: list[EmailCandidate] = []

    for email in emails:
        supported = supported_authors_for_email(
            authors,
            email,
        )

        if len(supported) == 1:
            distance, author = supported[0]
            if distance <= 350:
                mapped.append(
                    {
                        "author": author,
                        "email": email,
                        "method": (
                            "pdf_correspondence"
                            if email.source == "pdf"
                            else "explicit_email_context"
                        ),
                        "confidence": 0.99,
                    }
                )
                continue

        elif len(supported) >= 2:
            best_distance, best_author = supported[0]
            second_distance, _ = supported[1]

            if (
                best_distance <= 220
                and second_distance - best_distance >= 100
            ):
                mapped.append(
                    {
                        "author": best_author,
                        "email": email,
                        "method": (
                            "pdf_correspondence_nearest"
                            if email.source == "pdf"
                            else "explicit_email_context_nearest"
                        ),
                        "confidence": 0.97,
                    }
                )
                continue

        unresolved.append(email)

    return mapped, unresolved


def rows_for_article(
    employee_name: str,
    employee_email: str,
    file_name: str,
    source_url: str,
    journal_name: str,
    article: Article,
    article_title: str,
    doi: str | None,
    pdf_url: str,
    authors: list[AuthorCandidate],
    emails: list[EmailCandidate],
    ollama: OllamaMapper,
    include_without_email: bool,
    include_unmapped: bool,
) -> list[OutputRow]:
    rows = []
    mapped, unresolved = deterministic_mapping(authors, emails)

    used_emails: set[str] = set()
    used_authors: set[str] = set()

    for item in mapped:
        author = item["author"]
        email = item["email"]

        rows.append(
            OutputRow(
                employee_name=employee_name,
                employee_email=employee_email,
                file_name=file_name,
                source_url=source_url,
                journal_name=journal_name,
                volume=article.volume,
                issue=article.issue,
                year=article.year,
                issue_url=article.issue_url,
                article_title=article_title,
                article_url=article.url,
                pdf_url=pdf_url,
                doi=doi,
                author_name=author.name,
                email=email.email,
                affiliation=author.affiliation,
                country=author.country,
                is_corresponding_author=email.is_corresponding_author,
                author_source=author.source,
                email_source=email.source,
                mapping_method=item["method"],
                mapping_confidence=item["confidence"],
                status="email_found",
            )
        )

        used_emails.add(email.email)
        used_authors.add(author.name.casefold())

    if unresolved and authors:
        try:
            llm_matches = ollama.map(article_title, authors, unresolved)
        except Exception as exc:
            print(f"      [WARN] Ollama mapping failed: {exc}")
            llm_matches = []

        unresolved_by_email = {
            x.email.casefold(): x
            for x in unresolved
        }

        for item in llm_matches:
            author_name = item.get("author_name")
            email_value = clean_text(
                item.get("email")
            ).casefold()
            confidence = float(item.get("confidence", 0))

            if (
                not author_name
                or confidence < 0.95
                or email_value in {
                    value.casefold()
                    for value in used_emails
                }
            ):
                continue

            author = next(
                (a for a in authors if a.name.casefold() == author_name.casefold()),
                None,
            )
            email = unresolved_by_email.get(email_value)

            if not author or not email:
                continue

            rows.append(
                OutputRow(
                    employee_name=employee_name,
                    employee_email=employee_email,
                    file_name=file_name,
                    source_url=source_url,
                    journal_name=journal_name,
                    volume=article.volume,
                    issue=article.issue,
                    year=article.year,
                    issue_url=article.issue_url,
                    article_title=article_title,
                    article_url=article.url,
                    pdf_url=pdf_url,
                    doi=doi,
                    author_name=author.name,
                    email=email.email,
                    affiliation=author.affiliation,
                    country=author.country,
                    is_corresponding_author=email.is_corresponding_author,
                    author_source=author.source,
                    email_source=email.source,
                    mapping_method="ollama",
                    mapping_confidence=confidence,
                    status="email_found",
                )
            )

            used_emails.add(email.email)
            used_authors.add(author.name.casefold())

    if include_unmapped:
        for email in emails:
            if email.email in used_emails:
                continue

            rows.append(
                OutputRow(
                    employee_name=employee_name,
                    employee_email=employee_email,
                    file_name=file_name,
                    source_url=source_url,
                    journal_name=journal_name,
                    volume=article.volume,
                    issue=article.issue,
                    year=article.year,
                    issue_url=article.issue_url,
                    article_title=article_title,
                    article_url=article.url,
                    pdf_url=pdf_url,
                    doi=doi,
                    email=email.email,
                    is_corresponding_author=email.is_corresponding_author,
                    email_source=email.source,
                    mapping_method="unmapped",
                    mapping_confidence=0.0,
                    status="email_unmapped",
                    error="Email found but author could not be safely mapped.",
                )
            )

    if include_without_email:
        for author in authors:
            if author.name.casefold() in used_authors:
                continue

            rows.append(
                OutputRow(
                    employee_name=employee_name,
                    employee_email=employee_email,
                    file_name=file_name,
                    source_url=source_url,
                    journal_name=journal_name,
                    volume=article.volume,
                    issue=article.issue,
                    year=article.year,
                    issue_url=article.issue_url,
                    article_title=article_title,
                    article_url=article.url,
                    pdf_url=pdf_url,
                    doi=doi,
                    author_name=author.name,
                    affiliation=author.affiliation,
                    country=author.country,
                    author_source=author.source,
                    mapping_method="no_email",
                    status="no_email",
                )
            )

    return rows


def read_input_excel(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [
        clean_text(x.value).lower()
        for x in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    if "url" not in headers:
        raise ValueError("Input Excel must contain a 'url' column.")

    records: list[dict[str, str]] = []

    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {
            headers[i]: clean_text(values[i])
            for i in range(min(len(headers), len(values)))
            if headers[i]
        }

        if not row.get("url"):
            continue

        records.append(
            {
                "url": row.get("url", ""),
                "employee_name": row.get("employee_name", ""),
                "employee_email": row.get("employee_email", ""),
                "file_name": row.get("file_name", ""),
            }
        )

    return records


def write_output(path: Path, rows: list[OutputRow], errors: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(OUTPUT_COLUMNS)

    for row in rows:
        data = asdict(row)
        ws.append([data.get(col, "") if data.get(col) is not None else "" for col in OUTPUT_COLUMNS])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 55, "B": 40, "C": 10, "D": 10, "E": 10, "F": 60,
        "G": 70, "H": 70, "I": 70, "J": 36, "K": 32, "L": 38,
        "M": 60, "N": 24, "O": 18, "P": 22, "Q": 22, "R": 24,
        "S": 20, "T": 18, "U": 55,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ews = wb.create_sheet("errors")
    ews.append(["source_url", "stage", "error"])
    for err in errors:
        ews.append([
            clean_text(err.get("source_url")),
            clean_text(err.get("stage")),
            clean_text(err.get("error")),
        ])

    ews.freeze_panes = "A2"
    ews.column_dimensions["A"].width = 60
    ews.column_dimensions["B"].width = 30
    ews.column_dimensions["C"].width = 100

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()

    parser.add_argument("--input", required=True)
    parser.add_argument(
        "--output-dir",
        default="output",
        help="Directory where one XLSX per input file_name will be written.",
    )
    parser.add_argument(
        "--combined-output",
        default="",
        help="Optional combined XLSX path.",
    )

    parser.add_argument("--cdp-url", default=DEFAULT_CDP_URL)
    parser.add_argument("--no-browser", action="store_true")

    parser.add_argument("--ollama-url", default=DEFAULT_OLLAMA_URL)
    parser.add_argument("--ollama-model", default=DEFAULT_OLLAMA_MODEL)
    parser.add_argument("--no-ollama", action="store_true")

    parser.add_argument("--max-issues", type=int, default=3)
    parser.add_argument("--max-articles", type=int, default=30)
    parser.add_argument("--pdf-max-pages", type=int, default=5)
    parser.add_argument("--no-pdf", action="store_true")
    parser.add_argument("--include-without-email", action="store_true")
    parser.add_argument(
        "--include-unmapped",
        action="store_true",
        help="Include emails that could not be safely mapped to an author.",
    )

    args = parser.parse_args()

    input_records = read_input_excel(Path(args.input))
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    browser_ctx = None
    browser = None
    all_rows: list[OutputRow] = []
    all_errors: list[dict[str, Any]] = []
    grouped_rows: dict[str, list[OutputRow]] = {}
    grouped_errors: dict[str, list[dict[str, Any]]] = {}

    ollama = OllamaMapper(
        args.ollama_url,
        args.ollama_model,
        enabled=not args.no_ollama,
    )

    try:
        if not args.no_browser:
            browser_ctx = BrowserSession(args.cdp_url)
            browser = browser_ctx.__enter__()
            print(f"[OK] Connected to Chrome CDP: {args.cdp_url}")

        fetcher = Fetcher(browser)
        crawler = GenericJournalCrawler(fetcher)
        extractor = ArticleExtractor(fetcher)

        for url_index, record in enumerate(input_records, start=1):
            source_url = record["url"]
            employee_name = record.get("employee_name", "")
            employee_email = record.get("employee_email", "")
            requested_file_name = record.get("file_name", "")
            file_key = sanitize_filename(
                requested_file_name,
                f"journal_{url_index}",
            )

            grouped_rows.setdefault(file_key, [])
            grouped_errors.setdefault(file_key, [])

            print(
                f"\n[{url_index}/{len(input_records)}] "
                f"{source_url} -> {file_key}.xlsx"
            )

            try:
                journal_name, issues = crawler.discover_issues(
                    source_url,
                    max(1, args.max_issues),
                )
            except Exception as exc:
                print(f"[ERROR] issue discovery: {exc}")
                all_errors.append(
                    {
                        "source_url": source_url,
                        "stage": "issue_discovery",
                        "error": str(exc),
                    }
                )
                continue

            print(f"  journal={journal_name!r} issues={len(issues)}")

            for issue_index, issue in enumerate(issues, start=1):
                try:
                    articles = crawler.discover_articles(
                        issue,
                        max(1, args.max_articles),
                    )
                except Exception as exc:
                    print(f"  [ERROR] article discovery: {exc}")
                    all_errors.append(
                        {
                            "source_url": source_url,
                            "stage": "article_discovery",
                            "error": f"{issue.url}: {exc}",
                        }
                    )
                    continue

                print(
                    f"  [{issue_index}/{len(issues)}] "
                    f"volume={issue.volume or '-'} "
                    f"issue={issue.issue or '-'} "
                    f"year={issue.year or '-'} "
                    f"articles={len(articles)}"
                )

                for article_index, article in enumerate(articles, start=1):
                    try:
                        title, doi, authors, emails, pdf_url = extractor.extract(
                            article,
                            use_pdf=not args.no_pdf,
                            pdf_max_pages=max(1, args.pdf_max_pages),
                        )

                        article_rows = rows_for_article(
                            employee_name=employee_name,
                            employee_email=employee_email,
                            file_name=file_key,
                            source_url=source_url,
                            journal_name=journal_name,
                            article=article,
                            article_title=title,
                            doi=doi,
                            pdf_url=pdf_url,
                            authors=authors,
                            emails=emails,
                            ollama=ollama,
                            include_without_email=args.include_without_email,
                            include_unmapped=args.include_unmapped,
                        )

                        all_rows.extend(article_rows)
                        grouped_rows[file_key].extend(article_rows)

                        print(
                            f"      [{article_index}/{len(articles)}] "
                            f"authors={len(authors)} emails={len(emails)} "
                            f"rows={len(article_rows)} "
                            f"{title[:90]}"
                        )

                    except Exception as exc:
                        print(f"      [ERROR] article parse: {exc}")
                        all_errors.append(
                            {
                                "source_url": source_url,
                                "stage": "article_parse",
                                "error": f"{article.url}: {exc}",
                            }
                        )

                    time.sleep(0.25)

    finally:
        if browser_ctx is not None:
            browser_ctx.__exit__(None, None, None)

    def dedupe_rows(items: list[OutputRow]) -> list[OutputRow]:
        result: list[OutputRow] = []
        seen: set[tuple[str, str]] = set()

        for row in items:
            if row.email:
                key = (
                    row.article_url.casefold(),
                    row.email.casefold(),
                )
                if key in seen:
                    continue
                seen.add(key)
            result.append(row)

        return result

    total_written = 0

    for file_key, file_rows in grouped_rows.items():
        clean_rows = dedupe_rows(file_rows)

        source_urls = {
            row.source_url
            for row in file_rows
        }
        file_errors = [
            err
            for err in all_errors
            if clean_text(err.get("source_url")) in source_urls
        ]

        output_path = output_dir / f"{file_key}.xlsx"

        write_output(
            output_path,
            clean_rows,
            file_errors,
        )

        total_written += len(clean_rows)

        print(
            f"[SAVED] {output_path.resolve()} "
            f"rows={len(clean_rows)}"
        )

    if args.combined_output:
        combined_rows = dedupe_rows(all_rows)
        write_output(
            Path(args.combined_output),
            combined_rows,
            all_errors,
        )

    print("\n========================================")
    print(f"URLs processed : {len(input_records)}")
    print(f"Output rows    : {total_written}")
    print(f"Errors         : {len(all_errors)}")
    print(f"Output folder  : {output_dir.resolve()}")
    if args.combined_output:
        print(
            f"Combined Excel : "
            f"{Path(args.combined_output).resolve()}"
        )
    print("========================================")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())